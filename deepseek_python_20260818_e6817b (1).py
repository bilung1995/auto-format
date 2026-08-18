import asyncio
import logging
import html
import os
import sqlite3
import re
from datetime import datetime
from pathlib import Path
from calendar import monthrange
from functools import lru_cache

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, Message, ReplyKeyboardRemove

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_IDS = {int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()}

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN belum diisi di Railway Variables")

DB_PATH = Path("bot.db")

DEFAULT_TEMPLATE = """📍 KAB : {KAB}
📍 KEC : {KEC}
📍 KEL : {KEL}

💰 SALDO : {SALDO}

🚻 KELAMIN : {KELAMIN}
💳 KPJ : {KPJ}
🎯 SENSOR : {SENSOR}
📆 IT : {IT}

🏛️ PT : {PT}

DPT JMO LASIK ✅"""


# ==================== OPTIMASI DATABASE ====================

class Database:
    """Singleton untuk manajemen koneksi database dengan connection pooling"""
    _instance = None
    _conn = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance
    
    def get_conn(self):
        if self._conn is None:
            self._conn = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
            self._conn.row_factory = sqlite3.Row
            self._conn.execute("PRAGMA journal_mode=WAL")
            self._conn.execute("PRAGMA synchronous=NORMAL")
            self._conn.execute("PRAGMA busy_timeout=10000")
            self._conn.execute("PRAGMA cache_size=-4000")
            self._conn.execute("PRAGMA temp_store=MEMORY")
        return self._conn
    
    def close(self):
        if self._conn:
            self._conn.close()
            self._conn = None


db_instance = Database()

def db():
    return db_instance.get_conn()


def init_database():
    """Inisialisasi database dengan indeks yang optimal"""
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        telegram_id INTEGER PRIMARY KEY,
        name TEXT,
        username TEXT,
        balance INTEGER DEFAULT 0,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS subscriptions (
        telegram_id INTEGER PRIMARY KEY,
        package_code TEXT,
        package_name TEXT,
        price INTEGER DEFAULT 0,
        start_date TEXT,
        expiry_date TEXT,
        status TEXT DEFAULT 'inactive'
    );
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        amount INTEGER NOT NULL,
        payment_method TEXT DEFAULT 'SEABANK/DANA',
        package_code TEXT,
        package_name TEXT,
        proof_file_id TEXT,
        status TEXT DEFAULT 'pending',
        created_at TEXT NOT NULL,
        processed_at TEXT
    );
    CREATE TABLE IF NOT EXISTS format_settings (
        telegram_id INTEGER PRIMARY KEY,
        template TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        input_text TEXT,
        result_text TEXT,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_codes (
        telegram_id INTEGER PRIMARY KEY,
        prefix TEXT NOT NULL,
        suffix TEXT DEFAULT '',
        current_number INTEGER NOT NULL,
        padding INTEGER NOT NULL,
        enabled INTEGER DEFAULT 1,
        updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        email TEXT,
        password TEXT,
        raw_text TEXT,
        result_id INTEGER,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        input_text TEXT,
        result_text TEXT,
        created_at TEXT NOT NULL,
        deleted_at TEXT NOT NULL
    );
    """)
    conn.executescript("""
    CREATE INDEX IF NOT EXISTS idx_transactions_status ON transactions(status);
    CREATE INDEX IF NOT EXISTS idx_transactions_user ON transactions(telegram_id);
    CREATE INDEX IF NOT EXISTS idx_format_results_user ON format_results(telegram_id);
    CREATE INDEX IF NOT EXISTS idx_format_results_created ON format_results(created_at DESC);
    CREATE INDEX IF NOT EXISTS idx_subscriptions_user ON subscriptions(telegram_id);
    """)
    conn.commit()


# ==================== OPTIMASI FUNGSI ====================

@lru_cache(maxsize=128)
def rupiah(value):
    return f"Rp {int(value):,}".replace(",", ".")

@lru_cache(maxsize=64)
def is_admin(user_id):
    return user_id in ADMIN_IDS


# ==================== OPTIMASI JMO SOLUTIONS (LOAD LAZY) ====================

# JMO_SOLUTIONS tetap dipertahankan tapi dengan struktur yang lebih ringan
# Solusi dipisahkan ke file terpisah untuk mengurangi memory

class JMOSolutionManager:
    """Manajer solusi JMO dengan lazy loading"""
    _solutions = None
    
    @classmethod
    def get_solutions(cls):
        if cls._solutions is None:
            cls._solutions = cls._load_solutions()
        return cls._solutions
    
    @staticmethod
    def _load_solutions():
        """Memuat solusi dengan struktur yang lebih ringan"""
        # Gunakan list untuk mengurangi overhead dictionary
        solutions = []
        
        # ERROR KODE
        error_solutions = [
            ("025", ["025", "error 025", "kode 025", "pesan 025", "025 jmo"],
             """🔎 <b>MASALAH KODE 025 - JMO</b>

❌ <b>Penyebab:</b>
• Data identitas tidak sesuai dengan database BPJS Ketenagakerjaan
• NIK, nama, atau tanggal lahir tidak match
• Kepesertaan belum terdaftar atau tidak aktif
• Sistem JMO sedang maintenance

✅ <b>Solusi:</b>
1️⃣ Periksa kembali NIK, Nama, dan Tanggal Lahir
2️⃣ Pastikan kepesertaan BPJS Ketenagakerjaan aktif
3️⃣ Update aplikasi JMO ke versi terbaru
4️⃣ Coba login ulang setelah 1x24 jam
5️⃣ Hubungi kantor cabang BPJS terdekat"""),
            
            ("026", ["026", "error 026", "kode 026"],
             """🔎 <b>MASALAH KODE 026 - JMO</b>

❌ <b>Penyebab:</b>
• Nomor KPJ tidak ditemukan atau tidak valid
• Kepesertaan sudah berakhir (resign/pensiun)
• Data kepesertaan belum sinkron

✅ <b>Solusi:</b>
1️⃣ Cek nomor KPJ di kartu kepesertaan fisik
2️⃣ Pastikan status kepesertaan masih aktif
3️⃣ Jika baru daftar, tunggu 1-2 hari kerja
4️⃣ Hubungi HRD/Perusahaan untuk cek kepesertaan"""),
            
            ("027", ["027", "error 027", "kode 027"],
             """🔎 <b>MASALAH KODE 027 - JMO</b>

❌ <b>Penyebab:</b>
• Email tidak terdaftar atau tidak aktif
• Email sudah digunakan oleh akun lain

✅ <b>Solusi:</b>
1️⃣ Gunakan email aktif (Gmail, Yahoo, dll)
2️⃣ Cek folder SPAM untuk email verifikasi
3️⃣ Gunakan email yang belum terdaftar"""),
            
            ("028", ["028", "error 028", "kode 028"],
             """🔎 <b>MASALAH KODE 028 - JMO</b>

❌ <b>Penyebab:</b>
• Nomor HP tidak aktif atau tidak terdaftar
• Format nomor HP salah

✅ <b>Solusi:</b>
1️⃣ Pastikan nomor HP aktif dan dapat SMS
2️⃣ Gunakan format 08xxxxxxxx (tanpa +62)
3️⃣ Cek sinyal dan jaringan HP"""),
            
            ("029", ["029", "error 029", "kode 029"],
             """🔎 <b>MASALAH KODE 029 - JMO</b>

❌ <b>Penyebab:</b>
• Verifikasi wajah (face recognition) gagal
• Pencahayaan kurang atau terlalu terang

✅ <b>Solusi:</b>
1️⃣ Cari tempat dengan pencahayaan cukup
2️⃣ Hapus kacamata/aksesoris yang menutupi wajah
3️⃣ Posisikan wajah di tengah frame
4️⃣ Jangan gunakan foto atau video"""),
        ]
        
        # MASALAH UMUM
        general_solutions = [
            ("login", ["login", "tidak bisa login", "gagal login", "masuk", "sign in"],
             """🔎 <b>MASALAH LOGIN JMO</b>

❌ <b>Penyebab:</b>
• Email/HP dan password tidak cocok
• Akun belum terdaftar
• Server JMO bermasalah

✅ <b>Solusi:</b>
1️⃣ Periksa kembali email/HP dan password
2️⃣ Gunakan fitur "Lupa Kata Sandi"
3️⃣ Cek koneksi internet
4️⃣ Update aplikasi ke versi terbaru"""),
            
            ("password", ["password", "kata sandi", "lupa sandi", "lupa password", "reset password"],
             """🔎 <b>LUPA PASSWORD JMO</b>

❌ <b>Penyebab:</b>
• Lupa password yang digunakan
• Password kadaluarsa

✅ <b>Solusi:</b>
1️⃣ Klik "Lupa Kata Sandi" di halaman login
2️⃣ Masukkan email atau HP terdaftar
3️⃣ Ikuti instruksi reset via email/SMS
4️⃣ Buat password baru yang kuat"""),
            
            ("wajah", ["verifikasi wajah", "face", "wajah", "face recognition", "selfie"],
             """🔎 <b>MASALAH VERIFIKASI WAJAH JMO</b>

❌ <b>Penyebab:</b>
• Pencahayaan kurang atau berlebihan
• Wajah tidak terlihat jelas

✅ <b>Solusi:</b>
1️⃣ Cari ruangan dengan pencahayaan cukup
2️⃣ Hapus kacamata, topi, masker
3️⃣ Posisi wajah di tengah frame
4️⃣ Jangan gunakan filter atau efek
5️⃣ Ikuti instruksi (blink, senyum, angkat alis)"""),
            
            ("kpj", ["kpj", "nomor kpj", "kpj tidak ditemukan", "kpj tidak valid"],
             """🔎 <b>MASALAH NOMOR KPJ JMO</b>

❌ <b>Penyebab:</b>
• Nomor KPJ yang dimasukkan salah
• Kepesertaan tidak aktif

✅ <b>Solusi:</b>
1️⃣ Periksa kembali nomor KPJ di kartu
2️⃣ Cek status kepesertaan di aplikasi
3️⃣ Hubungi HRD untuk verifikasi"""),
            
            ("jht", ["saldo", "jht tidak muncul", "saldo jht", "cek saldo", "jht"],
             """🔎 <b>SALDO JHT TIDAK MUNCUL</b>

❌ <b>Penyebab:</b>
• Data tidak sinkron dengan server
• Kepesertaan tidak aktif

✅ <b>Solusi:</b>
1️⃣ Tarik layar ke bawah untuk refresh
2️⃣ Login ulang ke aplikasi
3️⃣ Update aplikasi ke versi terbaru"""),
            
            ("otp", ["otp", "kode otp", "verifikasi otp", "sms otp"],
             """🔎 <b>MASALAH OTP JMO</b>

❌ <b>Penyebab:</b>
• Nomor HP tidak aktif
• Sinyal buruk atau tidak ada

✅ <b>Solusi:</b>
1️⃣ Pastikan sinyal HP bagus
2️⃣ Cek folder spam/blocked SMS
3️⃣ Tunggu 1-2 menit, jangan spam request
4️⃣ Restart HP dan coba lagi"""),
            
            ("email", ["email", "ubah email", "ganti email", "verifikasi email"],
             """🔎 <b>MASALAH EMAIL JMO</b>

❌ <b>Penyebab:</b>
• Email tidak aktif
• Folder spam penuh

✅ <b>Solusi:</b>
1️⃣ Cek folder SPAM/Trash
2️⃣ Pastikan email aktif dan bisa diakses
3️⃣ Gunakan email yang selalu dipakai"""),
            
            ("aktivasi", ["aktivasi", "belum terdaftar", "registrasi", "daftar", "register"],
             """🔎 <b>MASALAH AKTIVASI/REGISTRASI JMO</b>

❌ <b>Penyebab:</b>
• Data kepesertaan belum terdaftar
• Status kepesertaan non-aktif

✅ <b>Solusi:</b>
1️⃣ Pastikan data kepesertaan sudah terdaftar
2️⃣ Cek status kepesertaan di BPJS
3️⃣ Tunggu 1-2 hari kerja setelah pendaftaran"""),
            
            ("klaim", ["klaim", "pengajuan klaim", "klaim jht", "cairkan jht"],
             """🔎 <b>PENCAIRAN JHT 2025</b>

📌 <b>Syarat Pencairan JHT:</b>
1️⃣ Peserta sudah berhenti bekerja
2️⃣ Masa kepesertaan minimal 5 tahun
3️⃣ Status kepesertaan non-aktif

📝 <b>Dokumen:</b>
• KTP asli dan fotokopi
• KK asli dan fotokopi
• Kartu kepesertaan BPJS
• Surat PHK/resign dari perusahaan
• Buku rekening bank aktif"""),
            
            ("server", ["server", "gangguan", "maintenance", "tidak dapat terhubung", "connection"],
             """🔎 <b>MASALAH KONEKSI/SERVER JMO</b>

❌ <b>Penyebab:</b>
• Server BPJS sedang maintenance
• Gangguan jaringan internet

✅ <b>Solusi:</b>
1️⃣ Cek koneksi internet
2️⃣ Tunggu 15-30 menit
3️⃣ Coba di jam non-sibuk
4️⃣ Update aplikasi ke versi terbaru"""),
            
            ("cara daftar", ["cara daftar", "pendaftaran", "registrasi", "daftar jmo"],
             """🔎 <b>CARA DAFTAR JMO</b>

📝 <b>Langkah-langkah:</b>
1️⃣ Download aplikasi JMO di Play Store/App Store
2️⃣ Buka aplikasi dan pilih "Daftar"
3️⃣ Input data: NIK, Nama, Tanggal lahir, HP, Email
4️⃣ Verifikasi OTP via SMS
5️⃣ Verifikasi email
6️⃣ Buat password
7️⃣ Login dan lengkapi data"""),
        ]
        
        # Gabungkan semua solusi
        all_solutions = error_solutions + general_solutions
        
        # Return sebagai list untuk mengurangi overhead
        return all_solutions
    
    @classmethod
    def get_solution(cls, text: str) -> str:
        """Mencari solusi berdasarkan keyword dengan optimasi"""
        text_lower = text.lower()
        solutions = cls.get_solutions()
        
        for _, keywords, solution in solutions:
            for keyword in keywords:
                if keyword in text_lower:
                    return solution
        
        return """🛠️ <b>BELUM DITEMUKAN SOLUSI KHUSUS</b>

Maaf, saya belum menemukan solusi yang tepat untuk masalah Anda.

📌 <b>Untuk mendapatkan solusi yang lebih akurat, silakan:</b>
1️⃣ Tulis ulang masalah dengan lebih DETAIL
2️⃣ Sertakan KODE ERROR yang muncul (025, 026, dst)
3️⃣ Jelaskan TAHAPAN yang gagal
4️⃣ Sebutkan PESAN ERROR lengkapnya

📞 <b>Atau hubungi langsung:</b>
• Admin: @Hambali1995
• WhatsApp: 083160776091"""


# ==================== OPTIMASI FUNGSI FORMAT ====================

def parse_kode_input(text):
    """Parsing kode dengan regex yang dioptimasi"""
    original = text.strip()
    if not original:
        return None
    
    # Cari angka di akhir
    m = re.search(r'(\d+)\s*[^\d]*$', original)
    if not m:
        return original, 1, 3, ""
    
    num_str = m.group(1)
    try:
        num = int(num_str)
    except:
        num = 1
    
    padding = max(len(num_str), 3)
    prefix = original[:m.start(1)].strip()
    suffix = original[m.end(1):].strip()
    
    if not prefix and not suffix:
        prefix = "JPG -"
    
    return prefix, num, padding, suffix


@lru_cache(maxsize=64)
def get_subscription(user_id):
    conn = db()
    try:
        row = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (user_id,)).fetchone()
        return row
    finally:
        conn.commit()


def has_auto_format_access(user_id):
    sub = get_subscription(user_id)
    if not sub:
        return False
    if sub["status"] == "unlimited":
        return True
    if sub["status"] == "active" and sub["expiry_date"]:
        try:
            return datetime.now() < datetime.fromisoformat(sub["expiry_date"])
        except ValueError:
            return False
    return False


def add_months(dt, months):
    idx = dt.month - 1 + months
    year = dt.year + idx // 12
    month = idx % 12 + 1
    day = min(dt.day, monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)


def generate_next_code(user_id):
    conn = db()
    try:
        row = conn.execute("SELECT * FROM format_codes WHERE telegram_id=?", (user_id,)).fetchone()
        if not row or not row["enabled"]:
            return None
        
        prefix = row["prefix"]
        suffix = row.get("suffix", "") or ""
        num = row["current_number"]
        padding = row["padding"]
        
        if suffix:
            code_str = f"{prefix} {str(num).zfill(padding)} {suffix}".strip()
        else:
            code_str = f"{prefix} {str(num).zfill(padding)}".strip()
            code_str = re.sub(r'\s+', ' ', code_str)
        
        conn.execute("UPDATE format_codes SET current_number=current_number+1 WHERE telegram_id=?", (user_id,))
        conn.commit()
        return code_str
    finally:
        conn.close()


# ==================== OPTIMASI MENU ====================

def main_menu(user_id=None):
    if user_id is None:
        user_id = 0
    
    rows = [
        [InlineKeyboardButton(text="👤 PROFIL USER", callback_data="profile")],
        [InlineKeyboardButton(text="💳 TOP UP", callback_data="topup")],
        [InlineKeyboardButton(text="📊 CEK STATUS", callback_data="status")],
        [InlineKeyboardButton(text="💡 SOLUSI JMO", callback_data="solusi_jmo")],
        [InlineKeyboardButton(text="📝 AUTO FORMAT", callback_data="auto_format")],
        [InlineKeyboardButton(text="📞 HUBUNGI ADMIN", callback_data="contact_admin")],
        [InlineKeyboardButton(text="ℹ️ INFO BOT", callback_data="info_bot")],
    ]
    if is_admin(user_id):
        rows.append([InlineKeyboardButton(text="🔐 PANEL ADMIN", callback_data="admin_panel")])
        rows.append([InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def auto_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 BUAT FORMAT", callback_data="format_create")],
        [InlineKeyboardButton(text="📄 HASIL TERKINI", callback_data="format_results")],
        [InlineKeyboardButton(text="📧 HASIL FORMAT+AKUN", callback_data="hasil_akun")],
        [InlineKeyboardButton(text="⚙️ SETTING FORMAT", callback_data="format_setting")],
        [InlineKeyboardButton(text="🕘 RIWAYAT", callback_data="format_history")],
        [InlineKeyboardButton(text="💳 TOP UP", callback_data="topup")],
        [InlineKeyboardButton(text="📊 STATUS", callback_data="status")],
        [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
    ])


def admin_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👥 CEK USER AKTIF", callback_data="admin_active")],
        [InlineKeyboardButton(text="💰 LIHAT TRANSAKSI PENDING", callback_data="admin_pending")],
        [InlineKeyboardButton(text="➕ TAMBAH SALDO USER", callback_data="admin_add_balance")],
        [InlineKeyboardButton(text="➖ KURANGI SALDO USER", callback_data="admin_sub_balance")],
        [InlineKeyboardButton(text="🗑️ HAPUS DATA USER", callback_data="admin_delete_user")],
        [InlineKeyboardButton(text="📢 KIRIM BROADCAST KE USER", callback_data="admin_broadcast")],
        [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
    ])


def back_main():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
    ])


# ==================== STATE CLASSES ====================

class PaymentState(StatesGroup):
    waiting_topup_amount = State()
    waiting_proof = State()


class JmoState(StatesGroup):
    waiting_question = State()


class FormatState(StatesGroup):
    waiting_kode = State()
    waiting_manual = State()
    waiting_excel = State()
    waiting_setting = State()
    waiting_search = State()
    waiting_search_akun = State()
    waiting_edit = State()
    waiting_edit_akun = State()


class AdminState(StatesGroup):
    waiting_user_amount = State()
    waiting_delete_user = State()
    waiting_broadcast = State()


# ==================== REGISTER USER ====================

def register_user(user):
    conn = db()
    try:
        conn.execute("""
            INSERT INTO users(telegram_id,name,username,balance,created_at)
            VALUES(?,?,?,0,?)
            ON CONFLICT(telegram_id) DO UPDATE SET
                name=excluded.name, username=excluded.username
        """, (user.id, user.full_name, user.username, datetime.now().isoformat()))
        conn.commit()
    finally:
        conn.close()


# ==================== DP ====================

dp = Dispatcher()


# ==================== COMMAND START ====================

@dp.message(CommandStart())
async def start(message: Message, state: FSMContext):
    await state.clear()
    register_user(message.from_user)
    try:
        await message.answer("Menghapus menu lama...", reply_markup=ReplyKeyboardRemove())
    except Exception:
        pass
    await message.answer(
        "🤖 <b>SAHABAT JHT 🤖</b>\n\n"
        f"👋 Selamat datang, <b>{message.from_user.full_name}</b>!\n"
        "Gimana kabarnya nih, saya berharap kabar baik-baik saja yah, "
        "tetap semangat dan jangan lupa bersyukur.\n"
        "Silahkan pilih menu di bawah ini : 👇",
        reply_markup=main_menu(message.from_user.id)
    )


# ==================== COMMAND PROFIL ====================

@dp.message(Command("profil"))
async def cmd_profil(message: Message, state: FSMContext):
    await state.clear()
    register_user(message.from_user)
    conn = db()
    try:
        user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (message.from_user.id,)).fetchone()
        sub = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (message.from_user.id,)).fetchone()
    finally:
        conn.close()
    
    username = f"@{user['username']}" if user["username"] else "-"
    package = sub["package_name"] if sub else "Belum ada"
    expiry = "-"
    status = "🔴 Tidak aktif"
    if sub:
        if sub["status"] == "unlimited":
            expiry, status = "SELAMANYA", "🟢 AKTIF"
        elif sub["expiry_date"]:
            exp = datetime.fromisoformat(sub["expiry_date"])
            expiry = exp.strftime("%d-%m-%Y")
            status = "🟢 AKTIF" if datetime.now() < exp else "🔴 EXPIRED"
    
    await message.answer(
        "👤 <b>PROFIL USER</b>\n\n"
        f"🆔 Telegram ID : <code>{user['telegram_id']}</code>\n"
        f"👤 Nama : {user['name']}\n"
        f"📱 Username : {username}\n"
        f"💰 Saldo : <b>{rupiah(user['balance'])}</b>\n\n"
        f"📦 Langganan : {package}\n"
        f"📅 Berakhir : {expiry}\n"
        f"📊 Status : {status}",
        reply_markup=main_menu(message.from_user.id)
    )


# ==================== COMMAND STATUS ====================

@dp.message(Command("status"))
async def cmd_status(message: Message, state: FSMContext):
    await state.clear()
    sub = get_subscription(message.from_user.id)
    if not sub:
        txt = "📊 <b>CEK STATUS LAYANAN</b>\n\n🔴 Belum ada langganan aktif.\nSilakan beli paket di menu AUTO FORMAT."
    else:
        expiry = sub["expiry_date"] if sub["expiry_date"] else "SELAMANYA"
        txt = f"📊 <b>CEK STATUS LAYANAN</b>\n\n📦 Paket : {sub['package_name']}\n📅 Berakhir : {expiry}\n📊 Status : {sub['status'].upper()}"
    await message.answer(txt, reply_markup=main_menu(message.from_user.id))


# ==================== COMMAND AUTO FORMAT ====================

@dp.message(Command("autoformat"))
@dp.message(Command("auto_format"))
async def cmd_autoformat(message: Message, state: FSMContext):
    await state.clear()
    if has_auto_format_access(message.from_user.id):
        await message.answer(
            "📝 <b>AUTO FORMAT</b>\n\n🔓 Akses kamu aktif.\nSilakan pilih menu:",
            reply_markup=auto_menu()
        )
    else:
        await message.answer(
            "🔒 <b>AUTO FORMAT TERKUNCI</b>\n\nUntuk membuka AUTO FORMAT, silakan pilih paket:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🟢 PAKET 6 BULAN — Rp50.000", callback_data="af_6m")],
                [InlineKeyboardButton(text="🔵 PAKET 1 TAHUN — Rp80.000", callback_data="af_1y")],
                [InlineKeyboardButton(text="🟣 PAKET UNLIMITED — Rp200.000", callback_data="af_unlimited")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )


# ==================== COMMAND SOLUSI JMO ====================

@dp.message(Command("solusijmo"))
@dp.message(Command("jmo"))
async def cmd_jmo(message: Message):
    await message.answer(
        "🛠️ <b>SOLUSI MASALAH JMO</b>\n\nKetik masalah kamu, contoh:\n<code>Kode 025 saat login</code>\n<code>JMO error 026</code>\n<code>Verifikasi wajah gagal</code>",
        reply_markup=back_main()
    )


# ==================== COMMAND BANTUAN ====================

@dp.message(Command("bantuan"))
@dp.message(Command("admin"))
async def cmd_bantuan(message: Message):
    await message.answer(
        "📞 <b>HUBUNGI ADMIN</b>\n\n👤 Admin : @Hambali1995\n📱 WhatsApp : 083160776091",
        reply_markup=back_main()
    )


# ==================== COMMAND INFO ====================

@dp.message(Command("info"))
async def cmd_info(message: Message):
    await message.answer(
        "ℹ️ <b>INFO BOT SABABAT JHT</b>\n\n🤖 Bot bantuan JHT, Auto Format, Top Up, dan Solusi JMO.",
        reply_markup=back_main()
    )


# ==================== CALLBACK: BACK MAIN ====================

@dp.callback_query(F.data == "back_main")
async def back_main_handler(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "🤖 <b>SABABAT JHT 🤖</b>\n\n"
        f"👋 Selamat datang, <b>{callback.from_user.full_name}</b>!\n"
        "Gimana kabarnya nih, saya berharap kabar baik-baik saja yah, "
        "tetap semangat dan jangan lupa bersyukur.\n"
        "Silahkan pilih menu di bawah ini : 👇",
        reply_markup=main_menu(callback.from_user.id)
    )
    await callback.answer()


# ==================== CALLBACK: INFO BOT ====================

@dp.callback_query(F.data == "info_bot")
async def info_bot(callback: CallbackQuery):
    await callback.message.edit_text(
        "🤖 <b>SABABAT JHT</b>\n\n"
        "Bot bantuan JHT, Auto Format, Top Up, dan Solusi JMO.\n"
        "Gunakan menu utama untuk melanjutkan.",
        reply_markup=back_main()
    )
    await callback.answer()


# ==================== CALLBACK: PROFILE ====================

@dp.callback_query(F.data == "profile")
async def profile(callback: CallbackQuery):
    register_user(callback.from_user)
    conn = db()
    try:
        user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
        sub = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
    finally:
        conn.close()

    username = f"@{user['username']}" if user["username"] else "-"
    package = sub["package_name"] if sub else "Belum ada"
    expiry = "-"
    status = "🔴 Tidak aktif"

    if sub:
        if sub["status"] == "unlimited":
            expiry, status = "SELAMANYA", "🟢 AKTIF"
        elif sub["expiry_date"]:
            exp = datetime.fromisoformat(sub["expiry_date"])
            expiry = exp.strftime("%d-%m-%Y")
            status = "🟢 AKTIF" if datetime.now() < exp else "🔴 EXPIRED"

    await callback.message.edit_text(
        "👤 <b>PROFIL USER</b>\n\n"
        f"🆔 Telegram ID : <code>{user['telegram_id']}</code>\n"
        f"👤 Nama : {user['name']}\n"
        f"📱 Username : {username}\n"
        f"💰 Saldo : <b>{rupiah(user['balance'])}</b>\n\n"
        f"📦 Langganan : {package}\n"
        f"📅 Berakhir : {expiry}\n"
        f"📊 Status : {status}",
        reply_markup=back_main()
    )
    await callback.answer()


# ==================== CALLBACK: TOPUP ====================

@dp.callback_query(F.data == "topup")
async def topup(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(PaymentState.waiting_topup_amount)
    await callback.message.edit_text(
        "💳 <b>TOP UP SALDO</b>\n\n"
        "Silakan transfer ke:\n\n"
        "🏦 <b>SEABANK</b>\n"
        "901040978290\n"
        "A/N HAMBALI\n\n"
        "💰 <b>DANA</b>\n"
        "083824101264\n"
        "A/N HAMBALI\n\n"
        "Atau bisa ketik nominal.\n"
        "Contoh : <code>50000</code>\n\n"
        "Setelah membayar, ketik nominal yang dibayar lalu tekan <b>SUDAH BAYAR</b>.",
        reply_markup=back_main()
    )
    await callback.answer()


# ==================== PAYMENT: AMOUNT ====================

@dp.message(PaymentState.waiting_topup_amount, F.text)
async def topup_amount(message: Message, state: FSMContext):
    raw = message.text.strip().lower().replace("rp", "").replace(".", "").replace(",", "").replace(" ", "")
    if not raw.isdigit() or int(raw) <= 0:
        await message.answer("❌ Nominal tidak valid. Contoh: <code>50000</code>")
        return

    amount = int(raw)
    await state.update_data(amount=amount, package_code=None, package_name=None)
    await state.set_state(PaymentState.waiting_proof)
    await message.answer(
        "✅ <b>NOMINAL TERCATAT</b>\n\n"
        f"💰 Nominal : <b>{rupiah(amount)}</b>\n\n"
        "Setelah transfer, tekan tombol:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ SUDAH BAYAR", callback_data="payment_done")],
            [InlineKeyboardButton(text="❌ BATAL", callback_data="back_main")]
        ])
    )


# ==================== AUTO PACKAGES ====================

AUTO_PACKAGES = {
    "af_6m": ("6M", "AUTO FORMAT 6 BULAN", 50000),
    "af_1y": ("1Y", "AUTO FORMAT 1 TAHUN", 80000),
    "af_unlimited": ("UNLIMITED", "AUTO FORMAT UNLIMITED", 200000)
}


@dp.callback_query(F.data == "auto_format")
async def auto_format(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    if has_auto_format_access(callback.from_user.id):
        await callback.message.edit_text(
            "📝 <b>AUTO FORMAT</b>\n\n🔓 Akses kamu aktif.\n\nSilakan pilih menu:",
            reply_markup=auto_menu()
        )
    else:
        await callback.message.edit_text(
            "🔒 <b>AUTO FORMAT TERKUNCI</b>\n\n"
            "Untuk membuka AUTO FORMAT, silakan pilih paket:\n\n"
            "🟢 6 Bulan — <b>Rp50.000</b>\n"
            "🔵 1 Tahun — <b>Rp80.000</b>\n"
            "🟣 Unlimited — <b>Rp200.000</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🟢 PAKET 6 BULAN — Rp50.000", callback_data="af_6m")],
                [InlineKeyboardButton(text="🔵 PAKET 1 TAHUN — Rp80.000", callback_data="af_1y")],
                [InlineKeyboardButton(text="🟣 PAKET UNLIMITED — Rp200.000", callback_data="af_unlimited")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )
    await callback.answer()


@dp.callback_query(F.data.in_(set(AUTO_PACKAGES.keys())))
async def auto_package(callback: CallbackQuery, state: FSMContext):
    code, name, price = AUTO_PACKAGES[callback.data]
    await state.update_data(amount=price, package_code=code, package_name=name)
    await state.set_state(PaymentState.waiting_proof)

    await callback.message.edit_text(
        "💳 <b>PEMBAYARAN AUTO FORMAT</b>\n\n"
        f"📦 Paket : <b>{name}</b>\n"
        f"💰 Harga : <b>{rupiah(price)}</b>\n\n"
        "Silakan transfer ke:\n\n"
        "🏦 <b>SEABANK</b>\n901040978290\nA/N HAMBALI\n\n"
        "💰 <b>DANA</b>\n083824101264\nA/N HAMBALI\n\n"
        "Setelah membayar, tekan <b>SUDAH BAYAR</b> lalu upload bukti.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ SUDAH BAYAR", callback_data="payment_done")],
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
        ])
    )
    await callback.answer()


# ==================== PAYMENT: DONE ====================

@dp.callback_query(F.data == "payment_done")
async def payment_done(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if not data.get("amount"):
        await callback.answer("❌ Nominal belum ditentukan.", show_alert=True)
        return
    await state.set_state(PaymentState.waiting_proof)
    await callback.message.edit_text(
        "📸 <b>UPLOAD BUKTI PEMBAYARAN</b>\n\n"
        f"💰 Nominal : <b>{rupiah(data['amount'])}</b>\n"
        f"📦 Paket : {data.get('package_name') or 'TOP UP SALDO'}\n\n"
        "Silakan kirim FOTO bukti transfer di chat ini."
    )
    await callback.answer()


# ==================== PAYMENT: PROOF ====================

@dp.message(PaymentState.waiting_proof, F.photo)
async def payment_proof(message: Message, state: FSMContext):
    data = await state.get_data()
    amount = int(data.get("amount", 0))
    if amount <= 0:
        await state.clear()
        await message.answer("❌ Nominal pembayaran tidak ditemukan.")
        return

    photo = message.photo[-1]
    conn = db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO transactions(
                telegram_id,amount,payment_method,package_code,package_name,
                proof_file_id,status,created_at
            ) VALUES(?,?,'SEABANK/DANA',?,?,?,'pending',?)
        """, (
            message.from_user.id, amount, data.get("package_code"),
            data.get("package_name"), photo.file_id, datetime.now().isoformat()
        ))
        tx_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()

    await message.answer(
        "✅ <b>BUKTI PEMBAYARAN DITERIMA</b>\n\n"
        f"🧾 Transaksi : #{tx_id}\n"
        f"💰 Nominal : {rupiah(amount)}\n"
        f"📦 Paket : {data.get('package_name') or 'TOP UP SALDO'}\n\n"
        "⏳ Menunggu konfirmasi Admin."
    )

    caption = (
        "💰 <b>PEMBAYARAN BARU</b>\n\n"
        f"🧾 Transaksi : #{tx_id}\n"
        f"👤 Nama : {message.from_user.full_name}\n"
        f"🆔 ID : <code>{message.from_user.id}</code>\n"
        f"📱 Username : @{message.from_user.username or '-'}\n"
        f"💰 Nominal : <b>{rupiah(amount)}</b>\n"
        f"📦 Paket : {data.get('package_name') or 'TOP UP SALDO'}\n"
        "🟡 Status : PENDING"
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ SETUJU", callback_data=f"approve_{tx_id}"),
        InlineKeyboardButton(text="❌ TOLAK", callback_data=f"reject_{tx_id}")
    ]])

    for admin_id in ADMIN_IDS:
        try:
            await message.bot.send_photo(admin_id, photo.file_id, caption=caption, reply_markup=keyboard)
        except Exception:
            logging.exception("Gagal mengirim transaksi ke admin")


@dp.message(PaymentState.waiting_proof)
async def payment_wrong_proof(message: Message):
    await message.answer("📸 Silakan kirim FOTO bukti pembayaran.")


# ==================== PROCESS PAYMENT ====================

async def process_payment(callback: CallbackQuery, tx_id: int, approve: bool):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Kamu bukan Admin.", show_alert=True)
        return

    conn = db()
    try:
        tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()

        if not tx or tx["status"] != "pending":
            await callback.answer("⚠️ Transaksi sudah diproses/tidak ditemukan.", show_alert=True)
            return

        now = datetime.now().isoformat()

        if approve:
            conn.execute(
                "UPDATE transactions SET status='approved', processed_at=? WHERE id=?",
                (now, tx_id)
            )

            expiry = None
            if tx["package_code"]:
                if tx["package_code"] == "6M":
                    expiry, sub_status = add_months(datetime.now(), 6), "active"
                elif tx["package_code"] == "1Y":
                    expiry, sub_status = add_months(datetime.now(), 12), "active"
                else:
                    sub_status = "unlimited"

                conn.execute("""
                    INSERT INTO subscriptions(
                        telegram_id,package_code,package_name,price,start_date,expiry_date,status
                    ) VALUES(?,?,?,?,?,?,?)
                    ON CONFLICT(telegram_id) DO UPDATE SET
                        package_code=excluded.package_code,
                        package_name=excluded.package_name,
                        price=excluded.price,
                        start_date=excluded.start_date,
                        expiry_date=excluded.expiry_date,
                        status=excluded.status
                """, (
                    tx["telegram_id"], tx["package_code"], tx["package_name"], tx["amount"],
                    now, expiry.isoformat() if expiry else None, sub_status
                ))
            else:
                conn.execute(
                    "UPDATE users SET balance=balance+? WHERE telegram_id=?",
                    (tx["amount"], tx["telegram_id"])
                )

            conn.commit()

            expiry_text = ""
            if tx["package_code"]:
                expiry_text = (
                    f"\n📅 Berakhir : "
                    f"{expiry.strftime('%d-%m-%Y') if expiry else 'SELAMANYA'}"
                )

            user_text = (
                "✅ <b>PEMBAYARAN DISETUJUI</b>\n\n"
                f"🧾 Transaksi : #{tx_id}\n"
                f"💰 Nominal : {rupiah(tx['amount'])}\n"
                f"📦 Paket : {tx['package_name'] or 'TOP UP SALDO'}"
                f"{expiry_text}\n\n"
                + (
                    "🔓 AUTO FORMAT SEKARANG SUDAH TERBUKA."
                    if tx["package_code"]
                    else "💰 Saldo kamu sudah ditambahkan."
                )
            )
        else:
            conn.execute(
                "UPDATE transactions SET status='rejected', processed_at=? WHERE id=?",
                (now, tx_id)
            )
            conn.commit()

            user_text = (
                "❌ <b>PEMBAYARAN DITOLAK</b>\n\n"
                f"🧾 Transaksi : #{tx_id}\n"
                f"💰 Nominal : {rupiah(tx['amount'])}\n"
                f"📦 Paket : {tx['package_name'] or 'TOP UP SALDO'}\n\n"
                "Silakan cek kembali bukti pembayaran atau hubungi Admin."
            )
    finally:
        conn.close()

    try:
        await callback.bot.send_message(tx["telegram_id"], user_text)
    except Exception:
        logging.exception("Gagal mengirim notifikasi pembayaran")

    try:
        if callback.message.photo:
            old_caption = callback.message.caption or ""
            label = "✅ DISETUJUI" if approve else "❌ DITOLAK"
            await callback.message.edit_caption(
                caption=old_caption + f"\n\n<b>{label}</b>",
                reply_markup=None
            )
        else:
            old_text = callback.message.text or ""
            label = "✅ DISETUJUI" if approve else "❌ DITOLAK"
            await callback.message.edit_text(
                old_text + f"\n\n<b>{label}</b>",
                reply_markup=None
            )
    except Exception:
        logging.exception("Gagal memperbarui pesan admin")

    await callback.answer("✅ Disetujui." if approve else "❌ Ditolak.")


@dp.callback_query(F.data.startswith("approve_"))
async def approve_payment(callback: CallbackQuery):
    try:
        tx_id = int(callback.data.split("_", 1)[1])
    except (ValueError, IndexError):
        await callback.answer("❌ ID transaksi tidak valid.", show_alert=True)
        return
    await process_payment(callback, tx_id, True)


@dp.callback_query(F.data.startswith("reject_"))
async def reject_payment(callback: CallbackQuery):
    try:
        tx_id = int(callback.data.split("_", 1)[1])
    except (ValueError, IndexError):
        await callback.answer("❌ ID transaksi tidak valid.", show_alert=True)
        return
    await process_payment(callback, tx_id, False)


# ==================== CALLBACK: STATUS ====================

@dp.callback_query(F.data == "status")
async def status(callback: CallbackQuery):
    conn = db()
    try:
        user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
        sub = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
    finally:
        conn.close()

    balance = user["balance"] if user else 0
    if not sub:
        text = f"📊 <b>STATUS USER</b>\n\n💰 Saldo : <b>{rupiah(balance)}</b>\n📦 Langganan : Belum ada\n📊 Status : 🔴 Tidak aktif"
    elif sub["status"] == "unlimited":
        text = f"📊 <b>STATUS USER</b>\n\n💰 Saldo : <b>{rupiah(balance)}</b>\n📦 Langganan : {sub['package_name']}\n📅 Berakhir : SELAMANYA\n📊 Status : 🟢 AKTIF"
    else:
        exp = datetime.fromisoformat(sub["expiry_date"])
        active = datetime.now() < exp
        text = (
            f"📊 <b>STATUS USER</b>\n\n💰 Saldo : <b>{rupiah(balance)}</b>\n"
            f"📦 Langganan : {sub['package_name']}\n"
            f"📅 Berakhir : {exp.strftime('%d-%m-%Y')}\n"
            f"📊 Status : {'🟢 AKTIF' if active else '🔴 EXPIRED'}"
        )

    await callback.message.edit_text(text, reply_markup=back_main())
    await callback.answer()


# ==================== SOLUSI JMO ====================

@dp.callback_query(F.data == "solusi_jmo")
async def solusi_start(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(JmoState.waiting_question)
    await callback.message.edit_text(
        "🛠️ <b>SOLUSI JMO</b>\n\n"
        "Silakan masukkan masalah Anda di sini.\n\n"
        "📌 <b>Contoh:</b>\n"
        "• <code>cara atasi solusi 026</code>\n"
        "• <code>JMO error 025</code>\n"
        "• <code>Verifikasi wajah gagal</code>\n"
        "• <code>Login tidak bisa</code>\n\n"
        "💡 <b>Tips:</b> Sebutkan kode error jika ada!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❓ SOLUSI ERROR 025", callback_data="jmo_025")],
            [InlineKeyboardButton(text="❓ SOLUSI ERROR 026", callback_data="jmo_026")],
            [InlineKeyboardButton(text="❓ SOLUSI ERROR 029 (WAJAH)", callback_data="jmo_029")],
            [InlineKeyboardButton(text="❓ CARA CAIRKAN JHT", callback_data="jmo_jht")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("jmo_"))
async def jmo_quick(callback: CallbackQuery, state: FSMContext):
    error_code = callback.data.split("_", 1)[1]
    # Gunakan JMOSolutionManager untuk mendapatkan solusi
    solution = JMOSolutionManager.get_solution(error_code)
    await callback.message.edit_text(
        solution,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 KEMBALI KE MENU SOLUSI", callback_data="solusi_jmo")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()


@dp.message(JmoState.waiting_question, F.text)
async def solusi_text(message: Message, state: FSMContext):
    solution = JMOSolutionManager.get_solution(message.text)
    await message.answer(
        solution,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 TANYA LAGI", callback_data="solusi_jmo")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )


@dp.message(JmoState.waiting_question, F.photo)
async def solusi_photo(message: Message, state: FSMContext):
    caption = message.caption or ""
    if caption:
        solution = JMOSolutionManager.get_solution(caption)
        await message.answer(
            solution,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔄 TANYA LAGI", callback_data="solusi_jmo")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )
    else:
        await message.answer(
            "📸 <b>FOTO DITERIMA</b>\n\n"
            "Silakan tuliskan kode/error yang terlihat pada foto, misalnya:\n"
            "• <code>025</code>\n"
            "• <code>026</code>\n"
            "• <code>verifikasi wajah</code>\n\n"
            "Saya akan mencari solusi yang tepat!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔄 KETIK MASALAH", callback_data="solusi_jmo")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )


# ==================== AUTO FORMAT: CREATE ====================

@dp.callback_query(F.data == "format_create")
async def format_create(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    await callback.message.edit_text(
        "📝 <b>BUAT FORMAT</b>\n\nPilih cara membuat format:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⌨️ KETIK DATA MANUAL", callback_data="format_manual")],
            [InlineKeyboardButton(text="📊 UPLOAD FILE EXCEL (.xlsx)", callback_data="format_excel")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )
    await callback.answer()


# ==================== FORMAT MANUAL ====================

@dp.callback_query(F.data == "format_manual")
async def format_manual_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(FormatState.waiting_manual)
    await callback.message.edit_text(
        "⌨️ <b>BUAT FORMAT</b>\n\n"
        "Ketik data <b>tanpa perlu menulis KAB/KEC/KEL</b>. Bot otomatis membaca urutannya:\n\n"
        "1️⃣ KAB\n2️⃣ KEC\n3️⃣ KEL\n4️⃣ SALDO\n5️⃣ KELAMIN\n6️⃣ KPJ\n7️⃣ SENSOR\n8️⃣ IT\n9️⃣ PT\n\n"
        "Contoh input:\n<code>DEPOK\nCILODONG\nKALIBARU\n10.000.000\nPEREMPUAN 1992\n2019\n23****\n01-07-2022\nINDONESIA MERDEKA</code>\n\n"
        "Hasilnya otomatis menjadi format JMO.",
        reply_markup=back_main()
    )
    await callback.answer()


def apply_template(template, raw_text, kode_header=None):
    """Aplikasikan template dengan data yang di-parse"""
    import re
    
    def parse_data(raw_text):
        data = {}
        lines = [l.strip() for l in raw_text.splitlines() if l.strip()]
        
        def get(idx, default="-"):
            if idx < len(lines):
                return lines[idx].strip()
            return default
        
        # Parse dengan regex untuk field yang sudah ada labelnya
        field_patterns = {
            "KAB": r'KAB\s*[:：]\s*([^\n]+)',
            "KEC": r'KEC\s*[:：]\s*([^\n]+)',
            "KEL": r'KEL\s*[:：]\s*([^\n]+)',
            "SALDO": r'(?:TOTAL\s*JHT|SALDO)\s*[:：]\s*([0-9\.\,]+)',
            "KELAMIN": r'KELAMIN\s*[:：]\s*([^\n]+)',
            "KPJ": r'KPJ\s*[:：]\s*([^\n]+)',
            "SENSOR": r'(?:KPJ\s*SENSOR|SENSOR)\s*[:：]\s*([^\n]+)',
            "IT": r'(?:IURAN\s*T|IT)\s*[:：]\s*([^\n]+)',
            "PT": r'PT\s*[:：]?\s*\*?\s*([^\n]+)',
        }
        
        # Coba dapatkan dari label terlebih dahulu
        for key, pattern in field_patterns.items():
            m = re.search(pattern, raw_text, re.I | re.M)
            if m:
                data[key] = m.group(1).strip()
        
        # Fallback ke urutan
        defaults = ["KAB", "KEC", "KEL", "SALDO", "KELAMIN", "KPJ", "SENSOR", "IT", "PT"]
        idx = 0
        for key in defaults:
            if key not in data or not data[key]:
                if idx < len(lines):
                    data[key] = lines[idx].strip()
                else:
                    data[key] = "-"
                idx += 1
        
        # Capitalize semua field
        for key in data:
            if data[key] and data[key] != "-":
                data[key] = data[key].upper()
        
        return data
    
    data = parse_data(raw_text)
    
    # Apply template
    if "{" in template and "}" in template:
        result = template
        for k, v in data.items():
            result = result.replace("{" + k + "}", v)
        result = re.sub(r'\{[A-Z_]+\}', '', result)
        if kode_header:
            centered = kode_header.center(27)
            result = f"{centered}\n━━━━━━━━━━━━━━━━━━━\n{result}"
        return result
    
    # Fallback: template tanpa placeholder
    result = template
    for k, v in data.items():
        result = result.replace(k, v)
    
    if kode_header:
        centered = kode_header.center(27)
        result = f"{centered}\n━━━━━━━━━━━━━━━━━━━\n{result}"
    
    return result


@dp.message(FormatState.waiting_manual, F.text)
async def format_manual_receive(message: Message, state: FSMContext):
    data = await state.get_data()
    edit_id = data.get("edit_result_id")

    conn = db()
    try:
        row = conn.execute(
            "SELECT template FROM format_settings WHERE telegram_id=?",
            (message.from_user.id,)
        ).fetchone()
        template = row["template"] if row else DEFAULT_TEMPLATE

        kode = generate_next_code(message.from_user.id)
        result = apply_template(template, message.text, kode_header=kode)

        if edit_id:
            conn.execute(
                "UPDATE format_results SET result_text=? WHERE id=? AND telegram_id=?",
                (result, edit_id, message.from_user.id)
            )
            result_id = edit_id
        else:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO format_results(telegram_id,input_text,result_text,created_at)
                VALUES(?,?,?,?)
            """, (message.from_user.id, message.text, result, datetime.now().isoformat()))
            result_id = cur.lastrowid
        
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()

    await message.answer(
        html.escape(result[:3900]),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 SALIN HASIL FORMAT", callback_data=f"copy_result_{result_id}")],
            [InlineKeyboardButton(text="📋 SALIN SEMUA HASIL", callback_data=f"copy_all_{result_id}")],
            [InlineKeyboardButton(text="✏️ EDIT HASIL FORMAT", callback_data=f"edit_result_{result_id}")],
            [InlineKeyboardButton(text="🗑️ HAPUS HASIL FORMAT", callback_data=f"delete_result_{result_id}")],
            [InlineKeyboardButton(text="💾 SIMPAN KE HISTORY", callback_data=f"save_result_{result_id}")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )


# ==================== FORMAT SETTING ====================

@dp.callback_query(F.data == "format_setting")
async def format_setting(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    conn = db()
    try:
        row = conn.execute(
            "SELECT template FROM format_settings WHERE telegram_id=?",
            (callback.from_user.id,)
        ).fetchone()
        code_row = conn.execute(
            "SELECT * FROM format_codes WHERE telegram_id=?",
            (callback.from_user.id,)
        ).fetchone()
    finally:
        conn.close()
    
    template = row["template"] if row else DEFAULT_TEMPLATE
    
    kode_info = "❌ Belum diatur"
    if code_row:
        kode_info = f"{code_row['prefix']} - {str(code_row['current_number']).zfill(code_row['padding'])} (Aktif)"
    
    await state.set_state(FormatState.waiting_setting)
    await callback.message.edit_text(
        "⚙️ <b>SETTING FORMAT</b>\n\n"
        f"🔢 <b>KODE SAAT INI:</b> {kode_info}\n\n"
        "Template saat ini:\n\n<pre>" + html.escape(template[:3000]) + "</pre>\n\n"
        "Kirim template baru untuk menggantinya, atau atur kode di bawah:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔢 SET KODE FORMAT", callback_data="set_kode_format")],
            [InlineKeyboardButton(text="❌ HAPUS KODE", callback_data="delete_kode_format")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()


@dp.message(FormatState.waiting_setting, F.text)
async def format_setting_receive(message: Message, state: FSMContext):
    raw = message.text.strip()
    # Coba deteksi template dengan placeholder
    if '{' in raw and '}' in raw:
        template_to_save = raw
    else:
        # Konversi template dari contoh
        template_to_save = raw
    
    conn = db()
    try:
        conn.execute("""
            INSERT INTO format_settings(telegram_id,template,updated_at)
            VALUES(?,?,?)
            ON CONFLICT(telegram_id) DO UPDATE SET
                template=excluded.template, updated_at=excluded.updated_at
        """, (message.from_user.id, template_to_save, datetime.now().isoformat()))
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()
    await message.answer(
        "✅ <b>SETTING FORMAT TERSIMPAN & AKTIF</b>\n\n"
        "Bot sekarang akan mengikuti template ini.",
        reply_markup=auto_menu()
    )


# ==================== SET KODE FORMAT ====================

@dp.callback_query(F.data == "set_kode_format")
async def set_kode_format(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 Terkunci", show_alert=True)
        return
    await state.set_state(FormatState.waiting_kode)
    await callback.message.edit_text(
        "🔢 <b>SET KODE FORMAT</b>\n\n"
        "Ketik kode awal yang kamu mau. Contoh:\n"
        "<code>JPG - 001</code>\n"
        "<code>ABC - 001</code>\n\n"
        "Bot akan otomatis urut: 001, 002, 003... tanpa duplikat dan muncul di atas format.\n\n"
        "Ketik kode sekarang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ BATAL", callback_data="format_setting")]
        ])
    )
    await callback.answer()


@dp.message(FormatState.waiting_kode, F.text)
async def receive_kode_format(message: Message, state: FSMContext):
    parsed = parse_kode_input(message.text)
    if not parsed:
        await message.answer("❌ Format kode salah. Contoh: <code>JPG - 001</code>")
        return
    if len(parsed) == 4:
        prefix, num, padding, suffix = parsed
    else:
        prefix, num, padding = parsed
        suffix = ""
    
    conn = db()
    try:
        conn.execute("""
            INSERT INTO format_codes(telegram_id,prefix,suffix,current_number,padding,enabled,updated_at)
            VALUES(?,?,?,?,?,1,?)
            ON CONFLICT(telegram_id) DO UPDATE SET
                prefix=excluded.prefix,
                suffix=excluded.suffix,
                current_number=excluded.current_number,
                padding=excluded.padding,
                enabled=1,
                updated_at=excluded.updated_at
        """, (message.from_user.id, prefix, suffix, num, padding, datetime.now().isoformat()))
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()
    demo_code_full = f"{prefix} {str(num).zfill(padding)} {suffix}".strip()
    await message.answer(
        f"✅ <b>KODE DISIMPAN</b>\n\n"
        f"Kode aktif: <b>{demo_code_full}</b>\n\n"
        f"Format selanjutnya akan otomatis berurutan tanpa duplikat.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ LIHAT SETTING", callback_data="format_setting")],
            [InlineKeyboardButton(text="📝 BUAT FORMAT", callback_data="format_create")]
        ])
    )


# ==================== DELETE KODE FORMAT ====================

@dp.callback_query(F.data == "delete_kode_format")
async def delete_kode_format(callback: CallbackQuery):
    conn = db()
    try:
        conn.execute("DELETE FROM format_codes WHERE telegram_id=?", (callback.from_user.id,))
        conn.commit()
    finally:
        conn.close()
    await callback.message.edit_text(
        "🗑️ <b>KODE DIHAPUS</b>\n\nKode format dinonaktifkan.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ KEMBALI KE SETTING", callback_data="format_setting")]
        ])
    )
    await callback.answer()


# ==================== FORMAT EXCEL ====================

@dp.callback_query(F.data == "format_excel")
async def format_excel_start(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    await state.set_state(FormatState.waiting_excel)
    await callback.message.edit_text(
        "📊 <b>FORMAT FILE EXCEL</b>\n\nSilakan upload file Excel <b>.xlsx</b>.",
        reply_markup=back_main()
    )
    await callback.answer()


@dp.message(FormatState.waiting_excel, F.document)
async def format_excel_receive(message: Message, state: FSMContext):
    if not (message.document.file_name or "").lower().endswith(".xlsx"):
        await message.answer("❌ File harus berformat .xlsx")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        await message.answer("❌ Tambahkan <code>openpyxl</code> ke requirements.txt lalu deploy ulang.")
        return

    file = await message.bot.get_file(message.document.file_id)
    local = Path("/tmp") / f"{message.from_user.id}_{message.document.file_name}"
    await message.bot.download_file(file.file_path, local)

    conn = db()
    try:
        row = conn.execute(
            "SELECT template FROM format_settings WHERE telegram_id=?",
            (message.from_user.id,)
        ).fetchone()
        template = row["template"] if row else DEFAULT_TEMPLATE
    finally:
        conn.close()

    wb = load_workbook(local, read_only=True, data_only=True)
    ws = wb.active
    results = []

    for values in ws.iter_rows(values_only=True):
        raw = " ".join(str(v) for v in values if v is not None).strip()
        if raw:
            results.append(apply_template(template, raw))

    if not results:
        await state.clear()
        await message.answer("❌ File Excel tidak memiliki data.")
        return

    result = "\n\n".join(results)
    conn = db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO format_results(telegram_id,input_text,result_text,created_at)
            VALUES(?,?,?,?)
        """, (message.from_user.id, message.document.file_name, result, datetime.now().isoformat()))
        rid = cur.lastrowid
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()

    await message.answer(
        "✅ <b>HASIL FORMAT EXCEL</b>\n\n" + html.escape(result[:3900]),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 SALIN SEMUA HASIL", callback_data=f"copy_all_{rid}")],
            [InlineKeyboardButton(text="💾 SIMPAN KE HISTORY", callback_data=f"save_result_{rid}")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )


@dp.message(FormatState.waiting_excel)
async def format_excel_wrong(message: Message):
    await message.answer("📊 Silakan upload file Excel .xlsx")


# ==================== FORMAT RESULTS ====================

@dp.callback_query(F.data == "format_results")
async def format_results(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    conn = db()
    try:
        rows = conn.execute(
            "SELECT id, result_text, created_at FROM format_results WHERE telegram_id=? ORDER BY id DESC LIMIT 5",
            (callback.from_user.id,)
        ).fetchall()
    finally:
        conn.close()
    
    if not rows:
        await callback.message.edit_text(
            "📄 <b>HASIL FORMAT - KOSONG</b>\n\nBelum ada hasil.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📝 BUAT FORMAT BARU", callback_data="format_create")],
                [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
            ])
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        f"📄 <b>HASIL FORMAT TERKINI - {len(rows)} TERBARU</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU", callback_data="auto_format")]
        ])
    )
    await callback.answer()

    for i, row in enumerate(rows, 1):
        rid = row["id"]
        full_text = row["result_text"].strip()
        if len(full_text) > 3800:
            full_text = full_text[:3800] + "..."
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="📋 SALIN", callback_data=f"copy_result_{rid}"),
                InlineKeyboardButton(text="✏️ EDIT", callback_data=f"edit_result_{rid}"),
                InlineKeyboardButton(text="🗑️ HAPUS", callback_data=f"delete_result_{rid}")
            ]
        ])
        try:
            await callback.message.bot.send_message(
                callback.from_user.id,
                f"<b>{i}. ID {rid}</b>\n<pre>{html.escape(full_text)}</pre>",
                reply_markup=kb
            )
            await asyncio.sleep(0.2)
        except Exception:
            pass

    await callback.message.bot.send_message(
        callback.from_user.id,
        "🔍 <b>MENU LANJUTAN</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔎 CARI HASIL FORMAT", callback_data="format_search")],
            [InlineKeyboardButton(text="📧 HASIL FORMAT+AKUN", callback_data="hasil_akun")],
            [InlineKeyboardButton(text="🗑️ HAPUS SEMUA HASIL", callback_data="clear_results")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )


@dp.callback_query(F.data == "clear_results")
async def clear_results(callback: CallbackQuery):
    conn = db()
    try:
        conn.execute("DELETE FROM format_results WHERE telegram_id=?", (callback.from_user.id,))
        conn.commit()
    finally:
        conn.close()
    await callback.message.edit_text("🗑️ Semua hasil format dihapus.", reply_markup=auto_menu())
    await callback.answer()


# ==================== FORMAT SEARCH ====================

@dp.callback_query(F.data == "format_search")
async def format_search(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    await state.set_state(FormatState.waiting_search)
    await callback.message.edit_text(
        "🔎 <b>CARI HASIL FORMAT</b>\n\n"
        "Ketik kata yang ingin dicari, misalnya:\n"
        "<code>SERANG</code> atau <code>CIPOCOK</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU HASIL FORMAT", callback_data="format_results")]
        ])
    )
    await callback.answer()


@dp.message(FormatState.waiting_search, F.text)
async def format_search_receive(message: Message, state: FSMContext):
    query = message.text.strip()
    if not query:
        await message.answer("❌ Kata pencarian tidak boleh kosong.")
        return

    conn = db()
    try:
        rows = conn.execute("""
            SELECT id, result_text, created_at
            FROM format_results
            WHERE telegram_id=? AND result_text LIKE ?
            ORDER BY id DESC LIMIT 20
        """, (message.from_user.id, f"%{query}%")).fetchall()
    finally:
        conn.close()
    
    await state.clear()

    if not rows:
        await message.answer(
            f"Tidak ada format dengan kata: <code>{query}</code>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔎 CARI LAGI", callback_data="format_search")],
                [InlineKeyboardButton(text="📄 LIHAT SEMUA", callback_data="format_results")]
            ])
        )
        return

    for row in rows:
        rid = row["id"]
        txt = row["result_text"]
        await message.bot.send_message(
            message.from_user.id,
            f"<pre>{html.escape(txt[:3800])}</pre>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="📋 SALIN", callback_data=f"copy_result_{rid}"),
                    InlineKeyboardButton(text="🗑️ HAPUS", callback_data=f"delete_result_{rid}"),
                    InlineKeyboardButton(text="✏️ EDIT", callback_data=f"edit_result_{rid}")
                ]
            ])
        )
        await asyncio.sleep(0.1)


# ==================== COPY / EDIT / DELETE RESULT ====================

@dp.callback_query(F.data.startswith("copy_result_"))
async def copy_result(callback: CallbackQuery):
    rid = int(callback.data.replace("copy_result_", ""))
    conn = db()
    try:
        row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id)).fetchone()
    finally:
        conn.close()
    
    if row:
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"📋 <b>HASIL FORMAT ID {rid}</b>\n\n<pre>{html.escape(row['result_text'][:3900])}</pre>"
        )
    await callback.answer("Disalin!")


@dp.callback_query(F.data.startswith("copy_all_"))
async def copy_all_result(callback: CallbackQuery):
    rid = int(callback.data.replace("copy_all_", ""))
    conn = db()
    try:
        row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id)).fetchone()
    finally:
        conn.close()
    
    if row:
        # Kirim semua hasil (potong jika terlalu panjang)
        text = row['result_text']
        for i in range(0, len(text), 3900):
            await callback.message.bot.send_message(
                callback.from_user.id,
                f"<pre>{html.escape(text[i:i+3900])}</pre>"
            )
    await callback.answer("Disalin!")


@dp.callback_query(F.data.startswith("edit_result_"))
async def edit_result_start(callback: CallbackQuery, state: FSMContext):
    rid = int(callback.data.replace("edit_result_", ""))
    await state.set_state(FormatState.waiting_manual)
    await state.update_data(edit_result_id=rid)
    
    conn = db()
    try:
        row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id)).fetchone()
    finally:
        conn.close()
    
    if row:
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"✏️ <b>EDIT FORMAT ID {rid}</b>\n\nKirim data baru (format manual):"
        )
    await callback.answer()


@dp.callback_query(F.data.startswith("delete_result_"))
async def delete_result(callback: CallbackQuery):
    rid = int(callback.data.split("_")[-1])
    conn = db()
    try:
        row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id)).fetchone()
        if row:
            conn.execute("""
                INSERT INTO format_history(telegram_id,input_text,result_text,created_at,deleted_at)
                VALUES(?,?,?,?,?)
            """, (callback.from_user.id, "", row["result_text"], datetime.now().isoformat(), datetime.now().isoformat()))
            conn.execute("DELETE FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id))
            conn.commit()
    finally:
        conn.close()
    
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("🗑️ Dipindahkan ke RIWAYAT")


@dp.callback_query(F.data.startswith("save_result_"))
async def save_result(callback: CallbackQuery):
    rid = int(callback.data.split("_")[-1])
    conn = db()
    try:
        row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id)).fetchone()
        if row:
            conn.execute("""
                INSERT INTO format_history(telegram_id,input_text,result_text,created_at,deleted_at)
                VALUES(?,?,?,?,?)
            """, (callback.from_user.id, "", row["result_text"], datetime.now().isoformat(), datetime.now().isoformat()))
            conn.commit()
    finally:
        conn.close()
    await callback.answer("💾 Tersimpan ke HISTORY")


# ==================== FORMAT HISTORY ====================

@dp.callback_query(F.data == "format_history")
async def format_history(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    conn = db()
    try:
        rows = conn.execute("""
            SELECT id,result_text,deleted_at FROM format_history
            WHERE telegram_id=? ORDER BY id DESC LIMIT 20
        """, (callback.from_user.id,)).fetchall()
    finally:
        conn.close()

    try:
        await callback.message.delete()
    except:
        pass

    if not rows:
        await callback.message.bot.send_message(
            callback.from_user.id,
            "🕘 <b>RIWAYAT HASIL FORMAT - KOSONG</b>\n\nBelum ada format yang dihapus.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📄 LIHAT HASIL FORMAT", callback_data="format_results")],
                [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
            ])
        )
        await callback.answer()
        return

    await callback.message.bot.send_message(
        callback.from_user.id,
        f"🕘 <b>RIWAYAT HASIL FORMAT - {len(rows)} DATA</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗑️ HAPUS SEMUA RIWAYAT", callback_data="clear_history")],
            [InlineKeyboardButton(text="📄 LIHAT HASIL FORMAT", callback_data="format_results")]
        ])
    )

    for r in rows:
        rid = r["id"]
        txt = r["result_text"]
        try:
            dt = r["deleted_at"][:16].replace("T", " ")
        except:
            dt = ""
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"<pre>{html.escape(txt[:3800])}</pre>\n<i>#{rid} • Dihapus: {dt}</i>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="♻️ PULIHKAN", callback_data=f"restore_history_{rid}"),
                    InlineKeyboardButton(text="🗑️ HAPUS PERMANEN", callback_data=f"delete_history_{rid}")
                ]
            ])
        )
        await asyncio.sleep(0.1)
    await callback.answer(f"📂 {len(rows)} riwayat")


@dp.callback_query(F.data.startswith("restore_history_"))
async def restore_history(callback: CallbackQuery):
    hid = int(callback.data.split("_")[-1])
    conn = db()
    try:
        row = conn.execute("SELECT * FROM format_history WHERE id=? AND telegram_id=?", (hid, callback.from_user.id)).fetchone()
        if not row:
            await callback.answer("Riwayat tidak ditemukan", show_alert=True)
            return
        conn.execute("""
            INSERT INTO format_results(telegram_id,input_text,result_text,created_at)
            VALUES(?,?,?,?)
        """, (callback.from_user.id, row["input_text"], row["result_text"], datetime.now().isoformat()))
        conn.execute("DELETE FROM format_history WHERE id=? AND telegram_id=?", (hid, callback.from_user.id))
        conn.commit()
    finally:
        conn.close()
    
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("♻️ Dipulihkan ke HASIL FORMAT")


@dp.callback_query(F.data.startswith("delete_history_"))
async def delete_history(callback: CallbackQuery):
    hid = int(callback.data.split("_")[-1])
    conn = db()
    try:
        conn.execute("DELETE FROM format_history WHERE id=? AND telegram_id=?", (hid, callback.from_user.id))
        conn.commit()
    finally:
        conn.close()
    
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("🗑️ Dihapus permanen dari riwayat")


@dp.callback_query(F.data == "clear_history")
async def clear_history(callback: CallbackQuery):
    conn = db()
    try:
        conn.execute("DELETE FROM format_history WHERE telegram_id=?", (callback.from_user.id,))
        conn.commit()
    finally:
        conn.close()
    await callback.answer("🗑️ Semua riwayat dihapus")
    await format_history(callback)


# ==================== HASIL AKUN (DISEDERHANAKAN) ====================

@dp.callback_query(F.data == "hasil_akun")
async def hasil_akun(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 Terkunci.", show_alert=True)
        return
    conn = db()
    try:
        rows = conn.execute(
            "SELECT id,email,password,created_at, result_id FROM format_accounts WHERE telegram_id=? ORDER BY id DESC LIMIT 10",
            (callback.from_user.id,)
        ).fetchall()
    finally:
        conn.close()
    
    if not rows:
        await callback.message.edit_text(
            "📧 <b>HASIL FORMAT+AKUN - KOSONG</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📝 BUAT FORMAT BARU", callback_data="format_create")],
                [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
            ])
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        f"📧 <b>HASIL FORMAT+AKUN - {len(rows)} TERBARU</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
        ])
    )
    await callback.answer()

    for i, row in enumerate(rows, 1):
        rid = row["id"]
        email = html.escape(row['email'] or '-')
        pwd = html.escape(row['password'] or '-')
        result_id = row['result_id']
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="📋 SALIN", callback_data=f"copy_akun_{rid}"),
                InlineKeyboardButton(text="✏️ EDIT", callback_data=f"edit_akun_{rid}"),
                InlineKeyboardButton(text="🗑️ HAPUS", callback_data=f"del_akun_{rid}")
            ]
        ])
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"<b>{i}. ID {rid} (Format ID: {result_id})</b>\n<pre>AKUN:\n{email}\n{pwd}</pre>",
            reply_markup=kb
        )
        await asyncio.sleep(0.15)


@dp.callback_query(F.data.startswith("copy_akun_"))
async def copy_akun(callback: CallbackQuery):
    aid = int(callback.data.replace("copy_akun_", ""))
    conn = db()
    try:
        row = conn.execute("SELECT email,password FROM format_accounts WHERE id=? AND telegram_id=?", (aid, callback.from_user.id)).fetchone()
    finally:
        conn.close()
    if row:
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"📋 <b>SALIN AKUN:</b>\n<code>{html.escape(row['email'])}\n{html.escape(row['password'])}</code>"
        )
    await callback.answer("Disalin!")


@dp.callback_query(F.data.startswith("del_akun_"))
async def del_akun(callback: CallbackQuery):
    aid = int(callback.data.replace("del_akun_", ""))
    conn = db()
    try:
        conn.execute("DELETE FROM format_accounts WHERE id=? AND telegram_id=?", (aid, callback.from_user.id))
        conn.commit()
    finally:
        conn.close()
    await callback.answer("Dihapus!")
    await hasil_akun(callback)


@dp.callback_query(F.data.startswith("edit_akun_"))
async def edit_akun_start(callback: CallbackQuery, state: FSMContext):
    aid = int(callback.data.replace("edit_akun_", ""))
    await state.set_state(FormatState.waiting_edit_akun)
    await state.update_data(edit_id=aid)
    conn = db()
    try:
        row = conn.execute("SELECT email,password FROM format_accounts WHERE id=? AND telegram_id=?", (aid, callback.from_user.id)).fetchone()
    finally:
        conn.close()
    if row:
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"✏️ <b>EDIT AKUN ID {aid}</b>\n\nLama:\n<code>{row['email']}\n{row['password']}</code>\n\nKirim baru:\nemail\npassword"
        )
    await callback.answer()


@dp.message(FormatState.waiting_edit_akun, F.text)
async def edit_akun_save(message: Message, state: FSMContext):
    data = await state.get_data()
    aid = data.get("edit_id")
    lines = [x.strip() for x in message.text.splitlines() if x.strip()]
    email = lines[0] if len(lines) >= 1 else ""
    password = lines[1] if len(lines) >= 2 else ""
    if "@" not in email:
        await message.answer("❌ Email tidak valid. Kirim lagi: email\npassword")
        return
    conn = db()
    try:
        conn.execute(
            "UPDATE format_accounts SET email=?, password=?, raw_text=? WHERE id=? AND telegram_id=?",
            (email, password, f"{email}\n{password}", aid, message.from_user.id)
        )
        conn.commit()
    finally:
        conn.close()
    await state.clear()
    await message.answer(f"✅ Akun ID {aid} diedit!", reply_markup=auto_menu())


# ==================== HUBUNGI ADMIN ====================

@dp.callback_query(F.data == "contact_admin")
async def contact_admin(callback: CallbackQuery):
    await callback.message.edit_text(
        "📞 <b>HUBUNGI ADMIN</b>\n\n"
        "Jika membutuhkan bantuan, silakan hubungi Admin:\n\n"
        "👤 <b>Telegram</b>\n"
        "@Hambali1995\n\n"
        "📱 <b>WhatsApp</b>\n"
        "083160776091",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💬 CHAT TELEGRAM", url="https://t.me/Hambali1995")],
            [InlineKeyboardButton(text="📱 CHAT WHATSAPP", url="https://wa.me/6283160776091")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()


# ==================== ADMIN PANEL ====================

@dp.message(Command("admin"))
async def admin_command(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("❌ Kamu bukan Admin.")
        return
    await state.clear()
    await message.answer("🔐 <b>PANEL ADMIN</b>\n\nPilih menu:", reply_markup=admin_menu())


@dp.callback_query(F.data == "admin_cancel")
async def admin_cancel(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await callback.message.edit_text(
        "🔐 <b>PANEL ADMIN</b>\n\nPilih menu:",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.callback_query(F.data == "admin_panel")
async def admin_panel(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Kamu bukan Admin.", show_alert=True)
        return
    await callback.message.edit_text("🔐 <b>PANEL ADMIN</b>\n\nPilih menu:", reply_markup=admin_menu())
    await callback.answer()


@dp.callback_query(F.data == "admin_active")
async def admin_active(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return

    conn = db()
    try:
        rows = conn.execute("""
            SELECT u.telegram_id,u.name,u.username,s.package_name,s.expiry_date,s.status
            FROM users u JOIN subscriptions s ON s.telegram_id=u.telegram_id
            WHERE s.status='unlimited' OR (s.status='active' AND s.expiry_date>?)
            ORDER BY u.telegram_id
        """, (datetime.now().isoformat(),)).fetchall()
    finally:
        conn.close()

    if not rows:
        text = "👥 <b>USER AKTIF</b>\n\nTidak ada user aktif."
    else:
        text = "👥 <b>USER AKTIF</b>\n\n" + "\n\n".join(
            f"👤 {r['name']}\n🆔 <code>{r['telegram_id']}</code>\n"
            f"📦 {r['package_name']}\n"
            f"📅 {'SELAMANYA' if r['status']=='unlimited' else r['expiry_date'][:10]}"
            for r in rows
        )
        text += f"\n\n📊 Total user aktif: <b>{len(rows)}</b>"

    await callback.message.edit_text(text[:3900], reply_markup=admin_menu())
    await callback.answer()


@dp.callback_query(F.data == "admin_pending")
async def admin_pending(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return

    conn = db()
    try:
        rows = conn.execute(
            "SELECT * FROM transactions WHERE status='pending' ORDER BY id DESC"
        ).fetchall()
    finally:
        conn.close()

    if not rows:
        await callback.message.edit_text(
            "💰 <b>TRANSAKSI PENDING</b>\n\nTidak ada transaksi pending.",
            reply_markup=admin_menu()
        )
        await callback.answer()
        return

    for row in rows:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ SETUJU", callback_data=f"approve_{row['id']}"),
            InlineKeyboardButton(text="❌ TOLAK", callback_data=f"reject_{row['id']}")
        ]])
        caption = (
            "💰 <b>TRANSAKSI PENDING</b>\n\n"
            f"🧾 #{row['id']}\n🆔 <code>{row['telegram_id']}</code>\n"
            f"💰 {rupiah(row['amount'])}\n📦 {row['package_name'] or 'TOP UP SALDO'}"
        )
        if row["proof_file_id"]:
            await callback.bot.send_photo(callback.from_user.id, row["proof_file_id"],
                                          caption=caption, reply_markup=keyboard)
        else:
            await callback.bot.send_message(callback.from_user.id, caption, reply_markup=keyboard)

    await callback.message.edit_text(
        f"💰 <b>TRANSAKSI PENDING</b>\n\nDitemukan <b>{len(rows)}</b> transaksi.",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.callback_query(F.data == "admin_add_balance")
async def admin_add_balance_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_user_amount)
    await state.update_data(action="add")
    await callback.message.edit_text(
        "➕ <b>TAMBAH SALDO</b>\n\nKirim:\n<code>ID_USER NOMINAL</code>\n\nContoh: <code>123456789 50000</code>",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.callback_query(F.data == "admin_sub_balance")
async def admin_sub_balance_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_user_amount)
    await state.update_data(action="sub")
    await callback.message.edit_text(
        "➖ <b>KURANGI SALDO</b>\n\nKirim:\n<code>ID_USER NOMINAL</code>\n\nContoh: <code>123456789 50000</code>",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.message(AdminState.waiting_user_amount, F.text)
async def admin_balance_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return

    data = await state.get_data()
    parts = message.text.strip().split()
    if len(parts) != 2 or not all(p.isdigit() for p in parts):
        await message.answer("❌ Format salah. Contoh: <code>123456789 50000</code>")
        return

    user_id, amount = int(parts[0]), int(parts[1])
    conn = db()
    try:
        user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (user_id,)).fetchone()
        if not user:
            await message.answer("❌ User tidak ditemukan.")
            return

        if data.get("action") == "add":
            conn.execute("UPDATE users SET balance=balance+? WHERE telegram_id=?", (amount, user_id))
            action_text = f"➕ Ditambah {rupiah(amount)}"
        else:
            conn.execute("UPDATE users SET balance=MAX(balance-?,0) WHERE telegram_id=?", (amount, user_id))
            action_text = f"➖ Dikurangi {rupiah(amount)}"

        new_balance = conn.execute(
            "SELECT balance FROM users WHERE telegram_id=?", (user_id,)
        ).fetchone()["balance"]
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()

    await message.answer(
        "✅ <b>BERHASIL</b>\n\n"
        f"👤 User : <code>{user_id}</code>\n{action_text}\n"
        f"💰 Saldo sekarang : <b>{rupiah(new_balance)}</b>",
        reply_markup=admin_menu()
    )
    try:
        await message.bot.send_message(
            user_id,
            "💰 <b>PERUBAHAN SALDO</b>\n\n"
            f"{action_text}\nSaldo sekarang : <b>{rupiah(new_balance)}</b>"
        )
    except Exception:
        pass


@dp.callback_query(F.data == "admin_delete_user")
async def admin_delete_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_delete_user)
    await callback.message.edit_text(
        "🗑️ <b>HAPUS USER</b>\n\nKirim Telegram ID user.\nContoh: <code>123456789</code>",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.message(AdminState.waiting_delete_user, F.text)
async def admin_delete_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return
    if not message.text.strip().isdigit():
        await message.answer("❌ Telegram ID harus angka.")
        return

    user_id = int(message.text.strip())
    if user_id in ADMIN_IDS:
        await message.answer("❌ User Admin tidak boleh dihapus.")
        return

    conn = db()
    try:
        user = conn.execute("SELECT telegram_id FROM users WHERE telegram_id=?", (user_id,)).fetchone()
        if not user:
            await message.answer("❌ User tidak ditemukan.")
            return

        for table in ["users", "subscriptions", "transactions", "format_settings", "format_results", "format_history"]:
            conn.execute(f"DELETE FROM {table} WHERE telegram_id=?", (user_id,))
        conn.commit()
    finally:
        conn.close()
    
    await state.clear()
    await message.answer(
        f"🗑️ User <code>{user_id}</code> berhasil dihapus.",
        reply_markup=admin_menu()
    )


@dp.callback_query(F.data == "admin_broadcast")
async def admin_broadcast_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_broadcast)
    await callback.message.edit_text(
        "📢 <b>BROADCAST</b>\n\nKirim pesan/foto/dokumen yang ingin dikirim ke semua user.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ BATAL", callback_data="admin_cancel")]
        ])
    )
    await callback.answer()


@dp.message(AdminState.waiting_broadcast)
async def admin_broadcast_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return

    conn = db()
    try:
        users = conn.execute("SELECT telegram_id FROM users").fetchall()
    finally:
        conn.close()

    success = failed = 0
    for row in users:
        try:
            if message.text:
                await message.bot.send_message(row["telegram_id"], message.text)
            elif message.photo:
                await message.bot.send_photo(row["telegram_id"], message.photo[-1].file_id, caption=message.caption or "")
            elif message.document:
                await message.bot.send_document(row["telegram_id"], message.document.file_id, caption=message.caption or "")
            else:
                failed += 1
                continue
            success += 1
            await asyncio.sleep(0.05)
        except Exception:
            failed += 1

    await state.clear()
    await message.answer(
        "📢 <b>BROADCAST SELESAI</b>\n\n"
        f"✅ Berhasil : {success}\n❌ Gagal : {failed}",
        reply_markup=admin_menu()
    )


# ==================== FALLBACK ====================

@dp.message(F.text)
async def fallback(message: Message):
    # Deteksi masalah JMO
    words = ["jmo", "error", "kode", "025", "026", "027", "028", "029", "030", "031", "032", "033", 
             "login", "verifikasi", "kpj", "jht", "otp", "wajah", "password", "email", "aktivasi", 
             "klaim", "bpu", "kamera", "server", "lemot", "notifikasi", "daftar", "cair"]
    if any(w in message.text.lower() for w in words):
        await message.answer(JMOSolutionManager.get_solution(message.text))
    else:
        await message.answer("Silakan gunakan menu utama dengan /start", reply_markup=main_menu())


# ==================== MAIN ====================

async def main():
    init_database()
    logging.basicConfig(level=logging.WARNING, format="%(asctime)s | %(levelname)s | %(message)s")
    logging.info("🤖 SABABAT JHT sedang berjalan")

    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML)
    )
    try:
        await dp.start_polling(bot, allowed_updates=["message", "callback_query", "my_chat_member"])
    finally:
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())