import asyncio
import logging
import html
import os
import sqlite3
import re
from aiohttp import web
from datetime import datetime
from pathlib import Path
from calendar import monthrange

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, Message

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_IDS = {int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()}

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN belum diisi di Railway Variables")

DB_PATH = Path("/data/bot.db") if Path("/data").exists() else Path("bot.db")

DEFAULT_TEMPLATE = """📍 KAB : {KAB}
📍 KEC : {KEC}
📍 KEL : {KEL}

💰 SALDO : {SALDO}

🚻 KELAMIN : {KELAMIN}
💳 KPJ : {KPJ}
🎯 SENSOR : {SENSOR}
📆 IT : {IT}

🏛️ PT : {PT}

DPT JMO LASIK ✅"""


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_database():
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        telegram_id INTEGER PRIMARY KEY,
        name TEXT,
        username TEXT,
        balance INTEGER DEFAULT 0,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS subscriptions (
        telegram_id INTEGER PRIMARY KEY,
        package_code TEXT,
        package_name TEXT,
        price INTEGER DEFAULT 0,
        start_date TEXT,
        expiry_date TEXT,
        status TEXT DEFAULT 'inactive'
    );
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        amount INTEGER NOT NULL,
        payment_method TEXT DEFAULT 'SEABANK/DANA',
        package_code TEXT,
        package_name TEXT,
        proof_file_id TEXT,
        status TEXT DEFAULT 'pending',
        created_at TEXT NOT NULL,
        processed_at TEXT
    );
    CREATE TABLE IF NOT EXISTS format_settings (
        telegram_id INTEGER PRIMARY KEY,
        template TEXT NOT NULL,
        updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        input_text TEXT,
        result_text TEXT,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_codes (
        telegram_id INTEGER PRIMARY KEY,
        prefix TEXT NOT NULL,
        suffix TEXT DEFAULT '',
        current_number INTEGER NOT NULL,
        padding INTEGER NOT NULL,
        enabled INTEGER DEFAULT 1,
        updated_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS admins (
        telegram_id INTEGER PRIMARY KEY,
        added_by INTEGER,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        email TEXT,
        password TEXT,
        raw_text TEXT,
        result_id INTEGER,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS user_kota (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        kota_name TEXT NOT NULL,
        kab TEXT DEFAULT '',
        kec TEXT DEFAULT '',
        kel TEXT DEFAULT '',
        catatan TEXT DEFAULT '',
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS blacklist_numbers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        phone TEXT NOT NULL UNIQUE,
        reason TEXT DEFAULT '',
        added_by INTEGER,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS format_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        telegram_id INTEGER NOT NULL,
        input_text TEXT,
        result_text TEXT,
        created_at TEXT NOT NULL,
        deleted_at TEXT NOT NULL
    );
    """)
    conn.executescript("""
    CREATE INDEX IF NOT EXISTS idx_transactions_status ON transactions(status);
    CREATE INDEX IF NOT EXISTS idx_transactions_user ON transactions(telegram_id);
    CREATE INDEX IF NOT EXISTS idx_format_results_user ON format_results(telegram_id);
    CREATE INDEX IF NOT EXISTS idx_user_kota_user ON user_kota(telegram_id);
    """)
    conn.commit()
    conn.close()


def register_user(user):
    conn = db()
    conn.execute("""
        INSERT INTO users(telegram_id,name,username,balance,created_at)
        VALUES(?,?,?,0,?)
        ON CONFLICT(telegram_id) DO UPDATE SET
            name=excluded.name, username=excluded.username
    """, (user.id, user.full_name, user.username, datetime.now().isoformat()))
    conn.commit()
    conn.close()


def rupiah(value):
    return f"Rp {int(value):,}".replace(",", ".")


def is_admin(user_id):
    if user_id in ADMIN_IDS:
        return True
    try:
        conn=db()
        row=conn.execute("SELECT telegram_id FROM admins WHERE telegram_id=?",(user_id,)).fetchone()
        conn.close()
        return row is not None
    except:
        return False

def get_all_admins():
    try:
        conn=db()
        rows=conn.execute("SELECT telegram_id FROM admins").fetchall()
        conn.close()
        return ADMIN_IDS.union({r['telegram_id'] for r in rows})
    except:
        return ADMIN_IDS



def get_subscription(user_id):
    conn = db()
    row = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (user_id,)).fetchone()
    conn.close()
    return row


def has_auto_format_access(user_id):
    if is_admin(user_id):
        return True
    sub = get_subscription(user_id)
    if not sub:
        return False
    if sub["status"] == "unlimited":
        return True
    if sub["status"] == "active" and sub["expiry_date"]:
        try:
            return datetime.now() < datetime.fromisoformat(sub["expiry_date"])
        except ValueError:
            return False
    return False


def add_months(dt, months):
    idx = dt.month - 1 + months
    year = dt.year + idx // 12
    month = idx % 12 + 1
    day = min(dt.day, monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)



# ==================== PAKET KHUSUS TAMBAH KOTA ====================
KOTA_PACKAGES = {
    "kota_1w": ("KOTA_1W", "TAMBAH KOTA 1 MINGGU", 35000, 7),
    "kota_1m": ("KOTA_1M", "TAMBAH KOTA 1 BULAN", 120000, 30),
    "kota_2m": ("KOTA_2M", "TAMBAH KOTA 2 BULAN", 200000, 60),
    "kota_6m": ("KOTA_6M", "TAMBAH KOTA 6 BULAN", 500000, 180),
    "kota_unlimited": ("KOTA_UNLIMITED", "TAMBAH KOTA UNLIMITED", 2000000, None),
}

# PAKET KHUSUS CARI KOTA LAIN
CARI_KOTA_PACKAGES = {
    "cari_1w": ("CARI_1W", "CARI KOTA LAIN 1 MINGGU", 15000, 7),
    "cari_1m": ("CARI_1M", "CARI KOTA LAIN 1 BULAN", 50000, 30),
    "cari_2m": ("CARI_2M", "CARI KOTA LAIN 2 BULAN", 80000, 60),
}

def get_cari_subscription(user_id):
    conn = db()
    rows = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=? AND package_code LIKE 'CARI_%' ORDER BY expiry_date DESC", (user_id,)).fetchall()
    for r in rows:
        if r["status"] == "unlimited":
            conn.close()
            return r
        if r["expiry_date"]:
            try:
                from datetime import datetime as dt
                if dt.now() < dt.fromisoformat(r["expiry_date"]):
                    conn.close()
                    return r
            except:
                pass
    conn.close()
    return None

def has_cari_access(user_id):
    sub = get_cari_subscription(user_id)
    if not sub:
        return False
    if sub["status"] == "unlimited":
        return True
    if sub["expiry_date"]:
        try:
            from datetime import datetime as dt
            return dt.now() < dt.fromisoformat(sub["expiry_date"])
        except:
            return False
    return False


def get_kota_subscription(user_id):
    conn = db()
    row = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=? AND package_code LIKE 'KOTA_%'", (user_id,)).fetchone()
    # cek yang terbaru yang masih aktif
    if not row:
        # cari semua kota packages yang aktif
        rows = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=? AND package_code LIKE 'KOTA_%' ORDER BY expiry_date DESC", (user_id,)).fetchall()
        for r in rows:
            if r["status"] == "unlimited":
                conn.close()
                return r
            if r["expiry_date"]:
                try:
                    from datetime import datetime as dt
                    if dt.now() < dt.fromisoformat(r["expiry_date"]):
                        conn.close()
                        return r
                except:
                    pass
        conn.close()
        return None
    conn.close()
    return row

def has_kota_access(user_id):
    sub = get_kota_subscription(user_id)
    if not sub:
        return False
    if sub["status"] == "unlimited":
        return True
    if sub["expiry_date"]:
        try:
            from datetime import datetime as dt
            return dt.now() < dt.fromisoformat(sub["expiry_date"])
        except:
            return False
    return False

def get_kota_usage(user_id):
    conn = db()
    # hitung berapa kali tambah kota setelah langganan terakhir
    sub = get_kota_subscription(user_id)
    if not sub:
        conn.close()
        return 0, 0
    start_date = sub["start_date"]
    # hitung jumlah kota yang ditambahkan setelah start_date
    count = conn.execute("SELECT COUNT(*) as c FROM user_kota WHERE telegram_id=? AND created_at >= ?", (user_id, start_date)).fetchone()["c"]
    conn.close()
    # quota 2 kali per paket
    quota = 2
    if sub["package_code"] == "KOTA_UNLIMITED":
        quota = 999999
    return count, quota

def check_kota_can_add(user_id):
    if not has_kota_access(user_id):
        return False, "belum_langganan"
    used, quota = get_kota_usage(user_id)
    if used >= quota:
        return False, "quota_habis"
    return True, f"{quota-used} sisa"


# ==================== PERBAIKAN TOMBOL MENU (DILEBARKAN) ====================


def get_format_code(user_id):
    conn = db()
    row = conn.execute("SELECT * FROM format_codes WHERE telegram_id=?", (user_id,)).fetchone()
    conn.close()
    return row

def generate_next_code(user_id):
    conn = db()
    row = conn.execute("SELECT * FROM format_codes WHERE telegram_id=?", (user_id,)).fetchone()
    if not row or not row["enabled"]:
        conn.close()
        return None
    prefix = row["prefix"]
    num = row["current_number"]
    padding = row["padding"]
    # format kode: JPG - 001
    code_str = f"{prefix} - {str(num).zfill(padding)}"
    # increment untuk selanjutnya
    conn.execute("UPDATE format_codes SET current_number=current_number+1 WHERE telegram_id=?", (user_id,))
    conn.commit()
    conn.close()
    return code_str


def parse_kode_input(text):
    """
    Kode BEBAS - boleh pakai icon atau polos, yang penting ada angka di akhirnya.
    Contoh valid:
    JPG - 001
    🤖 JPG - 001 🤖
    🚀 JPG 001
    JPG 001 🔥
    """
    original = text.strip()
    if not original:
        return None
    
    # Cari angka di akhir (misal 001)
    m = re.search(r'(\d+)\s*[^\d]*$', original)
    if not m:
        # Kalau tidak ada angka, anggap 001
        return original, 1, 3, ""  # prefix, num, padding, suffix
    
    num_str = m.group(1)
    num_start = m.start(1)
    num_end = m.end(1)
    
    prefix = original[:num_start].strip()
    suffix = original[num_end:].strip()
    
    # Bersihkan prefix dari spasi/ dash berlebih di akhir tapi pertahankan icon
    # prefix misal "🤖 JPG - " -> tetap
    try:
        num = int(num_str)
    except:
        num = 1
    
    padding = len(num_str)
    if padding < 3:
        padding = 3
    
    # Jika prefix kosong, pakai JPG
    if not prefix and not suffix:
        prefix = "JPG -"
    
    return prefix, num, padding, suffix

def get_format_code(user_id):
    conn = db()
    row = conn.execute("SELECT * FROM format_codes WHERE telegram_id=?", (user_id,)).fetchone()
    conn.close()
    return row

def generate_next_code(user_id):
    conn = db()
    row = conn.execute("SELECT * FROM format_codes WHERE telegram_id=?", (user_id,)).fetchone()
    if not row or not row["enabled"]:
        conn.close()
        return None
    
    prefix = row["prefix"]
    suffix = row["suffix"] if "suffix" in row.keys() and row["suffix"] else ""
    num = row["current_number"]
    padding = row["padding"]
    
    # Gabung prefix + number + suffix
    if suffix:
        code_str = f"{prefix} {str(num).zfill(padding)} {suffix}".strip()
    else:
        code_str = f"{prefix} {str(num).zfill(padding)}".strip()
        # rapikan double spasi dan dash
        code_str = re.sub(r'\s+', ' ', code_str)
        # kalau sudah ada " -" jangan double
        if not any(c.isdigit() for c in prefix[-3:]):
            # sudah ada pemisah
            pass
    
    # increment
    conn.execute("UPDATE format_codes SET current_number=current_number+1 WHERE telegram_id=?", (user_id,))
    conn.commit()
    conn.close()
    return code_str


def main_menu(user_id=None):
    if user_id is None:
        user_id = 0
    # ORDER FINAL SESUAI REQUEST: HAPUS KOTA SAYA setelah KOTA SAYA
    rows = [
        [
            InlineKeyboardButton(text="👤 PROFIL", callback_data="profile"),
            InlineKeyboardButton(text="🏙️ TAMBAH KOTA", callback_data="kota_add")
        ],
        [
            InlineKeyboardButton(text="🌆 KOTA SAYA", callback_data="kota_list"),
            InlineKeyboardButton(text="🗑️ HAPUS KOTA SAYA", callback_data="hapus_kota_saya")
        ],
        [
            InlineKeyboardButton(text="🔍 CARI KOTA LAIN", callback_data="kota_search_lain"),
            InlineKeyboardButton(text="🚫 NO BLACKLIST", callback_data="no_blacklist")
        ],
        [
            InlineKeyboardButton(text="📊 STATUS", callback_data="status"),
            InlineKeyboardButton(text="💳 TOP UP", callback_data="topup")
        ],
        [
            InlineKeyboardButton(text="📝 AUTO FORMAT", callback_data="auto_format"),
            InlineKeyboardButton(text="💡 SOLUSI JMO", callback_data="solusi_jmo")
        ],
        [
            InlineKeyboardButton(text="📞 HUBUNGI ADMIN", callback_data="contact_admin"),
            InlineKeyboardButton(text="ℹ️ BANTUAN", callback_data="info_bot")
        ],
    ]
    if is_admin(user_id):
        rows.append([
            InlineKeyboardButton(text="🔐 PANEL ADMIN", callback_data="admin_panel")
        ])
    return InlineKeyboardMarkup(inline_keyboard=rows)



def main_menu_with_colors_note():
    return main_menu()



def main_menu_with_colors_note():
    # Untuk ReplyKeyboard yang bisa warna (jika pakai WebApp), tapi tetap fallback ke inline
    return main_menu()



def back_main():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
    ])


def auto_menu():
    # UPDATE: 2 KOTAK PER BARIS
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📝 BUAT FORMAT", callback_data="format_create"),
            InlineKeyboardButton(text="📄 HASIL TERKINI", callback_data="format_results")
        ],
        [
            InlineKeyboardButton(text="📧 FORMAT+AKUN", callback_data="hasil_akun"),
            InlineKeyboardButton(text="⚙️ SETTING", callback_data="format_setting")
        ],
        [
            InlineKeyboardButton(text="🕘 RIWAYAT", callback_data="format_history"),
            InlineKeyboardButton(text="🏙️ TAMBAH KOTA", callback_data="kota_add")
        ],
        [
            InlineKeyboardButton(text="🌆 KOTA SAYA", callback_data="kota_list"),
            InlineKeyboardButton(text="🔍 CARI KOTA", callback_data="kota_search_lain")
        ],
        [
            InlineKeyboardButton(text="💳 TOP UP", callback_data="topup"),
            InlineKeyboardButton(text="📊 STATUS", callback_data="status")
        ],
        [
            InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")
        ]
    ])

class KotaState(StatesGroup):
    waiting_kota_add = State()
    waiting_lain_search = State()
    waiting_provinsi = State()
    waiting_kabupaten = State()
    waiting_kecamatan = State()

class WilayahState(StatesGroup):
    provinsi = State()
    kabupaten = State()
    kecamatan_multi = State()

# DATA PROVINSI LENGKAP
PROVINSI_LIST = [
    "Aceh", "Sumatera Utara", "Sumatera Barat", "Riau", "Jambi",
    "Sumatera Selatan", "Bengkulu", "Lampung", "Kep. Bangka Belitung", "Kep. Riau",
    "DKI Jakarta", "Jawa Barat", "Jawa Tengah", "DI Yogyakarta", "Jawa Timur",
    "Banten", "Bali", "Nusa Tenggara Barat", "Nusa Tenggara Timur",
    "Kalimantan Barat", "Kalimantan Tengah", "Kalimantan Selatan", "Kalimantan Timur", "Kalimantan Utara",
    "Sulawesi Utara", "Sulawesi Tengah", "Sulawesi Selatan", "Sulawesi Tenggara", "Gorontalo", "Sulawesi Barat",
    "Maluku", "Maluku Utara", "Papua", "Papua Barat", "Papua Tengah", "Papua Pegunungan", "Papua Selatan", "Papua Barat Daya"
]

# Dummy data kabupaten per provinsi (bisa diisi lengkap nanti)
KABUPATEN_DATA = {
    "Aceh": ["Banda Aceh", "Langsa", "Lhokseumawe", "Sabang", "Subulussalam", "Aceh Barat", "Aceh Besar"],
    "Sumatera Utara": ["Medan", "Binjai", "Tebing Tinggi", "Pematang Siantar", "Deli Serdang", "Langkat"],
    "Banten": ["Serang", "Cilegon", "Tangerang", "Tangerang Selatan", "Kab. Serang", "Kab. Tangerang", "Lebak", "Pandeglang"],
    "Jawa Barat": ["Bandung", "Bekasi", "Bogor", "Depok", "Cimahi", "Kab. Bandung", "Kab. Bogor", "Karawang", "Cirebon"],
    "Jawa Tengah": ["Semarang", "Solo", "Magelang", "Pekalongan", "Kab. Semarang", "Kab. Banyumas"],
    "Jawa Timur": ["Surabaya", "Malang", "Kediri", "Mojokerto", "Sidoarjo", "Gresik"],
    "DKI Jakarta": ["Jakarta Pusat", "Jakarta Utara", "Jakarta Barat", "Jakarta Selatan", "Jakarta Timur", "Kep. Seribu"],
}

# Dummy kecamatan per kabupaten
KECAMATAN_DATA = {
    "Serang": ["Serang", "Cipocok Jaya", "Curug", "Kasemen", "Taktakan", "Walantaka", "Ciruas", "Kramatwatu", "Waringinkurung", "Pontang", "Tirtayasa", "Tanara", "Carenang", "Binuang", "Petir", "Tunjung Teja", "Cikeusal", "Pamarayan", "Kopo", "Jawilan", "Ciomas", "Padarincang", "Gunung Sari", "Baros", "Pabuaran", "Lebak Wangi", "Bandung", "Cikande", "Kibin", "Kragilan"],
    "Kab. Serang": ["Ciruas", "Kramatwatu", "Waringinkurung", "Pontang", "Tirtayasa", "Tanara", "Carenang", "Binuang", "Petir", "Tunjung Teja", "Cikeusal", "Pamarayan", "Kopo", "Jawilan", "Ciomas", "Padarincang"],
    "Tangerang": ["Tangerang", "Ciledug", "Cipondoh", "Karawaci", "Jatiuwung", "Benda", "Neglasari", "Pinang", "Karawaci"],
    "Jakarta Pusat": ["Gambir", "Sawah Besar", "Kemayoran", "Senen", "Cempaka Putih", "Menteng", "Tanah Abang", "Johar Baru"],
}

def get_kabupaten_list(provinsi):
    return KABUPATEN_DATA.get(provinsi, [f"Kab. {provinsi} 1", f"Kab. {provinsi} 2", f"Kota {provinsi}"])

def get_kecamatan_list(kabupaten):
    # kalau ada data spesifik, pakai itu, kalau tidak generate dummy
    if kabupaten in KECAMATAN_DATA:
        return KECAMATAN_DATA[kabupaten]
    return [f"{kabupaten} - Kec 1", f"{kabupaten} - Kec 2", f"{kabupaten} - Kec 3", f"{kabupaten} - Kec 4", f"{kabupaten} - Kec 5", f"{kabupaten} - Kec 6", f"{kabupaten} - Kec 7", f"{kabupaten} - Kec 8"]

def build_single_column_keyboard(items, prefix, back_callback="kota_add"):
    # 1 KOTAK PER BARIS - BIAR JELAS
    kb = []
    for item in items:
        # callback data harus pendek, pakai index
        kb.append([InlineKeyboardButton(text=item, callback_data=f"{prefix}_{item[:30]}")])
    kb.append([InlineKeyboardButton(text="⬅️ KEMBALI", callback_data=back_callback)])
    return InlineKeyboardMarkup(inline_keyboard=kb)



def admin_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👥 CEK USER AKTIF", callback_data="admin_active")],
        [InlineKeyboardButton(text="💰 TRANSAKSI PENDING", callback_data="admin_pending")],
        [InlineKeyboardButton(text="➕ TAMBAH SALDO USER", callback_data="admin_add_balance")],
        [InlineKeyboardButton(text="➖ KURANGI SALDO USER", callback_data="admin_sub_balance")],
        [InlineKeyboardButton(text="🚫 LIHAT NO BLACKLIST", callback_data="no_blacklist")],
        [InlineKeyboardButton(text="➕ TAMBAH NO BLACKLIST", callback_data="admin_blacklist_add")],
        [InlineKeyboardButton(text="🗑️ HAPUS NO BLACKLIST", callback_data="admin_blacklist_del")],
        [InlineKeyboardButton(text="👥 DAFTAR ADMIN", callback_data="admin_list_admin")],
        [InlineKeyboardButton(text="➕ TAMBAH ADMIN", callback_data="admin_tambah_admin")],
        [InlineKeyboardButton(text="➖ HAPUS ADMIN", callback_data="admin_hapus_admin")],
        [InlineKeyboardButton(text="🗑️ HAPUS USER", callback_data="admin_delete_user")],
        [InlineKeyboardButton(text="📢 BROADCAST", callback_data="admin_broadcast")],
        [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
    ])



class PaymentState(StatesGroup):
    waiting_topup_amount = State()
    waiting_proof = State()


class JmoState(StatesGroup):
    waiting_question = State()


class FormatState(StatesGroup):
    waiting_kode = State()
    waiting_manual = State()
    waiting_excel = State()
    waiting_setting = State()
    waiting_search = State()
    waiting_search_akun = State()
    waiting_edit = State()
    waiting_edit_akun = State()


class AdminState(StatesGroup):
    waiting_user_amount = State()
    waiting_delete_user = State()
    waiting_broadcast = State()
    waiting_add_admin = State()
    waiting_remove_admin = State()


dp = Dispatcher()


# ==================== DATABASE SOLUSI JMO SUPER LENGKAP ====================

JMO_SOLUTIONS = {
    # ERROR KODE
    "025": {
        "keywords": ["025", "error 025", "kode 025", "pesan 025", "025 jmo"],
        "solution": """🔎 <b>MASALAH KODE 025 - JMO</b>

❌ <b>Penyebab:</b>
• Data identitas tidak sesuai dengan database BPJS Ketenagakerjaan
• NIK, nama, atau tanggal lahir tidak match
• Kepesertaan belum terdaftar atau tidak aktif
• Sistem JMO sedang maintenance

✅ <b>Solusi:</b>
1️⃣ Periksa kembali NIK, Nama, dan Tanggal Lahir
2️⃣ Pastikan kepesertaan BPJS Ketenagakerjaan aktif
3️⃣ Update aplikasi JMO ke versi terbaru
4️⃣ Coba login ulang setelah 1x24 jam
5️⃣ Hubungi kantor cabang BPJS terdekat jika masih gagal

📌 <b>Tips:</b> Pastikan data yang diinput SAMA PERSIS dengan kartu kepesertaan."""
    },
    
    "026": {
        "keywords": ["026", "error 026", "kode 026"],
        "solution": """🔎 <b>MASALAH KODE 026 - JMO</b>

❌ <b>Penyebab:</b>
• Nomor KPJ tidak ditemukan atau tidak valid
• Kepesertaan sudah berakhir (resign/pensiun)
• Data kepesertaan belum sinkron

✅ <b>Solusi:</b>
1️⃣ Cek nomor KPJ di kartu kepesertaan fisik
2️⃣ Pastikan status kepesertaan masih aktif
3️⃣ Jika baru daftar, tunggu 1-2 hari kerja
4️⃣ Hubungi HRD/Perusahaan untuk cek kepesertaan
5️⃣ Kunjungi kantor BPJS terdekat untuk update data

📌 <b>Tips:</b> Simpan screenshot error untuk dibawa ke kantor BPJS."""
    },
    
    "027": {
        "keywords": ["027", "error 027", "kode 027"],
        "solution": """🔎 <b>MASALAH KODE 027 - JMO</b>

❌ <b>Penyebab:</b>
• Email tidak terdaftar atau tidak aktif
• Email sudah digunakan oleh akun lain
• Format email tidak valid

✅ <b>Solusi:</b>
1️⃣ Gunakan email aktif (Gmail, Yahoo, dll)
2️⃣ Cek folder SPAM untuk email verifikasi
3️⃣ Gunakan email yang belum terdaftar
4️⃣ Pastikan format email benar (contoh: nama@gmail.com)

📌 <b>Tips:</b> Gunakan email yang selalu diakses untuk memudahkan pemulihan akun."""
    },
    
    "028": {
        "keywords": ["028", "error 028", "kode 028"],
        "solution": """🔎 <b>MASALAH KODE 028 - JMO</b>

❌ <b>Penyebab:</b>
• Nomor HP tidak aktif atau tidak terdaftar
• Format nomor HP salah
• Nomor HP sudah digunakan oleh akun lain

✅ <b>Solusi:</b>
1️⃣ Pastikan nomor HP aktif dan dapat SMS
2️⃣ Gunakan format 08xxxxxxxx (tanpa +62)
3️⃣ Cek sinyal dan jaringan HP
4️⃣ Tunggu beberapa saat lalu coba lagi

📌 <b>Tips:</b> Gunakan nomor HP yang terdaftar di bank untuk memudahkan verifikasi."""
    },
    
    "029": {
        "keywords": ["029", "error 029", "kode 029"],
        "solution": """🔎 <b>MASALAH KODE 029 - JMO</b>

❌ <b>Penyebab:</b>
• Verifikasi wajah (face recognition) gagal
• Pencahayaan kurang atau terlalu terang
• Wajah tidak terlihat jelas
• Menggunakan foto, bukan wajah asli

✅ <b>Solusi:</b>
1️⃣ Cari tempat dengan pencahayaan cukup
2️⃣ Hapus kacamata/aksesoris yang menutupi wajah
3️⃣ Posisikan wajah di tengah frame
4️⃣ Jangan gunakan foto atau video
5️⃣ Pastikan izin kamera aktif di pengaturan HP

📌 <b>Tips:</b> 
• Lakukan di ruangan terang tapi tidak silau
• Wajah harus menghadap kamera
• Ikuti instruksi gerakan (blink, tersenyum)"""
    },
    
    "030": {
        "keywords": ["030", "error 030", "kode 030"],
        "solution": """🔎 <b>MASALAH KODE 030 - JMO</b>

❌ <b>Penyebab:</b>
• Waktu sesi login habis (timeout)
• Koneksi internet tidak stabil
• Server JMO sedang sibuk

✅ <b>Solusi:</b>
1️⃣ Periksa koneksi internet (WiFi/Data)
2️⃣ Login ulang ke aplikasi
3️⃣ Tunggu 5-10 menit lalu coba lagi
4️⃣ Gunakan jaringan yang stabil
5️⃣ Clear cache aplikasi JMO

📌 <b>Tips:</b> Biasanya terjadi saat jam sibuk (08:00-10:00)."""
    },
    
    "031": {
        "keywords": ["031", "error 031", "kode 031"],
        "solution": """🔎 <b>MASALAH KODE 031 - JMO</b>

❌ <b>Penyebab:</b>
• Aplikasi JMO versi lama
• Ada bug/error pada aplikasi
• Konflik dengan aplikasi lain

✅ <b>Solusi:</b>
1️⃣ Update aplikasi JMO ke versi terbaru
2️⃣ Restart HP dan buka ulang aplikasi
3️⃣ Clear cache dan data aplikasi (hati-hati)
4️⃣ Reinstall aplikasi JMO
5️⃣ Cek kompatibilitas dengan sistem HP

📌 <b>Tips:</b> Backup data penting sebelum reinstall."""
    },
    
    "032": {
        "keywords": ["032", "error 032", "kode 032"],
        "solution": """🔎 <b>MASALAH KODE 032 - JMO</b>

❌ <b>Penyebab:</b>
• Data peserta tidak ditemukan
• Status kepesertaan non-aktif
• Perubahan data belum sinkron

✅ <b>Solusi:</b>
1️⃣ Cek status kepesertaan di aplikasi
2️⃣ Hubungi HRD untuk verifikasi kepesertaan
3️⃣ Kunjungi kantor BPJS terdekat
4️⃣ Tunggu proses sinkronisasi data

📌 <b>Tips:</b> Bawa KTP dan kartu kepesertaan saat ke kantor BPJS."""
    },
    
    "033": {
        "keywords": ["033", "error 033", "kode 033"],
        "solution": """🔎 <b>MASALAH KODE 033 - JMO</b>

❌ <b>Penyebab:</b>
• Password yang dimasukkan salah
• Password telah kadaluarsa
• Akun terkunci karena terlalu banyak percobaan

✅ <b>Solusi:</b>
1️⃣ Gunakan fitur "Lupa Password"
2️⃣ Reset password melalui email/HP terdaftar
3️⃣ Tunggu 15 menit jika akun terkunci
4️⃣ Gunakan password dengan kombinasi huruf/angka
5️⃣ Pastikan CapsLock tidak aktif

📌 <b>Tips:</b> Gunakan password yang mudah diingat tetapi aman."""
    },
    
    # MASALAH UMUM
    "login": {
        "keywords": ["login", "tidak bisa login", "gagal login", "masuk", "sign in"],
        "solution": """🔎 <b>MASALAH LOGIN JMO</b>

❌ <b>Penyebab:</b>
• Email/HP dan password tidak cocok
• Akun belum terdaftar
• Server JMO bermasalah
• Koneksi internet tidak stabil

✅ <b>Solusi:</b>
1️⃣ Periksa kembali email/HP dan password
2️⃣ Gunakan fitur "Lupa Kata Sandi"
3️⃣ Cek koneksi internet
4️⃣ Update aplikasi ke versi terbaru
5️⃣ Coba login di waktu lain (off-peak)

📌 <b>Tips:</b> 
• Simpan password di tempat aman
• Gunakan login dengan Face ID/Fingerprint jika tersedia"""
    },
    
    "lupa password": {
        "keywords": ["password", "kata sandi", "lupa sandi", "lupa password", "reset password"],
        "solution": """🔎 <b>LUPA PASSWORD JMO</b>

❌ <b>Penyebab:</b>
• Lupa password yang digunakan
• Password kadaluarsa
• Akun dibajak

✅ <b>Solusi:</b>
1️⃣ Klik "Lupa Kata Sandi" di halaman login
2️⃣ Masukkan email atau HP terdaftar
3️⃣ Ikuti instruksi reset via email/SMS
4️⃣ Buat password baru yang kuat
5️⃣ Simpan password di tempat aman

📌 <b>Tips:</b> 
• Gunakan kombinasi huruf besar, kecil, angka & simbol
• Jangan gunakan password yang sama dengan akun lain
• Aktifkan 2FA jika tersedia"""
    },
    
    "verifikasi wajah": {
        "keywords": ["verifikasi wajah", "face", "wajah", "face recognition", "selfie"],
        "solution": """🔎 <b>MASALAH VERIFIKASI WAJAH JMO</b>

❌ <b>Penyebab:</b>
• Pencahayaan kurang atau berlebihan
• Wajah tidak terlihat jelas
• Menggunakan foto, bukan selfie langsung
• Kacamata/penutup wajah
• Ekspresi wajah tidak sesuai instruksi

✅ <b>Solusi:</b>
1️⃣ Cari ruangan dengan pencahayaan cukup
2️⃣ Hapus kacamata, topi, masker
3️⃣ Posisi wajah di tengah frame
4️⃣ Jangan gunakan filter atau efek
5️⃣ Ikuti instruksi (blink, senyum, angkat alis)
6️⃣ Pastikan izin kamera aktif

📌 <b>Tips:</b> 
• Lakukan di tempat tenang
• Gunakan kamera belakang (kualitas lebih baik)
• Pastikan tidak ada bayangan di wajah"""
    },
    
    "kpj tidak ditemukan": {
        "keywords": ["kpj", "nomor kpj", "kpj tidak ditemukan", "kpj tidak valid"],
        "solution": """🔎 <b>MASALAH NOMOR KPJ JMO</b>

❌ <b>Penyebab:</b>
• Nomor KPJ yang dimasukkan salah
• Kepesertaan tidak aktif
• Data kepesertaan belum sinkron
• Peserta baru (belum terdaftar sistem)

✅ <b>Solusi:</b>
1️⃣ Periksa kembali nomor KPJ di kartu
2️⃣ Cek status kepesertaan di aplikasi
3️⃣ Hubungi HRD untuk verifikasi
4️⃣ Tunggu 1-2 hari kerja jika baru daftar
5️⃣ Kunjungi kantor BPJS untuk update data

📌 <b>Tips:</b> 
• Foto kartu kepesertaan untuk cadangan
• Simpan nomor KPJ di catatan HP"""
    },
    
    "saldo jht tidak muncul": {
        "keywords": ["saldo", "jht tidak muncul", "saldo jht", "cek saldo", "jht"],
        "solution": """🔎 <b>SALDO JHT TIDAK MUNCUL</b>

❌ <b>Penyebab:</b>
• Data tidak sinkron dengan server
• Kepesertaan tidak aktif
• Aplikasi versi lama
• Proses refresh gagal

✅ <b>Solusi:</b>
1️⃣ Tarik layar ke bawah untuk refresh
2️⃣ Login ulang ke aplikasi
3️⃣ Update aplikasi ke versi terbaru
4️⃣ Cek di website BPJS (bukan hanya aplikasi)
5️⃣ Tunggu 1-2 jam dan coba lagi

📌 <b>Tips:</b> 
• Saldo JHT biasanya update setiap bulan
• Cek secara berkala (tanggal 1-5 setiap bulan)
• Simpan bukti potongan gaji untuk verifikasi"""
    },
    
    "otp": {
        "keywords": ["otp", "kode otp", "verifikasi otp", "sms otp"],
        "solution": """🔎 <b>MASALAH OTP JMO</b>

❌ <b>Penyebab:</b>
• Nomor HP tidak aktif
• Sinyal buruk atau tidak ada
• SMS terblokir
• Provider mengalami gangguan
• Request OTP terlalu sering

✅ <b>Solusi:</b>
1️⃣ Pastikan sinyal HP bagus
2️⃣ Cek folder spam/blocked SMS
3️⃣ Tunggu 1-2 menit, jangan spam request
4️⃣ Restart HP dan coba lagi
5️⃣ Gunakan nomor HP lain jika masih gagal

📌 <b>Tips:</b> 
• Minta OTP di waktu sinyal stabil
• Jangan berbagi kode OTP dengan siapapun
• OTP berlaku 3 menit"""
    },
    
    "email": {
        "keywords": ["email", "ubah email", "ganti email", "verifikasi email"],
        "solution": """🔎 <b>MASALAH EMAIL JMO</b>

❌ <b>Penyebab:</b>
• Email tidak aktif
• Folder spam penuh
• Email sudah digunakan
• Link verifikasi kadaluarsa

✅ <b>Solusi:</b>
1️⃣ Cek folder SPAM/Trash
2️⃣ Pastikan email aktif dan bisa diakses
3️⃣ Gunakan email yang selalu dipakai
4️⃣ Minta kirim ulang email verifikasi
5️⃣ Hubungi admin jika email tidak terkirim

📌 <b>Tips:</b> 
• Gunakan Gmail untuk hasil terbaik
• Tandai email JMO sebagai "Important"
• Jangan gunakan email kantor"""
    },
    
    "aktivasi": {
        "keywords": ["aktivasi", "belum terdaftar", "registrasi", "daftar", "register"],
        "solution": """🔎 <b>MASALAH AKTIVASI/REGISTRASI JMO</b>

❌ <b>Penyebab:</b>
• Data kepesertaan belum terdaftar
• Status kepesertaan non-aktif
• Sistem maintenance
• Proses registrasi belum lengkap

✅ <b>Solusi:</b>
1️⃣ Pastikan data kepesertaan sudah terdaftar
2️⃣ Cek status kepesertaan di BPJS
3️⃣ Tunggu 1-2 hari kerja setelah pendaftaran
4️⃣ Lengkapi semua data yang diminta
5️⃣ Hubungi HRD jika masih terkendala

📌 <b>Tips:</b> 
• Siapkan KTP dan KK saat registrasi
• Data harus SAMA PERSIS dengan dokumen"""
    },
    
    "kartu digital": {
        "keywords": ["kartu digital", "kartu kepesertaan", "e-card", "virtual card"],
        "solution": """🔎 <b>MASALAH KARTU DIGITAL JMO</b>

❌ <b>Penyebab:</b>
• Data tidak sinkron
• Aplikasi versi lama
• Tidak ada akses internet
• Server BPJS bermasalah

✅ <b>Solusi:</b>
1️⃣ Refresh/update aplikasi
2️⃣ Login ulang ke aplikasi
3️⃣ Update ke versi terbaru
4️⃣ Coba di jam non-sibuk
5️⃣ Screenshot kartu digital untuk cadangan

📌 <b>Tips:</b> 
• Download kartu digital saat jaringan stabil
• Simpan di galeri HP untuk akses cepat
• Bawa kartu fisik sebagai cadangan"""
    },
    
    "klaim": {
        "keywords": ["klaim", "pengajuan klaim", "klaim jht", "cairkan jht"],
        "solution": """🔎 <b>MASALAH PENGAJUAN KLAIM JHT</b>

❌ <b>Penyebab:</b>
• Data tidak lengkap
• Dokumen tidak sesuai
• Kepesertaan tidak aktif
• Belum memenuhi syarat

✅ <b>Solusi:</b>
1️⃣ Periksa syarat dan ketentuan klaim
2️⃣ Lengkapi semua dokumen yang diminta
3️⃣ Pastikan kepesertaan aktif
4️⃣ Periksa masa kerja minimal (5 tahun)
5️⃣ Baca pesan error dengan teliti

📌 <b>Tips:</b> 
• Siapkan dokumen: KTP, KK, buku rekening
• Klaim bisa dilakukan online/offline
• Proses klaim 7-14 hari kerja"""
    },
    
    "bpu": {
        "keywords": ["bpu", "bukan penerima upah", "bpumandiri"],
        "solution": """🔎 <b>MASALAH KEPESERTAAN BPU</b>

❌ <b>Penyebab:</b>
• Data kepesertaan tidak sesuai
• Status non-aktif
• Pembayaran iuran tertunda
• Data sinkronisasi gagal

✅ <b>Solusi:</b>
1️⃣ Cek status kepesertaan di aplikasi
2️⃣ Pastikan pembayaran iuran lancar
3️⃣ Update data jika ada perubahan
4️⃣ Hubungi BPJS untuk verifikasi
5️⃣ Bawa KTP dan bukti pembayaran

📌 <b>Tips:</b> 
• BPU = Bukan Penerima Upah (mandiri)
• Bayar iuran tepat waktu
• Update data jika pindah alamat"""
    },
    
    "data tidak sesuai": {
        "keywords": ["data tidak sesuai", "identitas tidak sesuai", "nik", "nama", "ttl"],
        "solution": """🔎 <b>MASALAH DATA IDENTITAS TIDAK SESUAI</b>

❌ <b>Penyebab:</b>
• Data di JMO berbeda dengan KTP/KK
• NIK tidak terdaftar di Dukcapil
• Perubahan data belum sinkron
• Kesalahan input saat registrasi

✅ <b>Solusi:</b>
1️⃣ Periksa data di KTP dan KK
2️⃣ Input data SAMA PERSIS dengan dokumen
3️⃣ Hubungi Dukcapil jika NIK bermasalah
4️⃣ Update data ke BPJS terdekat
5️⃣ Tunggu proses sinkronisasi (1-2 hari)

📌 <b>Tips:</b> 
• Bawa fotokopi KTP/KK ke kantor BPJS
• Pastikan e-KTP sudah aktif
• Data harus match dengan database kependudukan"""
    },
    
    "kamera": {
        "keywords": ["kamera", "izin kamera", "akses kamera", "permission camera"],
        "solution": """🔎 <b>MASALAH KAMERA JMO</b>

❌ <b>Penyebab:</b>
• Izin kamera tidak aktif
• Aplikasi tidak memiliki akses kamera
• Kamera HP rusak/tertutup
• HP kotor atau lemot

✅ <b>Solusi:</b>
1️⃣ Aktifkan izin kamera di pengaturan HP
2️⃣ Restart aplikasi JMO
3️⃣ Bersihkan lensa kamera
4️⃣ Coba gunakan kamera depan
5️⃣ Update aplikasi JMO

📌 <b>Tips:</b> 
• Cek di Settings > Apps > JMO > Permissions
• Izinkan akses kamera secara permanen
• Gunakan HP dengan kamera yang bagus"""
    },
    
    "server": {
        "keywords": ["server", "gangguan", "maintenance", "tidak dapat terhubung", "connection"],
        "solution": """🔎 <b>MASALAH KONEKSI/SERVER JMO</b>

❌ <b>Penyebab:</b>
• Server BPJS sedang maintenance
• Gangguan jaringan internet
• Aplikasi tidak update
• Traffic pengguna tinggi

✅ <b>Solusi:</b>
1️⃣ Cek koneksi internet
2️⃣ Tunggu 15-30 menit
3️⃣ Coba di jam non-sibuk
4️⃣ Update aplikasi ke versi terbaru
5️⃣ Coba via website jika aplikasi error

📌 <b>Tips:</b> 
• Biasanya maintenance jam 00:00-04:00
• Cek akun media sosial JMO untuk info maintenance
• Gunakan WiFi stabil untuk hasil terbaik"""
    },
    
    "lemot": {
        "keywords": ["lemot", "lambat", "loading", "error loading", "blank"],
        "solution": """🔎 <b>MASALAH APLIKASI LEMOT/BLANK</b>

❌ <b>Penyebab:</b>
• Cache aplikasi penuh
• RAM HP penuh
• Aplikasi versi lama
• Koneksi internet lemot

✅ <b>Solusi:</b>
1️⃣ Clear cache aplikasi JMO
2️⃣ Restart HP
3️⃣ Tutup aplikasi lain yang berjalan
4️⃣ Update aplikasi ke versi terbaru
5️⃣ Gunakan WiFi 4G/5G

📌 <b>Tips:</b> 
• Hapus aplikasi yang jarang dipakai
• Aktifkan mode hemat data jika perlu
• Reinstall jika masih lemot"""
    },
    
    "notifikasi": {
        "keywords": ["notifikasi", "push notif", "pemberitahuan", "alert"],
        "solution": """🔎 <b>MASALAH NOTIFIKASI JMO</b>

❌ <b>Penyebab:</b>
• Izin notifikasi tidak aktif
• Aplikasi di mode silent
• HP mode Do Not Disturb
• Notifikasi diblock di pengaturan

✅ <b>Solusi:</b>
1️⃣ Aktifkan notifikasi di pengaturan HP
2️⃣ Cek setting notifikasi di aplikasi
3️⃣ Matikan mode Do Not Disturb
4️⃣ Allow notifikasi di Settings > Apps > JMO

📌 <b>Tips:</b> 
• Aktifkan notifikasi untuk info penting
• Cek secara manual jika notifikasi tidak muncul
• Allow semua izin notifikasi"""
    },
    
    "lupa email": {
        "keywords": ["lupa email", "email lupa", "tidak ingat email"],
        "solution": """🔎 <b>LUPA EMAIL JMO</b>

❌ <b>Penyebab:</b>
• Lupa email yang digunakan
• Email tidak aktif
• Email kantor sudah tidak digunakan

✅ <b>Solusi:</b>
1️⃣ Coba semua email yang mungkin
2️⃣ Gunakan fitur "Lupa Email" jika ada
3️⃣ Hubungi admin JMO
4️⃣ Gunakan email personal (bukan kantor)
5️⃣ Buat akun baru dengan email aktif

📌 <b>Tips:</b> 
• Gunakan email yang selalu diakses
• Catat email di tempat aman
• Gunakan 1 email khusus untuk JMO"""
    },
    
    "lupa nomor hp": {
        "keywords": ["lupa nomor hp", "nomor hp lupa", "ganti nomor"],
        "solution": """🔎 <b>LUPA/GANTI NOMOR HP JMO</b>

❌ <b>Penyebab:</b>
• Ganti nomor HP tanpa update
• Nomor HP tidak aktif
• Lupa nomor yang digunakan

✅ <b>Solusi:</b>
1️⃣ Update nomor HP di pengaturan akun
2️⃣ Hubungi admin untuk verifikasi
3️⃣ Bawa KTP ke kantor BPJS
4️⃣ Gunakan email untuk reset jika lupa
5️⃣ Pastikan nomor baru aktif

📌 <b>Tips:</b> 
• Update nomor segera jika ganti HP
• Gunakan nomor yang selalu aktif
• Simpan nomor cadangan jika perlu"""
    },
    
    "cara daftar": {
        "keywords": ["cara daftar", "pendaftaran", "registrasi", "daftar jmo"],
        "solution": """🔎 <b>CARA DAFTAR JMO</b>

📝 <b>Langkah-langkah:</b>

1️⃣ Download aplikasi JMO di Play Store/App Store
2️⃣ Buka aplikasi dan pilih "Daftar"
3️⃣ Input data:
   • NIK (KTP)
   • Nama lengkap
   • Tanggal lahir
   • Nomor HP aktif
   • Email aktif
4️⃣ Verifikasi OTP via SMS
5️⃣ Verifikasi email
6️⃣ Buat password
7️⃣ Login dan lengkapi data

📌 <b>Persyaratan:</b>
• WNI dengan e-KTP
• Memiliki kepesertaan BPJS Ketenagakerjaan
• Nomor HP aktif
• Email aktif

💡 <b>Tips:</b> 
• Siapkan KTP/KK sebelum daftar
• Gunakan jaringan stabil
• Data harus SAMA PERSIS dengan KTP"""
    },
    
    "jht cair": {
        "keywords": ["jht cair", "cairkan jht", "pencairan jht", "jht 2025"],
        "solution": """🔎 <b>PENCAIRAN JHT 2025</b>

📌 <b>Syarat Pencairan JHT:</b>
1️⃣ Peserta sudah berhenti bekerja (resign/di-PHK)
2️⃣ Masa kepesertaan minimal 5 tahun (156 bulan)
3️⃣ Status kepesertaan non-aktif
4️⃣ Mengisi form pengajuan klaim JHT

📝 <b>Dokumen yang dibutuhkan:</b>
• KTP asli dan fotokopi
• KK asli dan fotokopi
• Kartu kepesertaan BPJS
• Surat PHK/resign dari perusahaan
• Buku rekening bank aktif

💡 <b>Proses:</b>
• Online: melalui aplikasi JMO atau website BPJS
• Offline: datang ke kantor BPJS terdekat
• Waktu proses: 7-14 hari kerja

📌 <b>Tips:</b> 
• Pastikan semua dokumen lengkap
• Cek saldo JHT terlebih dahulu di aplikasi
• Ajukan klaim segera setelah resign
• Simpan bukti pengajuan klaim"""
    }
,

    "OTP_LENGKAP": {"keywords": ["otp","kode otp","otp tidak masuk"], "solution": """🔎 <b>MASALAH OTP JMO VERSI LENGKAP</b>

✅ Cek pulsa, sinyal, no HP aktif, tunggu 60 detik, cek Spam, restart HP, ganti jaringan"""},
    "AMALGAMASI_LENGKAP": {"keywords": ["amalgamasi","gabung kpj","kpj ganda"], "solution": """🔎 <b>AMALGAMASI - GABUNG KPJ GANDA VERSI LENGKAP</b>

1. Pengkinian Data dulu 2. Menu Penggabungan Saldo 3. Masukkan KPJ lama 4. Syarat Nama & NIK sama persis 5. Proses 3x24 jam kerja"""},
    "REKENING_LENGKAP": {"keywords": ["rekening","buku tabungan"], "solution": """🔎 <b>REKENING GAGAL - LENGKAP</b>

✅ Rekening atas nama sendiri, bukan e-wallet, foto buku tabungan halaman 1 jelas"""},
    "KARTU_LENGKAP": {"keywords": ["kartu","kpj hilang"], "solution": """🔎 <b>KPJ HILANG / LUPA - LENGKAP</b>

✅ Tanya HRD, cek di JMO, atau datang ke cabang bawa KTP"""},

}



def get_jmo_solution(text: str) -> str:
    """Mencari solusi JMO berdasarkan keyword yang cocok"""
    text_lower = text.lower()
    
    # Cari kecocokan di database
    for key, data in JMO_SOLUTIONS.items():
        for keyword in data["keywords"]:
            if keyword in text_lower:
                return data["solution"]
    
    # Jika tidak ada yang cocok, berikan respon default dengan saran
    return """🛠️ <b>BELUM DITEMUKAN SOLUSI KHUSUS</b>

Maaf, saya belum menemukan solusi yang tepat untuk masalah Anda.

📌 <b>Untuk mendapatkan solusi yang lebih akurat, silakan:</b>

1️⃣ Tulis ulang masalah dengan lebih DETAIL
2️⃣ Sertakan KODE ERROR yang muncul (contoh: 025, 026, dst)
3️⃣ Jelaskan TAHAPAN yang gagal (login, verifikasi, dll)
4️⃣ Sebutkan PESAN ERROR lengkapnya
5️⃣ Foto/screenshot ERROR bisa diupload

💡 <b>Contoh pertanyaan yang baik:</b>
• "Kode 025 saat login JMO, data sudah sesuai"
• "JMO error 026, nomor KPJ tidak ditemukan"
• "Verifikasi wajah gagal terus, sudah dicoba berkali-kali"

📞 <b>Atau hubungi langsung:</b>
• Admin: @Hambali1995
• WhatsApp: 083160776091
• Kantor BPJS terdekat

⚠️ Jangan lupa screenshot error untuk dibawa ke kantor BPJS!"""


# ==================== COMMAND START ====================

@dp.message(CommandStart())
async def start(message: Message, state: FSMContext):
    await state.clear()
    register_user(message.from_user)
    # FIX: Hapus ReplyKeyboard kotak abu-abu yang lama
    try:
        await message.answer("Menghapus menu lama...", reply_markup=ReplyKeyboardRemove())
    except Exception:
        pass
    await message.answer(
        "🤖 <b>SAHABAT JHT 🤖</b>\n\n"
        f"👋 Selamat datang, <b>{message.from_user.full_name}</b>!\n"
        "Gimana kabarnya nih, saya berharap kabar baik-baik saja yah, "
        "tetap semangat dan jangan lupa bersyukur.\n"
        "Silahkan pilih menu di bawah ini : 👇",
        reply_markup=main_menu(message.from_user.id)
    )



@dp.message(Command("profil"))
async def cmd_profil(message: Message, state: FSMContext):
    await state.clear()
    register_user(message.from_user)
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (message.from_user.id,)).fetchone()
    sub = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (message.from_user.id,)).fetchone()
    conn.close()
    username = f"@{user['username']}" if user["username"] else "-"
    package = sub["package_name"] if sub else "Belum ada"
    expiry = "-"
    status = "🔴 Tidak aktif"
    if sub:
        if sub["status"] == "unlimited":
            expiry, status = "SELAMANYA", "🟢 AKTIF"
        elif sub["expiry_date"]:
            from datetime import datetime as dt
            exp = dt.fromisoformat(sub["expiry_date"])
            expiry = exp.strftime("%d-%m-%Y")
            status = "🟢 AKTIF" if dt.now() < exp else "🔴 EXPIRED"
    def rupiah2(v): return f"Rp {int(v):,}".replace(",", ".")
    await message.answer(
        "👤 <b>PROFIL USER</b>\n\n"
        f"🆔 Telegram ID : <code>{user['telegram_id']}</code>\n"
        f"👤 Nama : {user['name']}\n"
        f"📱 Username : {username}\n"
        f"💰 Saldo : <b>{rupiah2(user['balance'])}</b>\n\n"
        f"📦 Langganan : {package}\n"
        f"📅 Berakhir : {expiry}\n"
        f"📊 Status : {status}",
        reply_markup=main_menu(message.from_user.id)
    )

@dp.message(Command("status"))
async def cmd_status(message: Message, state: FSMContext):
    await state.clear()
    sub = get_subscription(message.from_user.id)
    if not sub:
        txt = "📊 <b>CEK STATUS LAYANAN</b>\n\n🔴 Belum ada langganan aktif.\nSilakan beli paket di menu AUTO FORMAT."
    else:
        expiry = sub["expiry_date"] if sub["expiry_date"] else "SELAMANYA"
        txt = f"📊 <b>CEK STATUS LAYANAN</b>\n\n📦 Paket : {sub['package_name']}\n📅 Berakhir : {expiry}\n📊 Status : {sub['status'].upper()}"
    await message.answer(txt, reply_markup=main_menu(message.from_user.id))

@dp.message(Command("autoformat"))
@dp.message(Command("auto_format"))
async def cmd_autoformat(message: Message, state: FSMContext):
    await state.clear()
    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
    if has_auto_format_access(message.from_user.id):
        await message.answer(
            "📝 <b>AUTO FORMAT</b>\n\n🔓 Akses kamu aktif.\nSilakan pilih menu:",
            reply_markup=auto_menu()
        )
    else:
        await message.answer(
            "🔒 <b>AUTO FORMAT TERKUNCI</b>\n\nUntuk membuka AUTO FORMAT, silakan pilih paket:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🟢 PAKET 6 BULAN — Rp50.000", callback_data="af_6m")],
                [InlineKeyboardButton(text="🔵 PAKET 1 TAHUN — Rp80.000", callback_data="af_1y")],
                [InlineKeyboardButton(text="🟣 PAKET UNLIMITED — Rp200.000", callback_data="af_unlimited")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )

@dp.message(Command("topup"))
async def cmd_topup(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(PaymentState.waiting_topup_amount)
    await message.answer(
        "💳 <b>TOP UP SALDO</b>\n\n"
        "Silakan transfer ke:\n\n"
        "🏦 <b>SEABANK</b>\n901040978290\nA/N HAMBALI\n\n"
        "💰 <b>DANA</b>\n083824101264\nA/N HAMBALI\n\n"
        "Ketik nominal. Contoh : <code>50000</code>",
        reply_markup=back_main()
    )

@dp.message(Command("solusijmo"))
@dp.message(Command("jmo"))
async def cmd_jmo(message: Message):
    await message.answer(
        "🛠️ <b>SOLUSI MASALAH JMO</b>\n\nKetik masalah kamu, contoh:\n<code>Kode 025 saat login</code>\n<code>JMO error 026</code>\n<code>Verifikasi wajah gagal</code>",
        reply_markup=back_main()
    )

@dp.message(Command("bantuan"))
@dp.message(Command("admin"))
async def cmd_bantuan(message: Message):
    await message.answer(
        "📞 <b>HUBUNGI ADMIN</b>\n\n👤 Admin : @Hambali1995\n📱 WhatsApp : 083160776091",
        reply_markup=back_main()
    )

@dp.message(Command("info"))
async def cmd_info(message: Message):
    await message.answer(
        "ℹ️ <b>INFO BOT SABABAT JHT</b>\n\n🤖 Bot bantuan JHT, Auto Format, Top Up, dan Solusi JMO.",
        reply_markup=back_main()
    )


@dp.callback_query(F.data == "back_main")
async def back_main_handler(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "🤖 <b>SABABAT JHT 🤖</b>\n\n"
        f"👋 Selamat datang, <b>{callback.from_user.full_name}</b>!\n"
        "Gimana kabarnya nih, saya berharap kabar baik-baik saja yah, "
        "tetap semangat dan jangan lupa bersyukur.\n"
        "Silahkan pilih menu di bawah ini : 👇",
        reply_markup=main_menu(callback.from_user.id)
    )
    await callback.answer()


@dp.callback_query(F.data == "info_bot")
async def info_bot(callback: CallbackQuery):
    await callback.message.edit_text(
        "🤖 <b>SABABAT JHT</b>\n\n"
        "Bot bantuan JHT, Auto Format, Top Up, dan Solusi JMO.\n"
        "Gunakan menu utama untuk melanjutkan.",
        reply_markup=back_main()
    )
    await callback.answer()


@dp.callback_query(F.data == "profile")
async def profile(callback: CallbackQuery):
    register_user(callback.from_user)
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
    sub = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
    conn.close()

    username = f"@{user['username']}" if user["username"] else "-"
    package = sub["package_name"] if sub else "Belum ada"
    expiry = "-"
    status = "🔴 Tidak aktif"

    if sub:
        if sub["status"] == "unlimited":
            expiry, status = "SELAMANYA", "🟢 AKTIF"
        elif sub["expiry_date"]:
            exp = datetime.fromisoformat(sub["expiry_date"])
            expiry = exp.strftime("%d-%m-%Y")
            status = "🟢 AKTIF" if datetime.now() < exp else "🔴 EXPIRED"

    await callback.message.edit_text(
        "👤 <b>PROFIL USER</b>\n\n"
        f"🆔 Telegram ID : <code>{user['telegram_id']}</code>\n"
        f"👤 Nama : {user['name']}\n"
        f"📱 Username : {username}\n"
        f"💰 Saldo : <b>{rupiah(user['balance'])}</b>\n\n"
        f"📦 Langganan : {package}\n"
        f"📅 Berakhir : {expiry}\n"
        f"📊 Status : {status}",
        reply_markup=back_main()
    )
    await callback.answer()


@dp.callback_query(F.data == "topup")
async def topup(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(PaymentState.waiting_topup_amount)
    await callback.message.edit_text(
        "💳 <b>TOP UP SALDO</b>\n\n"
        "Silakan transfer ke:\n\n"
        "🏦 <b>SEABANK</b>\n"
        "901040978290\n"
        "A/N HAMBALI\n\n"
        "💰 <b>DANA</b>\n"
        "083824101264\n"
        "A/N HAMBALI\n\n"
        "Atau bisa ketik nominal.\n"
        "Contoh : <code>50000</code>\n\n"
        "Setelah membayar, ketik nominal yang dibayar lalu tekan <b>SUDAH BAYAR</b>.",
        reply_markup=back_main()
    )
    await callback.answer()


@dp.message(PaymentState.waiting_topup_amount, F.text)
async def topup_amount(message: Message, state: FSMContext):
    raw = message.text.strip().lower().replace("rp", "").replace(".", "").replace(",", "").replace(" ", "")
    if not raw.isdigit() or int(raw) <= 0:
        await message.answer("❌ Nominal tidak valid. Contoh: <code>50000</code>")
        return

    amount = int(raw)
    await state.update_data(amount=amount, package_code=None, package_name=None)
    await state.set_state(PaymentState.waiting_proof)
    await message.answer(
        "✅ <b>NOMINAL TERCATAT</b>\n\n"
        f"💰 Nominal : <b>{rupiah(amount)}</b>\n\n"
        "Setelah transfer, tekan tombol:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ SUDAH BAYAR", callback_data="payment_done")],
            [InlineKeyboardButton(text="❌ BATAL", callback_data="back_main")]
        ])
    )


AUTO_PACKAGES = {
    "af_6m": ("6M", "AUTO FORMAT 6 BULAN", 50000),
    "af_1y": ("1Y", "AUTO FORMAT 1 TAHUN", 80000),
    "af_unlimited": ("UNLIMITED", "AUTO FORMAT UNLIMITED", 200000)
}


@dp.callback_query(F.data == "auto_format")
async def auto_format(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    if has_auto_format_access(callback.from_user.id):
        await callback.message.edit_text(
            "📝 <b>AUTO FORMAT</b>\n\n🔓 Akses kamu aktif.\n\nSilakan pilih menu:",
            reply_markup=auto_menu()
        )
    else:
        await callback.message.edit_text(
            "🔒 <b>AUTO FORMAT TERKUNCI</b>\n\n"
            "Untuk membuka AUTO FORMAT, silakan pilih paket:\n\n"
            "🟢 6 Bulan — <b>Rp50.000</b>\n"
            "🔵 1 Tahun — <b>Rp80.000</b>\n"
            "🟣 Unlimited — <b>Rp200.000</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🟢 PAKET 6 BULAN — Rp50.000", callback_data="af_6m")],
                [InlineKeyboardButton(text="🔵 PAKET 1 TAHUN — Rp80.000", callback_data="af_1y")],
                [InlineKeyboardButton(text="🟣 PAKET UNLIMITED — Rp200.000", callback_data="af_unlimited")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )
    await callback.answer()


@dp.callback_query(F.data.in_(set(AUTO_PACKAGES.keys())))
async def auto_package(callback: CallbackQuery, state: FSMContext):
    code, name, price = AUTO_PACKAGES[callback.data]
    await state.update_data(amount=price, package_code=code, package_name=name)
    await state.set_state(PaymentState.waiting_proof)

    await callback.message.edit_text(
        "💳 <b>PEMBAYARAN AUTO FORMAT</b>\n\n"
        f"📦 Paket : <b>{name}</b>\n"
        f"💰 Harga : <b>{rupiah(price)}</b>\n\n"
        "Silakan transfer ke:\n\n"
        "🏦 <b>SEABANK</b>\n901040978290\nA/N HAMBALI\n\n"
        "💰 <b>DANA</b>\n083824101264\nA/N HAMBALI\n\n"
        "Setelah membayar, tekan <b>SUDAH BAYAR</b> lalu upload bukti.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ SUDAH BAYAR", callback_data="payment_done")],
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
        ])
    )
    await callback.answer()


@dp.callback_query(F.data == "payment_done")
async def payment_done(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if not data.get("amount"):
        await callback.answer("❌ Nominal belum ditentukan.", show_alert=True)
        return
    await state.set_state(PaymentState.waiting_proof)
    await callback.message.edit_text(
        "📸 <b>UPLOAD BUKTI PEMBAYARAN</b>\n\n"
        f"💰 Nominal : <b>{rupiah(data['amount'])}</b>\n"
        f"📦 Paket : {data.get('package_name') or 'TOP UP SALDO'}\n\n"
        "Silakan kirim FOTO bukti transfer di chat ini."
    )
    await callback.answer()


@dp.message(PaymentState.waiting_proof, F.photo)
async def payment_proof(message: Message, state: FSMContext):
    data = await state.get_data()
    amount = int(data.get("amount", 0))
    if amount <= 0:
        await state.clear()
        await message.answer("❌ Nominal pembayaran tidak ditemukan.")
        return

    photo = message.photo[-1]
    conn = db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO transactions(
            telegram_id,amount,payment_method,package_code,package_name,
            proof_file_id,status,created_at
        ) VALUES(?,?,'SEABANK/DANA',?,?,?,'pending',?)
    """, (
        message.from_user.id, amount, data.get("package_code"),
        data.get("package_name"), photo.file_id, datetime.now().isoformat()
    ))
    tx_id = cur.lastrowid
    conn.commit()
    conn.close()
    await state.clear()

    await message.answer(
        "✅ <b>BUKTI PEMBAYARAN DITERIMA</b>\n\n"
        f"🧾 Transaksi : #{tx_id}\n"
        f"💰 Nominal : {rupiah(amount)}\n"
        f"📦 Paket : {data.get('package_name') or 'TOP UP SALDO'}\n\n"
        "⏳ Menunggu konfirmasi Admin."
    )

    caption = (
        "💰 <b>PEMBAYARAN BARU</b>\n\n"
        f"🧾 Transaksi : #{tx_id}\n"
        f"👤 Nama : {message.from_user.full_name}\n"
        f"🆔 ID : <code>{message.from_user.id}</code>\n"
        f"📱 Username : @{message.from_user.username or '-'}\n"
        f"💰 Nominal : <b>{rupiah(amount)}</b>\n"
        f"📦 Paket : {data.get('package_name') or 'TOP UP SALDO'}\n"
        "🟡 Status : PENDING"
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ SETUJU", callback_data=f"approve_{tx_id}"),
        InlineKeyboardButton(text="❌ TOLAK", callback_data=f"reject_{tx_id}")
    ]])

    for admin_id in ADMIN_IDS:
        try:
            await message.bot.send_photo(admin_id, photo.file_id, caption=caption, reply_markup=keyboard)
        except Exception:
            logging.exception("Gagal mengirim transaksi ke admin")


@dp.message(PaymentState.waiting_proof)
async def payment_wrong_proof(message: Message):
    await message.answer("📸 Silakan kirim FOTO bukti pembayaran.")


async def process_payment(callback: CallbackQuery, tx_id: int, approve: bool):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Kamu bukan Admin.", show_alert=True)
        return

    conn = db()
    tx = conn.execute("SELECT * FROM transactions WHERE id=?", (tx_id,)).fetchone()

    if not tx or tx["status"] != "pending":
        conn.close()
        await callback.answer("⚠️ Transaksi sudah diproses/tidak ditemukan.", show_alert=True)
        return

    now = datetime.now().isoformat()

    # Cek jika ini paket KOTA, hitung expiry berdasarkan days
    if tx["package_code"] and (tx["package_code"].startswith("KOTA_") or tx["package_code"].startswith("CARI_")):
        # ambil days dari KOTA_PACKAGES
        days_map = {"KOTA_1W": 7, "KOTA_1M": 30, "KOTA_2M": 60, "KOTA_6M": 180, "KOTA_UNLIMITED": None, "CARI_1W": 7, "CARI_1M": 30, "CARI_2M": 60}
        d = days_map.get(tx["package_code"])
        if d is None:
            expiry, sub_status = None, "unlimited"
        else:
            expiry, sub_status = datetime.now() + __import__('datetime').timedelta(days=d), "active"
    else:
        expiry = None

    if approve:
        conn.execute(
            "UPDATE transactions SET status='approved', processed_at=? WHERE id=?",
            (now, tx_id)
        )

        # expiry already set for KOTA above, handle non-KOTA
        if not tx["package_code"] or not tx["package_code"].startswith("KOTA_"):
            expiry = None
        if tx["package_code"] and not tx["package_code"].startswith("KOTA_"):
            if tx["package_code"] == "6M":
                expiry, sub_status = add_months(datetime.now(), 6), "active"
            elif tx["package_code"] == "1Y":
                expiry, sub_status = add_months(datetime.now(), 12), "active"
            else:
                sub_status = "unlimited"

            conn.execute("""
                INSERT INTO subscriptions(
                    telegram_id,package_code,package_name,price,start_date,expiry_date,status
                ) VALUES(?,?,?,?,?,?,?)
                ON CONFLICT(telegram_id) DO UPDATE SET
                    package_code=excluded.package_code,
                    package_name=excluded.package_name,
                    price=excluded.price,
                    start_date=excluded.start_date,
                    expiry_date=excluded.expiry_date,
                    status=excluded.status
            """, (
                tx["telegram_id"], tx["package_code"], tx["package_name"], tx["amount"],
                now, expiry.isoformat() if expiry else None, sub_status
            ))
        else:
            conn.execute(
                "UPDATE users SET balance=balance+? WHERE telegram_id=?",
                (tx["amount"], tx["telegram_id"])
            )

        conn.commit()
        conn.close()

        expiry_text = ""
        if tx["package_code"]:
            expiry_text = (
                f"\n📅 Berakhir : "
                f"{expiry.strftime('%d-%m-%Y') if expiry else 'SELAMANYA'}"
            )

        user_text = (
            "✅ <b>PEMBAYARAN DISETUJUI</b>\n\n"
            f"🧾 Transaksi : #{tx_id}\n"
            f"💰 Nominal : {rupiah(tx['amount'])}\n"
            f"📦 Paket : {tx['package_name'] or 'TOP UP SALDO'}"
            f"{expiry_text}\n\n"
            + (
                "🔓 AUTO FORMAT SEKARANG SUDAH TERBUKA."
                if tx["package_code"]
                else "💰 Saldo kamu sudah ditambahkan."
            )
        )
    else:
        conn.execute(
            "UPDATE transactions SET status='rejected', processed_at=? WHERE id=?",
            (now, tx_id)
        )
        conn.commit()
        conn.close()

        user_text = (
            "❌ <b>PEMBAYARAN DITOLAK</b>\n\n"
            f"🧾 Transaksi : #{tx_id}\n"
            f"💰 Nominal : {rupiah(tx['amount'])}\n"
            f"📦 Paket : {tx['package_name'] or 'TOP UP SALDO'}\n\n"
            "Silakan cek kembali bukti pembayaran atau hubungi Admin."
        )

    try:
        await callback.bot.send_message(tx["telegram_id"], user_text)
    except Exception:
        logging.exception("Gagal mengirim notifikasi pembayaran")

    try:
        if callback.message.photo:
            old_caption = callback.message.caption or ""
            label = "✅ DISETUJUI" if approve else "❌ DITOLAK"
            await callback.message.edit_caption(
                caption=old_caption + f"\n\n<b>{label}</b>",
                reply_markup=None
            )
        else:
            old_text = callback.message.text or ""
            label = "✅ DISETUJUI" if approve else "❌ DITOLAK"
            await callback.message.edit_text(
                old_text + f"\n\n<b>{label}</b>",
                reply_markup=None
            )
    except Exception:
        logging.exception("Gagal memperbarui pesan admin")

    await callback.answer("✅ Disetujui." if approve else "❌ Ditolak.")


@dp.callback_query(F.data.startswith("approve_"))
async def approve_payment(callback: CallbackQuery):
    try:
        tx_id = int(callback.data.split("_", 1)[1])
    except (ValueError, IndexError):
        await callback.answer("❌ ID transaksi tidak valid.", show_alert=True)
        return
    await process_payment(callback, tx_id, True)


@dp.callback_query(F.data.startswith("reject_"))
async def reject_payment(callback: CallbackQuery):
    try:
        tx_id = int(callback.data.split("_", 1)[1])
    except (ValueError, IndexError):
        await callback.answer("❌ ID transaksi tidak valid.", show_alert=True)
        return
    await process_payment(callback, tx_id, False)


@dp.callback_query(F.data == "status")
async def status(callback: CallbackQuery):
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
    sub = conn.execute("SELECT * FROM subscriptions WHERE telegram_id=?", (callback.from_user.id,)).fetchone()
    conn.close()

    balance = user["balance"] if user else 0
    if not sub:
        text = f"📊 <b>STATUS USER</b>\n\n💰 Saldo : <b>{rupiah(balance)}</b>\n📦 Langganan : Belum ada\n📊 Status : 🔴 Tidak aktif"
    elif sub["status"] == "unlimited":
        text = f"📊 <b>STATUS USER</b>\n\n💰 Saldo : <b>{rupiah(balance)}</b>\n📦 Langganan : {sub['package_name']}\n📅 Berakhir : SELAMANYA\n📊 Status : 🟢 AKTIF"
    else:
        exp = datetime.fromisoformat(sub["expiry_date"])
        active = datetime.now() < exp
        text = (
            f"📊 <b>STATUS USER</b>\n\n💰 Saldo : <b>{rupiah(balance)}</b>\n"
            f"📦 Langganan : {sub['package_name']}\n"
            f"📅 Berakhir : {exp.strftime('%d-%m-%Y')}\n"
            f"📊 Status : {'🟢 AKTIF' if active else '🔴 EXPIRED'}"
        )

    await callback.message.edit_text(text, reply_markup=back_main())
    await callback.answer()


# ==================== SOLUSI JMO ====================

@dp.callback_query(F.data == "solusi_jmo")
async def solusi_start(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(JmoState.waiting_question)
    await callback.message.edit_text(
        "🛠️ <b>SOLUSI JMO</b>\n\n"
        "Silakan masukkan masalah Anda di sini.\n\n"
        "📌 <b>Contoh:</b>\n"
        "• <code>cara atasi solusi 026</code>\n"
        "• <code>JMO error 025</code>\n"
        "• <code>Verifikasi wajah gagal</code>\n"
        "• <code>Login tidak bisa</code>\n\n"
        "💡 <b>Tips:</b> Sebutkan kode error jika ada!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❓ SOLUSI ERROR 025", callback_data="jmo_025")],
            [InlineKeyboardButton(text="❓ SOLUSI ERROR 026", callback_data="jmo_026")],
            [InlineKeyboardButton(text="❓ SOLUSI ERROR 029 (WAJAH)", callback_data="jmo_029")],
            [InlineKeyboardButton(text="❓ CARA CAIRKAN JHT", callback_data="jmo_jht cair")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()


# Quick access untuk error umum
@dp.callback_query(F.data.startswith("jmo_"))
async def jmo_quick(callback: CallbackQuery, state: FSMContext):
    error_code = callback.data.split("_", 1)[1]
    # Cari solusi untuk error code
    for key, data in JMO_SOLUTIONS.items():
        if key == error_code:
            await callback.message.edit_text(
                data["solution"],
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🔄 KEMBALI KE MENU SOLUSI", callback_data="solusi_jmo")],
                    [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
                ])
            )
            await callback.answer()
            return
    
    # Jika tidak ditemukan
    await callback.answer("Solusi tidak ditemukan", show_alert=True)


@dp.message(JmoState.waiting_question, F.text)
async def solusi_text(message: Message, state: FSMContext):
    # Cari solusi di database
    solution = get_jmo_solution(message.text)
    
    # Kirim solusi dengan menu kembali
    await message.answer(
        solution,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔄 TANYA LAGI", callback_data="solusi_jmo")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )


@dp.message(JmoState.waiting_question, F.photo)
async def solusi_photo(message: Message, state: FSMContext):
    caption = message.caption or ""
    if caption:
        solution = get_jmo_solution(caption)
        await message.answer(
            solution,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔄 TANYA LAGI", callback_data="solusi_jmo")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )
    else:
        await message.answer(
            "📸 <b>FOTO DITERIMA</b>\n\n"
            "Silakan tuliskan kode/error yang terlihat pada foto, misalnya:\n"
            "• <code>025</code>\n"
            "• <code>026</code>\n"
            "• <code>verifikasi wajah</code>\n\n"
            "Saya akan mencari solusi yang tepat!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔄 KETIK MASALAH", callback_data="solusi_jmo")],
                [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
            ])
        )


# ==================== AUTO FORMAT (OTOMATIS + KAPITAL) ====================

@dp.callback_query(F.data == "format_create")
async def format_create(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    await callback.message.edit_text(
        "📝 <b>BUAT FORMAT</b>\n\nPilih cara membuat format:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⌨️ KETIK DATA MANUAL", callback_data="format_manual")],
            [InlineKeyboardButton(text="📊 UPLOAD FILE EXCEL (.xlsx)", callback_data="format_excel")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )
    await callback.answer()


@dp.callback_query(F.data == "format_manual")
async def format_manual_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(FormatState.waiting_manual)
    await callback.message.edit_text(
        "⌨️ <b>BUAT FORMAT</b>\n\n"
        "Ketik data <b>tanpa perlu menulis KAB/KEC/KEL</b>. Bot otomatis membaca urutannya:\n\n"
        "1️⃣ KAB\n2️⃣ KEC\n3️⃣ KEL\n4️⃣ SALDO\n5️⃣ KELAMIN\n6️⃣ KPJ\n7️⃣ SENSOR\n8️⃣ IT\n9️⃣ PT\n\n"
        "Contoh input:\n<code>DEPOK\nCILODONG\nKALIBARU\n10.000.000\nPEREMPUAN 1992\n2019\n23****\n01-07-2022\nINDONESIA MERDEKA</code>\n\n"
        "Hasilnya otomatis menjadi format JMO.",
        reply_markup=back_main()
    )
    await callback.answer()




@dp.callback_query(F.data == "lanjut_bikin")
async def lanjut_bikin_handler(callback: CallbackQuery, state: FSMContext):
    await state.set_state(FormatState.waiting_manual)
    await callback.message.answer(
        "📝 <b>LANJUT BIKIN FORMAT</b>\n\nSilakan kirim data format baru di bawah:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )
    await callback.answer()



def make_template_from_example(example_text: str) -> str:
    import re
from aiohttp import web
    lines = example_text.splitlines()
    out=[]
    for line in lines:
        l=line
        if re.search(r'KAB\s*:', l, re.I):
            l=re.sub(r'(KAB\s*:).*', r'\1 {KAB}', l, flags=re.I)
        elif re.search(r'KEC\s*:', l, re.I):
            l=re.sub(r'(KEC\s*:).*', r'\1 {KEC}', l, flags=re.I)
        elif re.search(r'KEL\s*:', l, re.I):
            l=re.sub(r'(KEL\s*:).*', r'\1 {KEL}', l, flags=re.I)
        elif re.search(r'(TOTAL\s*JHT|SALDO)', l, re.I) and ':' in l:
            l=re.sub(r'(:).*', r'\1 {SALDO}', l)
        elif re.search(r'KELAMIN\s*:', l, re.I):
            l=re.sub(r'(KELAMIN\s*:).*', r'\1 {KELAMIN}', l, flags=re.I)
        elif re.search(r'KPJ\s+SENSOR', l, re.I) or (re.search(r'SENSOR\s*:', l, re.I) and 'KPJ' in l.upper()):
            l=re.sub(r'(:).*', r'\1 {SENSOR}', l)
        elif re.search(r'^[^:]*SENSOR\s*:', l, re.I):
            l=re.sub(r'(:).*', r'\1 {SENSOR}', l)
        elif re.search(r'(?<!SENSOR\s)KPJ\s*:', l, re.I):
            l=re.sub(r'(KPJ\s*:).*', r'\1 {KPJ}', l, flags=re.I)
        elif re.search(r'IURAN\s*T|\bIT\s*:', l, re.I) and ':' in l:
            l=re.sub(r'(:).*', r'\1 {IT}', l)
        elif re.search(r'\bPT\b', l, re.I):
            if ':' in l:
                l=re.sub(r'(:).*', r'\1 {PT}', l)
            else:
                m=re.search(r'(PT\s*\*?\s*)(.*)', l, re.I)
                if m:
                    l=m.group(1)+'{PT}'
        out.append(l)
    return "\n".join(out).strip()

def parse_data_with_akun(raw_text: str) -> dict:
    import re
from aiohttp import web
    data={}
    def get(pat):
        m=re.search(pat, raw_text, re.I | re.M)
        return m.group(1).strip() if m else ""
    data["KAB"]=get(r'KAB\s*:\s*([^\n]+)')
    data["KEC"]=get(r'KEC\s*:\s*([^\n]+)')
    data["KEL"]=get(r'KEL\s*:\s*([^\n]+)')
    m=re.search(r'(?:TOTAL\s*JHT|SALDO)[^:]*:\s*([0-9\.\,]+)', raw_text, re.I)
    data["SALDO"]=m.group(1).strip() if m else get(r'SALDO\s*:\s*([^\n]+)')
    data["KELAMIN"]=get(r'KELAMIN\s*:\s*([^\n]+)')
    kpjs=re.findall(r'KPJ\s*:\s*([^\n]+)', raw_text, re.I)
    data["KPJ"]=""
    for k in kpjs:
        if 'SENSOR' not in k.upper():
            if re.search(r'\d{4}', k):
                data["KPJ"]=k.strip()
                break
    if not data["KPJ"]:
        for k in kpjs:
            if 'SENSOR' not in k.upper():
                data["KPJ"]=k.strip()
                break
    data["SENSOR"]=get(r'(?:KPJ\s*SENSOR|SENSOR)\s*:\s*([^\n]+)')
    data["IT"]=get(r'(?:IURAN\s*T|IT)\s*:\s*([^\n]+)')
    m=re.search(r'PT\s*:?\s*\*?\s*([^\n]+)', raw_text, re.I)
    data["PT"]=m.group(1).strip() if m else ""
    email=""; password=""
    m=re.search(r'AKUN\s*:\s*\n+([^\n@\s]+@[^\n\s]+)\s*\n+([^\n]+)', raw_text, re.I)
    if m:
        email=m.group(1).strip(); password=m.group(2).strip()
    else:
        em=re.findall(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}', raw_text)
        if em:
            email=em[0]
            after=raw_text.split(email)[-1].splitlines()
            for line in after:
                line=line.strip()
                if line and '@' not in line and len(line)>=3 and 'SAHABAT' not in line.upper():
                    if not line.startswith('🎼') and not line.startswith('🇯🇵'):
                        password=line; break
    data["EMAIL"]=email; data["PASSWORD"]=password
    return data


def apply_template(template, raw_text, kode_header=None):
    akun_data = parse_data_with_akun(raw_text)
    lines = [l.strip() for l in str(raw_text).strip().splitlines() if l.strip()!=""]
    jht_lines=[]
    for l in lines:
        if '@' in l and '.' in l and not __import__('re').search(r'KAB|KEC|KEL|PT|KPJ|SALDO', l, __import__('re').I):
            continue
        if __import__('re').search(r'AKUN\s*:', l, __import__('re').I):
            break
        jht_lines.append(l)
    cleaned=jht_lines
    while len(cleaned)<9:
        cleaned.append("-")
    if len(cleaned)>9:
        cleaned=cleaned[:8]+[" ".join(cleaned[8:])]
    def up(x):
        x=x.strip()
        if x=="-" or x=="": return "-"
        return x.upper()
    data = {
        "KAB": up(akun_data.get("KAB") or (cleaned[0] if len(cleaned)>0 else "-")),
        "KEC": up(akun_data.get("KEC") or (cleaned[1] if len(cleaned)>1 else "-")),
        "KEL": up(akun_data.get("KEL") or (cleaned[2] if len(cleaned)>2 else "-")),
        "SALDO": up(akun_data.get("SALDO") or (cleaned[3] if len(cleaned)>3 else "-")),
        "KELAMIN": up(akun_data.get("KELAMIN") or (cleaned[4] if len(cleaned)>4 else "-")),
        "KPJ": up(akun_data.get("KPJ") or (cleaned[5] if len(cleaned)>5 else "-")),
        "SENSOR": up(akun_data.get("SENSOR") or (cleaned[6] if len(cleaned)>6 else "-")),
        "IT": up(akun_data.get("IT") or (cleaned[7] if len(cleaned)>7 else "-")),
        "PT": up(akun_data.get("PT") or (cleaned[8] if len(cleaned)>8 else "-")),
    }
    if "{" in template and "}" in template:
        result=template
        for k,v in data.items():
            result=result.replace("{"+k+"}", v)
            result=result.replace("{"+k.lower()+"}", v)
            result=result.replace("{"+k.title()+"}", v)
        import re
from aiohttp import web
        result=re.sub(r'\{[A-Z_]+\}', '', result)
        if kode_header:
            centered=kode_header.center(27)
            result=f"{centered}\n━━━━━━━━━━━━━━━━━━━\n{result}"
        return result
    templ_converted=make_template_from_example(template)
    if "{" in templ_converted:
        result=templ_converted
        for k,v in data.items():
            result=result.replace("{"+k+"}", v)
        import re
from aiohttp import web
        result=re.sub(r'\{[A-Z_]+\}', '', result)
        if kode_header:
            centered=kode_header.center(27)
            result=f"{centered}\n━━━━━━━━━━━━━━━━━━━\n{result}"
        return result
    out_lines=[]; pin_count=0
    import re
from aiohttp import web
    for line in template.splitlines():
        upper=line.upper()
        new_line=line
        if "📍" in line:
            pin_count+=1
            if pin_count==1: new_line=re.sub(r"📍\s*.*", f"📍 {data['KAB']}", line)
            elif pin_count==2: new_line=re.sub(r"📍\s*.*", f"📍 {data['KEC']}", line)
            elif pin_count==3: new_line=re.sub(r"📍\s*.*", f"📍 {data['KEL']}", line)
            out_lines.append(new_line); continue
        if "KELAMIN" in upper and ":" in line:
            new_line=re.sub(r":\s*.*", f": {data['KELAMIN']}", line)
        elif "SALDO" in upper and ":" in line:
            if "RP" in upper: new_line=re.sub(r":\s*.*", f": Rp. {data['SALDO']}", line)
            else: new_line=re.sub(r":\s*.*", f": {data['SALDO']}", line)
        elif "KPJ" in upper and "SENSOR" not in upper and ":" in line:
            new_line=re.sub(r":\s*.*", f": {data['KPJ']}", line)
        elif "SENSOR" in upper and ":" in line:
            new_line=re.sub(r":\s*.*", f": {data['SENSOR']}", line)
        elif ("IT" in upper or "IURAN" in upper) and ":" in line:
            new_line=re.sub(r":\s*.*", f": {data['IT']}", line)
        elif "PT" in upper and ":" in line:
            new_line=re.sub(r":\s*.*", f": {data['PT']}", line)
        out_lines.append(new_line)
    result="\n".join(out_lines)
    if kode_header:
        result=f"           {kode_header} \n━━━━━━━━━━━━━━━━━━━\n{result}"
    return result



@dp.message(FormatState.waiting_manual, F.text)

def extract_phone_from_text(text):
    import re
    # cari nomor hp Indonesia
    patterns = [
        r'08\d{8,11}',
        r'\+628\d{8,11}',
        r'628\d{8,11}'
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            num = m.group(0)
            # normalisasi ke 08
            if num.startswith('+62'):
                num = '0' + num[3:]
            elif num.startswith('62'):
                num = '0' + num[2:]
            return num
    return None

def build_inbox_keyboard(result_id, result_text):
    phone = extract_phone_from_text(result_text)
    kb = []
    # Baris 1: Salin dan Hapus
    kb.append([
        InlineKeyboardButton(text="📋 Salin", callback_data=f"copy_result_{result_id}"),
        InlineKeyboardButton(text="🗑️ Hapus", callback_data=f"delete_result_{result_id}")
    ])
    # Baris 2: Chat pengirim di WA jika ada nomor
    if phone:
        # buat link WA
        wa_num = phone
        if wa_num.startswith('0'):
            wa_num = '62' + wa_num[1:]
        wa_link = f"https://wa.me/{wa_num}"
        kb.append([
            InlineKeyboardButton(text="💬 Chat Pengirim di WA", url=wa_link)
        ])
    kb.append([
        InlineKeyboardButton(text="✏️ Edit", callback_data=f"edit_result_{result_id}"),
        InlineKeyboardButton(text="➕ Lanjut Bikin", callback_data="lanjut_bikin")
    ])
    kb.append([
        InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")
    ])
    return InlineKeyboardMarkup(inline_keyboard=kb)


async def format_manual_receive(message: Message, state: FSMContext):
    data = await state.get_data()
    edit_id = data.get("edit_result_id")

    conn = db()
    
    # Ambil template dari Database (SETTING FORMAT user)
    row = conn.execute(
        "SELECT template FROM format_settings WHERE telegram_id=?",
        (message.from_user.id,)
    ).fetchone()
    template = row["template"] if row else DEFAULT_TEMPLATE

    kode = generate_next_code(message.from_user.id)
    result = apply_template(template, message.text, kode_header=kode)

    if edit_id:
        conn.execute(
            "UPDATE format_results SET result_text=? WHERE id=? AND telegram_id=?",
            (result, edit_id, message.from_user.id)
        )
        result_id = edit_id
    else:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO format_results(telegram_id,input_text,result_text,created_at)
            VALUES(?,?,?,?)
        """, (message.from_user.id, message.text, result, datetime.now().isoformat()))
        result_id = cur.lastrowid
        # FIX: Jika ada kata AKUN di bawah format, auto masuk ke HASIL FORMAT+AKUN
        akun = parse_data_with_akun(message.text)
        has_akun_keyword = 'AKUN' in message.text.upper()
        if akun.get('EMAIL') or has_akun_keyword:
            raw_upper_idx = message.text.upper().find('AKUN')
            full_akun_block = message.text[raw_upper_idx:].strip() if raw_upper_idx != -1 else ""
            akun_lines_all = full_akun_block.splitlines()
            if akun_lines_all and 'AKUN' in akun_lines_all[0].upper():
                full_akun_content = "\n".join(akun_lines_all[1:]).strip()
            else:
                full_akun_content = full_akun_block
            content_lines = [l.strip() for l in full_akun_content.splitlines() if l.strip()]
            email_save = content_lines[0] if content_lines else (akun.get('EMAIL') or 'AKUN')
            pass_save = "\n".join(content_lines[1:]) if len(content_lines) > 1 else (akun.get('PASSWORD') or full_akun_content[:1000])
            raw_text_save = full_akun_content[:2000]
            cur.execute('''INSERT INTO format_accounts(telegram_id,email,password,raw_text,result_id,created_at) VALUES(?,?,?,?,?,?)''', (message.from_user.id, email_save, pass_save, raw_text_save, result_id, datetime.now().isoformat()))

    conn.commit()
    conn.close()
    await state.set_state(FormatState.waiting_manual)

    await message.answer(
        html.escape(result[:3900]),
        reply_markup=build_inbox_keyboard(result_id, result)
    )



@dp.callback_query(F.data == "format_setting")
async def format_setting(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    conn = db()
    row = conn.execute(
        "SELECT template FROM format_settings WHERE telegram_id=?",
        (callback.from_user.id,)
    ).fetchone()
    code_row = conn.execute(
        "SELECT * FROM format_codes WHERE telegram_id=?",
        (callback.from_user.id,)
    ).fetchone()
    conn.close()
    template = row["template"] if row else DEFAULT_TEMPLATE
    
    kode_info = "❌ Belum diatur"
    if code_row:
        kode_info = f"{code_row['prefix']} - {str(code_row['current_number']).zfill(code_row['padding'])} (Aktif)"
    
    await state.set_state(FormatState.waiting_setting)
    await callback.message.edit_text(
        "⚙️ <b>SETTING FORMAT</b>\n\n"
        f"🔢 <b>KODE SAAT INI:</b> {kode_info}\n\n"
        "Template saat ini:\n\n<pre>" + html.escape(template[:3000]) + "</pre>\n\n"
        "Kirim template baru untuk menggantinya, atau atur kode di bawah:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔢 SET KODE FORMAT", callback_data="set_kode_format")],
            [InlineKeyboardButton(text="❌ HAPUS KODE", callback_data="delete_kode_format")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()



@dp.message(FormatState.waiting_setting, F.text)
async def format_setting_receive(message: Message, state: FSMContext):
    raw = message.text.strip()
    if '{' in raw and '}' in raw:
        template_to_save = raw
    else:
        template_to_save = make_template_from_example(raw)
    conn = db()
    conn.execute("""
        INSERT INTO format_settings(telegram_id,template,updated_at)
        VALUES(?,?,?)
        ON CONFLICT(telegram_id) DO UPDATE SET
            template=excluded.template, updated_at=excluded.updated_at
    """, (message.from_user.id, template_to_save, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    await state.clear()
    # Test preview dengan data dummy
    dummy_input = "JAKARTA\nPENJARINGAN\nBEBAS\n14000\nPEREMPUAN 1992\n2021\n22* 23*\n02-03-2025\nSAKTI MULYA"
    preview = apply_template(template_to_save, dummy_input)
    await message.answer(
        "✅ <b>SETTING FORMAT TERSIMPAN & AKTIF</b>\n\n"
        "Bot sekarang akan <b>ngikutin 100%</b> template ini:\n\n"
        f"<pre>{html.escape(preview[:3800])}</pre>\n\n"
        "Coba BUAT FORMAT baru, hasilnya pasti ngikutin template ini!",
        reply_markup=auto_menu()
    )




@dp.callback_query(F.data == "set_kode_format")
async def set_kode_format(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 Terkunci", show_alert=True)
        return
    await state.set_state(FormatState.waiting_kode)
    await callback.message.edit_text(
        "🔢 <b>SET KODE FORMAT</b>\n\n"
        "Ketik kode awal yang kamu mau. Contoh:\n"
        "<code>JPG - 001</code>\n"
        "<code>ABC - 001</code>\n\n"
        "Bot akan otomatis urut: 001, 002, 003... tanpa duplikat dan muncul di atas format.\n\n"
        "Ketik kode sekarang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ BATAL", callback_data="format_setting")]
        ])
    )
    await callback.answer()


@dp.message(FormatState.waiting_kode, F.text)
async def receive_kode_format(message: Message, state: FSMContext):
    parsed = parse_kode_input(message.text)
    if not parsed:
        await message.answer("❌ Format kode salah. Contoh: <code>JPG - 001</code> atau <code>🤖 JPG - 001 🤖</code>")
        return
    if len(parsed) == 4:
        prefix, num, padding, suffix = parsed
    else:
        prefix, num, padding = parsed
        suffix = ""
    
    conn = db()
    conn.execute("""
        INSERT INTO format_codes(telegram_id,prefix,suffix,current_number,padding,enabled,updated_at)
        VALUES(?,?,?,?,?,1,?)
        ON CONFLICT(telegram_id) DO UPDATE SET
            prefix=excluded.prefix,
            suffix=excluded.suffix,
            current_number=excluded.current_number,
            padding=excluded.padding,
            enabled=1,
            updated_at=excluded.updated_at
    """, (message.from_user.id, prefix, suffix, num, padding, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    await state.clear()
    demo_code_full = f"{prefix} {str(num).zfill(padding)} {suffix}".strip()
    centered_demo = demo_code_full.center(27)
    await message.answer(
        f"✅ <b>KODE DISIMPAN</b>\n\n"
        f"Kode aktif: <b>{demo_code_full}</b>\n\n"
        f"Format selanjutnya akan otomatis jadi:\n"
        f"<pre>{centered_demo}\n━━━━━━━━━━━━━━━━━━━\n[FORMAT KAMU]</pre>\n\n"
        f"Dan berikutnya otomatis berurutan tanpa duplikat. Posisi kode selalu di TENGAH.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ LIHAT SETTING", callback_data="format_setting")],
            [InlineKeyboardButton(text="📝 BUAT FORMAT", callback_data="format_create")]
        ])
    )



@dp.callback_query(F.data == "delete_kode_format")
async def delete_kode_format(callback: CallbackQuery):
    conn = db()
    conn.execute("DELETE FROM format_codes WHERE telegram_id=?", (callback.from_user.id,))
    conn.commit()
    conn.close()
    await callback.message.edit_text(
        "🗑️ <b>KODE DIHAPUS</b>\n\nKode format dinonaktifkan. Format tidak akan pakai kode lagi.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⚙️ KEMBALI KE SETTING", callback_data="format_setting")]
        ])
    )
    await callback.answer()



@dp.callback_query(F.data == "format_excel")
async def format_excel_start(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    await state.set_state(FormatState.waiting_excel)
    await callback.message.edit_text(
        "📊 <b>FORMAT FILE EXCEL</b>\n\nSilakan upload file Excel <b>.xlsx</b>.",
        reply_markup=back_main()
    )
    await callback.answer()


@dp.message(FormatState.waiting_excel, F.document)
async def format_excel_receive(message: Message, state: FSMContext):
    if not (message.document.file_name or "").lower().endswith(".xlsx"):
        await message.answer("❌ File harus berformat .xlsx")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        await message.answer("❌ Tambahkan <code>openpyxl</code> ke requirements.txt lalu deploy ulang.")
        return

    file = await message.bot.get_file(message.document.file_id)
    local = Path("/tmp") / f"{message.from_user.id}_{message.document.file_name}"
    await message.bot.download_file(file.file_path, local)

    conn = db()
    row = conn.execute(
        "SELECT template FROM format_settings WHERE telegram_id=?",
        (message.from_user.id,)
    ).fetchone()
    template = row["template"] if row else DEFAULT_TEMPLATE

    wb = load_workbook(local, read_only=True, data_only=True)
    ws = wb.active
    results = []

    for values in ws.iter_rows(values_only=True):
        raw = " ".join(str(v) for v in values if v is not None).strip()
        if raw:
            results.append(apply_template(template, raw))

    if not results:
        conn.close()
        await state.clear()
        await message.answer("❌ File Excel tidak memiliki data.")
        return

    result = "\n\n".join(results)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO format_results(telegram_id,input_text,result_text,created_at)
        VALUES(?,?,?,?)
    """, (message.from_user.id, message.document.file_name, result, datetime.now().isoformat()))
    rid = cur.lastrowid
    conn.commit()
    conn.close()
    await state.clear()

    await message.answer(
        "✅ <b>HASIL FORMAT EXCEL</b>\n\n" + html.escape(result[:3900]),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 SALIN SEMUA HASIL", callback_data=f"copy_all_{rid}")],
            [InlineKeyboardButton(text="💾 SIMPAN KE HISTORY", callback_data=f"save_result_{rid}")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )


@dp.message(FormatState.waiting_excel)
async def format_excel_wrong(message: Message):
    await message.answer("📊 Silakan upload file Excel .xlsx")




@dp.callback_query(F.data == "format_results")
async def format_results(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    conn = db()
    rows = conn.execute(
        "SELECT id, result_text, created_at FROM format_results WHERE telegram_id=? ORDER BY id ASC LIMIT 10",
        (callback.from_user.id,)
    ).fetchall()
    conn.close()
    if not rows:
        await callback.message.edit_text(
            "📄 <b>HASIL FORMAT - KOSONG</b>\n\nBelum ada hasil.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📝 BUAT FORMAT BARU", callback_data="format_create")],
                [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
            ])
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        f"📄 <b>HASIL FORMAT TERKINI - {len(rows)} TERBARU</b>\n\nMenampilkan {len(rows)} format di bawah, tiap format ada tombolnya sendiri:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU", callback_data="auto_format")]
        ])
    )
    await callback.answer()

    for i, row in enumerate(rows, 1):
        rid = row["id"]
        full_text = row["result_text"].strip()
        if len(full_text) > 3800:
            full_text = full_text[:3800] + "..."
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="📋 SALIN", callback_data=f"copy_result_{rid}"),
                InlineKeyboardButton(text="✏️ EDIT", callback_data=f"edit_result_{rid}"),
                InlineKeyboardButton(text="🗑️ HAPUS", callback_data=f"delete_result_{rid}")
            ]
        ])
        try:
            await callback.message.bot.send_message(
                callback.from_user.id,
                f"<b>{i}. ID {rid}</b>\n<pre>{html.escape(full_text)}</pre>",
                reply_markup=kb
            )
            await asyncio.sleep(0.2)
        except Exception:
            pass

    await callback.message.bot.send_message(
        callback.from_user.id,
        "🔍 <b>MENU LANJUTAN</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔎 CARI HASIL FORMAT", callback_data="format_search")],
            [InlineKeyboardButton(text="📧 HASIL FORMAT+AKUN", callback_data="hasil_akun")],
            [InlineKeyboardButton(text="🗑️ HAPUS SEMUA HASIL", callback_data="clear_results")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )

@dp.callback_query(F.data == "clear_results")
async def clear_results(callback: CallbackQuery):
    conn = db()
    conn.execute("DELETE FROM format_results WHERE telegram_id=?", (callback.from_user.id,))
    conn.commit(); conn.close()
    await callback.message.edit_text("🗑️ Semua hasil format dihapus.", reply_markup=auto_menu())
    await callback.answer()



@dp.callback_query(F.data == "format_search")
async def format_search(callback: CallbackQuery, state: FSMContext):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    await state.set_state(FormatState.waiting_search)
    await callback.message.edit_text(
        "🔎 <b>CARI HASIL FORMAT</b>\n\n"
        "Ketik kata yang ingin dicari, misalnya:\n"
        "<code>SERANG</code> atau <code>CIPOCOK</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU HASIL FORMAT", callback_data="format_results")]
        ])
    )
    await callback.answer()




@dp.message(FormatState.waiting_search, F.text)
async def format_search_receive(message: Message, state: FSMContext):
    query = message.text.strip()
    if not query:
        await message.answer("❌ Kata pencarian tidak boleh kosong.")
        return

    conn = db()
    rows = conn.execute("""
        SELECT id, result_text, created_at
        FROM format_results
        WHERE telegram_id=? AND result_text LIKE ?
        ORDER BY id ASC LIMIT 20
    """, (message.from_user.id, f"%{query}%")).fetchall()
    conn.close()
    # JANGAN clear state, biar bisa cari terus menerus
    # await state.clear()  # dihapus

    if not rows:
        await message.answer(
            f"🔍 Tidak ada format untuk '<code>{query}</code>'\n\nCoba kata lain:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔎 CARI LAGI", callback_data="format_search")],
                [InlineKeyboardButton(text="📄 LIHAT SEMUA", callback_data="format_results")],
                [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
            ])
        )
        return

    # Langsung kirim format saja, tanpa kata HASIL CARI
    for row in rows:
        rid = row["id"]
        txt = row["result_text"]
        await message.bot.send_message(
            message.from_user.id,
            f"<pre>{html.escape(txt[:3800])}</pre>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="📋 SALIN", callback_data=f"copy_result_{rid}"),
                    InlineKeyboardButton(text="🗑️ HAPUS", callback_data=f"delete_result_{rid}"),
                    InlineKeyboardButton(text="✏️ EDIT", callback_data=f"edit_result_{rid}")
                ]
            ])
        )
        await asyncio.sleep(0.1)




@dp.callback_query(F.data == "hasil_akun")
async def hasil_akun(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 Terkunci.", show_alert=True)
        return
    conn = db()
    rows = conn.execute("""
        SELECT a.id, a.email, a.password, a.created_at, a.result_id,
               r.result_text, r.input_text
        FROM format_accounts a
        LEFT JOIN format_results r ON r.id = a.result_id
        WHERE a.telegram_id=?
        ORDER BY a.id ASC LIMIT 10
    """, (callback.from_user.id,)).fetchall()
    conn.close()
    if not rows:
        await callback.message.edit_text(
            "📧 <b>HASIL FORMAT+AKUN - KOSONG</b>\n\nBelum ada akun. Buat format dengan tambahan:\n<code>AKUN:\nemail@gmail.com\npassword123</code>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📝 BUAT FORMAT BARU", callback_data="format_create")],
                [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
            ])
        )
        await callback.answer()
        return

    await callback.message.edit_text(
        f"📧 <b>HASIL FORMAT+AKUN - {len(rows)} TERBARU</b>\n\nMenampilkan {len(rows)} akun di bawah:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")]
        ])
    )
    await callback.answer()

    for i, row in enumerate(rows, 1):
        rid = row["id"]
        email = html.escape(row['email'] or '-')
        pwd = html.escape(row['password'] or '-')
        result_id = row['result_id']
        # Ambil SEX CODE dari result_text
        result_text = (row["result_text"] or "").strip()
        sex_code = "SEX -"
        if result_text:
            first = result_text.splitlines()[0].strip()
            # kalau baris pertama ada SEX, pakai itu
            if "SEX" in first.upper() or len(first) < 30:
                sex_code = first
            else:
                # cari baris yang ada SEX
                for ln in result_text.splitlines()[:3]:
                    if "SEX" in ln.upper():
                        sex_code = ln.strip()
                        break
        sex_code = html.escape(sex_code)

        kb = InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="📋 SALIN", callback_data=f"copy_akun_{rid}"),
                InlineKeyboardButton(text="✏️ EDIT", callback_data=f"edit_akun_{rid}"),
                InlineKeyboardButton(text="🗑️ HAPUS", callback_data=f"del_akun_{rid}")
            ]
        ])
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"<b>{i}. ID {rid} (Format ID: {result_id})</b>\n"
            f"<pre>{sex_code}\n"
            f"───────────────────\n"
            f"AKUN:\n{email}\n{pwd}</pre>",
            reply_markup=kb
        )
        await asyncio.sleep(0.15)

    await callback.message.bot.send_message(
        callback.from_user.id,
        "🔍 <b>MENU AKUN LAINNYA</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔎 CARI AKUN", callback_data="search_akun")],
            [InlineKeyboardButton(text="📄 HASIL FORMAT TERKINI", callback_data="format_results")],
            [InlineKeyboardButton(text="🗑️ HAPUS SEMUA AKUN", callback_data="hapus_akun_all")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )

@dp.callback_query(F.data.startswith("copy_akun_"))
async def copy_akun(callback: CallbackQuery):
    aid = int(callback.data.replace("copy_akun_",""))
    conn = db()
    row = conn.execute("""
        SELECT a.email, a.password, r.result_text 
        FROM format_accounts a 
        LEFT JOIN format_results r ON r.id = a.result_id 
        WHERE a.id=? AND a.telegram_id=?
    """, (aid, callback.from_user.id)).fetchone()
    conn.close()
    if row:
        result_text = (row["result_text"] or "").strip()
        sex_code = result_text.splitlines()[0].strip() if result_text else "SEX -"
        for ln in result_text.splitlines()[:3]:
            if "SEX" in ln.upper():
                sex_code = ln.strip()
                break
        await callback.message.bot.send_message(
            callback.from_user.id, 
            f"📋 <b>SALIN AKUN:</b>\n<code>{html.escape(sex_code)}\n───────────────────\nAKUN:\n{html.escape(row['email'])}\n{html.escape(row['password'])}</code>"
        )
    await callback.answer("Disalin!")

@dp.callback_query(F.data.startswith("del_akun_"))
async def del_akun(callback: CallbackQuery):
    aid = int(callback.data.replace("del_akun_",""))
    conn = db(); conn.execute("DELETE FROM format_accounts WHERE id=? AND telegram_id=?", (aid, callback.from_user.id)); conn.commit(); conn.close()
    await callback.answer("Dihapus!")
    await hasil_akun(callback)

@dp.callback_query(F.data.startswith("edit_akun_"))
async def edit_akun_start(callback: CallbackQuery, state: FSMContext):
    aid = int(callback.data.replace("edit_akun_",""))
    await state.set_state(FormatState.waiting_edit_akun)
    await state.update_data(edit_id=aid)
    conn = db(); row = conn.execute("SELECT email,password FROM format_accounts WHERE id=? AND telegram_id=?", (aid, callback.from_user.id)).fetchone(); conn.close()
    if row:
        await callback.message.bot.send_message(callback.from_user.id, f"✏️ <b>EDIT AKUN ID {aid}</b>\n\nLama:\n<code>{row['email']}\n{row['password']}</code>\n\nKirim baru:\nemail\npassword")
    await callback.answer()

@dp.message(FormatState.waiting_edit_akun, F.text)
async def edit_akun_save(message: Message, state: FSMContext):
    data = await state.get_data()
    aid = data.get("edit_id")
    lines = [x.strip() for x in message.text.splitlines() if x.strip()]
    email = lines[0] if len(lines)>=1 else ""
    password = lines[1] if len(lines)>=2 else ""
    if "@" not in email:
        await message.answer("❌ Email tidak valid. Kirim lagi: email\npassword")
        return
    conn = db(); conn.execute("UPDATE format_accounts SET email=?, password=?, raw_text=? WHERE id=? AND telegram_id=?", (email, password, f"{email}\n{password}", aid, message.from_user.id)); conn.commit(); conn.close()
    await state.clear()
    await message.answer(f"✅ Akun ID {aid} diedit!", reply_markup=auto_menu())

@dp.callback_query(F.data == "search_akun")
async def search_akun_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(FormatState.waiting_search_akun)
    await callback.message.bot.send_message(callback.from_user.id, "🔍 Kirim kata kunci untuk cari di HASIL AKUN+EMAIL:")
    await callback.answer()

@dp.message(FormatState.waiting_search_akun, F.text)
async def search_akun_do(message: Message, state: FSMContext):
    keyword = message.text.strip()
    conn = db()
    rows = conn.execute("""
        SELECT a.id, a.email, a.password, a.raw_text, r.result_text 
        FROM format_accounts a 
        LEFT JOIN format_results r ON r.id = a.result_id 
        WHERE a.telegram_id=? AND (a.email LIKE ? OR a.password LIKE ? OR a.raw_text LIKE ? OR r.result_text LIKE ?) 
        ORDER BY a.id ASC LIMIT 20
    """, (message.from_user.id, f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")).fetchall()
    conn.close()
    # JANGAN clear, biar bisa cari terus
    if not rows:
        await message.answer(
            f"🔍 Tidak ada akun untuk '<code>{keyword}</code>'\n\nCoba kata lain:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔎 CARI AKUN LAGI", callback_data="search_akun")],
                [InlineKeyboardButton(text="📄 HASIL AKUN", callback_data="hasil_akun")],
                [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
            ])
        )
        return
    await message.answer(f"🔍 Hasil '{keyword}' - {len(rows)} akun:")
    for r in rows:
        kb = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="📋 Salin", callback_data=f"copy_akun_{r['id']}"), InlineKeyboardButton(text="🗑️ Hapus", callback_data=f"del_akun_{r['id']}")]])
        await message.bot.send_message(message.from_user.id, f"<code>{html.escape(r['email'])}\n{html.escape(r['password'])}</code>", reply_markup=kb)

    await message.bot.send_message(
        message.from_user.id,
        f"✅ <b>{len(rows)} AKUN DITEMUKAN</b> untuk '<code>{keyword}</code>'\n\nKetik lagi untuk cari terus:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔎 CARI AKUN LAGI", callback_data="search_akun")],
            [InlineKeyboardButton(text="📄 HASIL AKUN", callback_data="hasil_akun")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )

@dp.callback_query(F.data == "hapus_akun_all")
async def hapus_akun_all(callback: CallbackQuery):
    conn = db(); conn.execute("DELETE FROM format_accounts WHERE telegram_id=?", (callback.from_user.id,)); conn.commit(); conn.close()
    await callback.message.edit_text("🗑️ Semua akun dihapus.", reply_markup=auto_menu())
    await callback.answer()






@dp.callback_query(F.data == "menu_warna")
async def menu_warna(callback: CallbackQuery):
    await callback.message.answer(
        "🎨 <b>MENU PROFESIONAL</b>\n\n"
        "⚠️ Warna background tidak bisa diubah di Telegram (limit API).\n"
        "Semua tombol ikut tema HP (hijau di screenshot).\n\n"
        "🟡 TOP UP = Emas Muda (pakai emoji)\n"
        "🔵 Lainnya = Biru",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🟡 TOP UP", callback_data="topup")],
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="back_main")]
        ])
    )
    await callback.answer()



@dp.callback_query(F.data == "format_history")
async def format_history(callback: CallbackQuery):
    if not has_auto_format_access(callback.from_user.id):
        await callback.answer("🔒 AUTO FORMAT masih terkunci.", show_alert=True)
        return
    conn = db()
    rows = conn.execute("""
        SELECT id,result_text,deleted_at FROM format_history
        WHERE telegram_id=? ORDER BY id ASC LIMIT 20
    """, (callback.from_user.id,)).fetchall()
    conn.close()

    try:
        await callback.message.delete()
    except:
        pass

    if not rows:
        await callback.message.bot.send_message(
            callback.from_user.id,
            "🕘 <b>RIWAYAT HASIL FORMAT - KOSONG</b>\n\nBelum ada format yang dihapus.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📄 LIHAT HASIL FORMAT", callback_data="format_results")],
                [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
            ])
        )
        await callback.answer()
        return

    await callback.message.bot.send_message(
        callback.from_user.id,
        f"🕘 <b>RIWAYAT HASIL FORMAT - {len(rows)} DATA</b>\n\nFormat yang dihapus ada di sini. Bisa dihapus permanen atau dipulihkan ke HASIL FORMAT.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗑️ HAPUS SEMUA RIWAYAT", callback_data="clear_history")],
            [InlineKeyboardButton(text="📄 LIHAT HASIL FORMAT", callback_data="format_results")]
        ])
    )

    for r in rows:
        rid = r["id"]
        txt = r["result_text"]
        try:
            dt = r["deleted_at"][:16].replace("T"," ")
        except:
            dt = ""
        await callback.message.bot.send_message(
            callback.from_user.id,
            f"<pre>{html.escape(txt[:3800])}</pre>\n<i>#{rid} • Dihapus: {dt}</i>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="♻️ PULIHKAN", callback_data=f"restore_history_{rid}"),
                    InlineKeyboardButton(text="🗑️ HAPUS PERMANEN", callback_data=f"delete_history_{rid}")
                ]
            ])
        )
        await asyncio.sleep(0.1)
    await callback.answer(f"📂 {len(rows)} riwayat")


@dp.callback_query(F.data.startswith("copy_result_"))
async def copy_result_handler(callback: CallbackQuery):
    result_id = int(callback.data.replace("copy_result_", ""))
    conn = db()
    row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (result_id, callback.from_user.id)).fetchone()
    conn.close()
    if row:
        # kirim ulang text biar bisa di-copy
        await callback.message.answer(
            f"📋 <b>SALIN TEXT:</b>\n\n<code>{html.escape(row['result_text'][:3800])}</code>\n\nTekan tahan untuk salin!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="format_results")]
            ])
        )
        await callback.answer("📋 Siap disalin!")
    else:
        await callback.answer("❌ Tidak ditemukan", show_alert=True)


@dp.callback_query(F.data.startswith("delete_result_"))
async def delete_result(callback: CallbackQuery):
    rid = int(callback.data.split("_")[-1])
    conn = db()
    row = conn.execute("SELECT result_text FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id)).fetchone()
    if row:
        conn.execute("""
            INSERT INTO format_history(telegram_id,input_text,result_text,created_at,deleted_at)
            VALUES(?,?,?,?,?)
        """, (callback.from_user.id, "", row["result_text"], datetime.now().isoformat(), datetime.now().isoformat()))
        conn.execute("DELETE FROM format_results WHERE id=? AND telegram_id=?", (rid, callback.from_user.id))
        conn.commit()
    conn.close()
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("🗑️ Dipindahkan ke RIWAYAT")

@dp.callback_query(F.data.startswith("restore_history_"))
async def restore_history(callback: CallbackQuery):
    hid = int(callback.data.split("_")[-1])
    conn = db()
    row = conn.execute("SELECT * FROM format_history WHERE id=? AND telegram_id=?", (hid, callback.from_user.id)).fetchone()
    if not row:
        conn.close()
        await callback.answer("Riwayat tidak ditemukan", show_alert=True)
        return
    conn.execute("""
        INSERT INTO format_results(telegram_id,input_text,result_text,created_at)
        VALUES(?,?,?,?)
    """, (callback.from_user.id, row["input_text"], row["result_text"], datetime.now().isoformat()))
    conn.execute("DELETE FROM format_history WHERE id=? AND telegram_id=?", (hid, callback.from_user.id))
    conn.commit()
    conn.close()
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("♻️ Dipulihkan ke HASIL FORMAT")
    await callback.message.bot.send_message(
        callback.from_user.id,
        f"<pre>{html.escape(row['result_text'][:3800])}</pre>\n✅ <b>Berhasil dipulihkan ke HASIL FORMAT</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📄 LIHAT HASIL FORMAT", callback_data="format_results")],
            [InlineKeyboardButton(text="🕘 LIHAT RIWAYAT", callback_data="format_history")]
        ])
    )

@dp.callback_query(F.data.startswith("delete_history_"))
async def delete_history(callback: CallbackQuery):
    hid = int(callback.data.split("_")[-1])
    conn = db()
    conn.execute("DELETE FROM format_history WHERE id=? AND telegram_id=?", (hid, callback.from_user.id))
    conn.commit()
    conn.close()
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("🗑️ Dihapus permanen dari riwayat")





@dp.callback_query(F.data == "clear_history")
async def clear_history(callback: CallbackQuery):
    conn = db()
    conn.execute("DELETE FROM format_history WHERE telegram_id=?", (callback.from_user.id,))
    conn.commit()
    conn.close()
    await callback.answer("🗑️ Riwayat dibersihkan")
    try:
        await callback.message.delete()
    except:
        pass
    await callback.message.bot.send_message(
        callback.from_user.id,
        "🕘 Riwayat sudah dibersihkan.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📄 LIHAT HASIL FORMAT", callback_data="format_results")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )



# ==================== HUBUNGI ADMIN ====================

@dp.callback_query(F.data == "contact_admin")
async def contact_admin(callback: CallbackQuery):
    await callback.message.edit_text(
        "📞 <b>HUBUNGI ADMIN</b>\n\n"
        "Jika membutuhkan bantuan, silakan hubungi Admin:\n\n"
        "👤 <b>Telegram</b>\n"
        "@Hambali1995\n\n"
        "📱 <b>WhatsApp</b>\n"
        "083160776091",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💬 CHAT TELEGRAM", url="https://t.me/Hambali1995")],
            [InlineKeyboardButton(text="📱 CHAT WHATSAPP", url="https://wa.me/6283160776091")],
            [InlineKeyboardButton(text="⬅️ KEMBALI KE MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer()






# ==================== FILTER KOTA & KECAMATAN - WA GRUP KE TELEGRAM USER ====================
# Tujuan bot: memfilter kota/kecamatan yang dipilih user/admin
# Ketika ada share di grup WA dengan kata kunci KOTA & KECAMATAN -> otomatis masuk ke tele user

def extract_phone_from_any(text):
    import re
    patterns = [r'08\d{8,11}', r'\+628\d{8,11}', r'628\d{8,11}']
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            num = m.group(0)
            if num.startswith('+62'):
                num = '0' + num[3:]
            elif num.startswith('62'):
                num = '0' + num[2:]
            return num
    return None

def is_match_kota(text, kota_row):
    # HARUS KOTA DAN KECAMATAN DUA-DUANYA - kalau kota saja tidak masuk!
    text_upper = text.upper()
    
    # Ambil data dari row
    # Format user_kota: kota_name = "Serang - Cipocok Jaya", kab = provinsi, kec = kecamatan, kab = kabupaten
    # Kita cek KOTA (kabupaten) dan KECAMATAN harus ada di text
    
    kab = (kota_row["kab"] or "").upper()  # Ini bisa berisi Provinsi atau Kabupaten
    kec = (kota_row["kec"] or "").upper()  # Kecamatan
    kota_name = (kota_row["kota_name"] or "").upper()  # Contoh: "Serang - Cipocok Jaya"
    kel = (kota_row["kel"] or "").upper()
    
    # Parse kota_name kalau format "KABUPATEN - KECAMATAN"
    kota_part = ""
    kec_part = ""
    if " - " in kota_row["kota_name"]:
        parts = kota_row["kota_name"].split(" - ")
        if len(parts) >= 2:
            kota_part = parts[0].strip().upper()
            kec_part = parts[1].strip().upper()
    
    # Cek KOTA ada di text?
    kota_match = False
    if kab and len(kab) > 2 and kab in text_upper:
        kota_match = True
    if kota_part and len(kota_part) > 2 and kota_part in text_upper:
        kota_match = True
    if kota_name and len(kota_name) > 2:
        # cek juga kalau kota_name sendiri ada
        # tapi untuk requirement KOTA + KECAMATAN, kita butuh 2 kategori
        pass
    
    # Cek KECAMATAN ada di text?
    kec_match = False
    if kec and len(kec) > 2 and kec in text_upper:
        kec_match = True
    if kec_part and len(kec_part) > 2 and kec_part in text_upper:
        kec_match = True
    
    # HARUS KOTA DAN KECAMATAN ADA DUA-DUANYA!
    # Contoh: User pilih KOTA=SERANG, KECAMATAN=CIPOCOK JAYA
    # Pesan harus mengandung SERANG dan CIPOCOK JAYA baru masuk
    # Kalau cuma SERANG saja -> TIDAK MASUK
    
    if kota_match and kec_match:
        return True
    
    # Fallback: kalau data lama tidak ada pemisah - cek apakah text mengandung minimal 2 kata dari kombinasi
    # Misal kota_name = "SERANG - CIPOCOK JAYA" -> harus ada SERANG dan CIPOCOK JAYA di text
    if kota_part and kec_part:
        if kota_part in text_upper and kec_part in text_upper:
            return True
    
    return False

def is_match_kota_strict(text):
    # Helper untuk debug: harus ada KOTA dan KECAMATAN kategori
    # Tidak boleh hanya kota saja
    return is_match_kota(text, {"kota_name": "", "kab": "", "kec": "", "kel": ""})

def build_filtered_inbox_keyboard(phone, message_id=None, chat_id=None):
    kb = []
    mid = message_id or 0
    cid = chat_id or 0
    kb.append([
        InlineKeyboardButton(text="📋 Salin", callback_data=f"inbox_copy_{cid}_{mid}"),
        InlineKeyboardButton(text="🗑️ Hapus", callback_data=f"inbox_del_{mid}")
    ])
    if phone:
        wa_num = phone
        if wa_num.startswith('0'):
            wa_num = '62' + wa_num[1:]
        wa_link = f"https://wa.me/{wa_num}"
        kb.append([InlineKeyboardButton(text="💬 Chat Pengirim di WA", url=wa_link)])
    return InlineKeyboardMarkup(inline_keyboard=kb)

# Handler untuk pesan yang di-forward dari WA (via bot atau via grup Tele yang mirroring WA)
@dp.message(F.chat.type.in_({"group", "supergroup"}))
async def group_message_filter_handler(message: Message):
    if message.from_user and message.from_user.is_bot:
        return
    
    text = message.text or message.caption or ""
    if not text or len(text) < 5:
        return
    
    phone = extract_phone_from_any(text)
    
    # Ambil info grup dan pengirim untuk format
    grup_name = message.chat.title or "GRUP WA"
    pengirim_name = message.from_user.full_name if message.from_user else "Unknown"
    no_hp = phone or "Tidak ada nomor"
    
    # AMBIL SEMUA USER DAN CEK FILTER KOTA MEREKA
    conn = db()
    # Ambil semua kota yang dipilih semua user
    all_kota = conn.execute("SELECT * FROM user_kota").fetchall()
    
    # Buat mapping user_id -> list kota mereka
    user_kota_map = {}
    for kota_row in all_kota:
        uid = kota_row["telegram_id"]
        if uid not in user_kota_map:
            user_kota_map[uid] = []
        user_kota_map[uid].append(kota_row)
    
    conn.close()
    
    # Untuk setiap user, cek apakah pesan ini match dengan kota mereka
    matched_users = []
    for uid, kota_list in user_kota_map.items():
        for kota_row in kota_list:
            if is_match_kota(text, kota_row):
                matched_users.append(uid)
                break  # cukup 1 kota match, langsung masuk list
    
    if not matched_users:
        # Tidak ada user yang match, jangan kirim
        return
    
    # Format persis sesuai request
    formatted_text = (
        f"🛡️GRUP: {grup_name} 🕷️\n"
        f"🧑‍💻PENGIRIM: {pengirim_name}\n"
        f"📱NO HP: {no_hp}\n"
        f"\n"
        f"{text}\n"
        f"\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"⚠️ PERHATIAN :\n"
        f"Tetap hati-hati dan waspada dalam bertransaksi untuk lebih aman gunakan jasa rekber.Terimakasih🙏\n"
        f"Sumber: https//t.me/seduluranjht_bot"
    )
    
    keyboard = build_filtered_inbox_keyboard(phone, message.message_id, message.chat.id)
    
    # Kirim hanya ke user yang match filter kota
    for uid in set(matched_users):
        try:
            await bot.send_message(
                chat_id=uid,
                text=formatted_text[:3900],
                reply_markup=keyboard,
                parse_mode="HTML"
            )
        except:
            continue

# Handler untuk pesan yang di-forward manual ke bot (simulasi dari WA)
@dp.message(F.chat.type == "private", F.text)
async def private_forward_filter(message: Message, state: FSMContext):
    # Jika sedang dalam state lain (misal setting, dll), skip
    current_state = await state.get_state()
    if current_state:
        return
    
    text = message.text
    if not text or len(text) < 5:
        return
    
    # Cek apakah ini pesan dengan format kota/kec (bukan perintah bot)
    if text.startswith("/") or text.startswith("👤") or text.startswith("🏙️"):
        return
    
    phone = extract_phone_from_any(text)
    if not phone and "Rp" not in text and "SERANG" not in text.upper() and "CIPOCOK" not in text.upper():
        # Bukan pesan jualan/kota, skip
        return
    
    # Ini adalah pesan yang mirip share dari WA grup, proses filter
    conn = db()
    all_kota = conn.execute("SELECT * FROM user_kota").fetchall()
    user_kota_map = {}
    for kota_row in all_kota:
        uid = kota_row["telegram_id"]
        if uid not in user_kota_map:
            user_kota_map[uid] = []
        user_kota_map[uid].append(kota_row)
    conn.close()
    
    matched_users = []
    for uid, kota_list in user_kota_map.items():
        for kota_row in kota_list:
            if is_match_kota(text, kota_row):
                matched_users.append(uid)
                break
    
    if not matched_users:
        # Tidak match dengan filter siapapun, tetap simpan sebagai hasil format jika ada akses
        # Tapi jangan spam
        return
    
    formatted_text = (
        f"🛡️GRUP: WA GROUP (Forward) 🕷️\n"
        f"🧑‍💻PENGIRIM: {message.from_user.full_name}\n"
        f"📱NO HP: {phone or 'Tidak ada'}\n"
        f"\n"
        f"{text}\n"
        f"\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"⚠️ PERHATIAN :\n"
        f"Tetap hati-hati dan waspada dalam bertransaksi untuk lebih aman gunakan jasa rekber.Terimakasih🙏\n"
        f"Sumber: https//t.me/seduluranjht_bot"
    )
    
    keyboard = build_filtered_inbox_keyboard(phone, message.message_id, message.chat.id)
    
    for uid in set(matched_users):
        if uid == message.from_user.id:
            continue  # jangan kirim ke pengirim sendiri
        try:
            await bot.send_message(chat_id=uid, text=formatted_text[:3900], reply_markup=keyboard, parse_mode="HTML")
        except:
            continue

@dp.callback_query(F.data.startswith("inbox_copy_"))
async def inbox_copy_handler(callback: CallbackQuery):
    try:
        text = callback.message.text or ""
        await callback.message.answer(f"📋 <b>SALIN TEXT:</b>\n\n<code>{html.escape(text[:3800])}</code>\n\nTekan tahan untuk salin!")
        await callback.answer("📋 Siap disalin!")
    except:
        await callback.answer("❌ Gagal salin", show_alert=True)

@dp.callback_query(F.data.startswith("inbox_del_"))
async def inbox_del_handler(callback: CallbackQuery):
    try:
        await callback.message.delete()
        await callback.answer("🗑️ Pesan dihapus")
    except:
        await callback.answer("❌ Gagal hapus", show_alert=True)


# ==================== NO BLACKLIST ====================
@dp.callback_query(F.data == "no_blacklist")
async def no_blacklist_handler(callback: CallbackQuery, state: FSMContext):
    conn = db()
    rows = conn.execute("SELECT * FROM blacklist_numbers ORDER BY id DESC LIMIT 50").fetchall()
    conn.close()
    
    if not rows:
        text = (
            "🚫 <b>NO BLACKLIST</b>\n\n"
            "Belum ada nomor blacklist.\n\n"
            "Untuk cari manual, ketik:\n"
            "<code>/cari 083812345678</code>\n\n"
            "Atau kirim nomor langsung di chat ini untuk cek."
        )
    else:
        text = f"🚫 <b>NO BLACKLIST - {len(rows)} NOMOR</b>\n\n"
        for i, r in enumerate(rows, 1):
            phone = r["phone"]
            reason = r["reason"] or "-"
            created = r["created_at"][:10] if r["created_at"] else "-"
            text += f"{i}. <code>{phone}</code>\n   Alasan: {reason} | {created}\n"
        if len(text) > 3800:
            text = text[:3800] + "\n..."
        text += "\n\n🔍 <b>CARI MANUAL:</b>\nKetik <code>/cari 0838xxxx</code> atau kirim nomornya langsung."
    
    await state.set_state(FormatState.waiting_search)
    kb = [
        [InlineKeyboardButton(text="🔍 CARI NOMOR", callback_data="blacklist_search")],
        [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
    ]
    # TAMBAH BLACKLIST hanya untuk admin, tidak tampil di user
    await callback.message.edit_text(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer()

class BlacklistState(StatesGroup):
    waiting_search = State()
    waiting_add = State()

@dp.callback_query(F.data == "blacklist_search")
async def blacklist_search_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(BlacklistState.waiting_search)
    await callback.message.edit_text(
        "🔍 <b>CARI NO BLACKLIST</b>\n\n"
        "Kirim nomor HP yang mau dicek.\n"
        "Contoh: <code>083812345678</code>\n\n"
        "Atau pakai perintah: <code>/cari 083812345678</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="no_blacklist")]
        ])
    )
    await callback.answer()

@dp.callback_query(F.data == "blacklist_add")
async def blacklist_add_start_DEPRECATED(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Hanya admin yang bisa tambah blacklist", show_alert=True)
        return
    await state.set_state(BlacklistState.waiting_add)
    await callback.message.edit_text(
        "➕ <b>TAMBAH BLACKLIST</b>\n\n"
        "Kirim format:\n"
        "<code>nomor|alasan</code>\n\n"
        "Contoh:\n"
        "<code>083812345678|Penipuan</code>\n"
        "<code>083812345678</code>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="no_blacklist")]
        ])
    )
    await callback.answer()

@dp.message(BlacklistState.waiting_search, F.text)
async def blacklist_search_process(message: Message, state: FSMContext):
    phone = "".join([c for c in message.text if c.isdigit()])
    if len(phone) < 10:
        await message.answer("❌ Nomor tidak valid. Contoh: 083812345678")
        return
    
    conn = db()
    row = conn.execute("SELECT * FROM blacklist_numbers WHERE phone LIKE ?", (f"%{phone}%",)).fetchone()
    rows_like = conn.execute("SELECT * FROM blacklist_numbers WHERE phone LIKE ? ORDER BY id DESC LIMIT 10", (f"%{phone}%",)).fetchall()
    conn.close()
    
    if rows_like:
        text = f"🔍 <b>HASIL CARI: {phone}</b>\n\n"
        for r in rows_like:
            text += f"🚫 <code>{r['phone']}</code> - BLACKLIST\n   Alasan: {r['reason'] or '-'}\n   Tgl: {r['created_at'][:10]}\n\n"
        text += f"⚠️ Nomor {phone} ada di blacklist!"
    else:
        text = f"✅ <b>HASIL CARI: {phone}</b>\n\nNomor <code>{phone}</code> TIDAK ada di blacklist.\n✅ AMAN!"
    
    await message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔍 CARI LAGI", callback_data="blacklist_search")],
            [InlineKeyboardButton(text="🚫 LIHAT BLACKLIST", callback_data="no_blacklist")],
            [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
        ])
    )

@dp.message(BlacklistState.waiting_add, F.text)
async def blacklist_add_process(message: Message, state: FSMContext):
    parts = message.text.split("|")
    phone = "".join([c for c in parts[0] if c.isdigit()])
    reason = parts[1].strip() if len(parts) > 1 else "Blacklist"
    
    if len(phone) < 10:
        await message.answer("❌ Nomor tidak valid")
        return
    
    conn = db()
    try:
        conn.execute(
            "INSERT INTO blacklist_numbers(phone,reason,added_by,created_at) VALUES(?,?,?,?)",
            (phone, reason, message.from_user.id, __import__('datetime').datetime.now().isoformat())
        )
        conn.commit()
        await message.answer(f"✅ Nomor <code>{phone}</code> berhasil ditambah ke blacklist!", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🚫 LIHAT BLACKLIST", callback_data="no_blacklist")]]))
    except Exception as e:
        if "UNIQUE" in str(e):
            await message.answer(f"⚠️ Nomor <code>{phone}</code> sudah ada di blacklist!")
        else:
            await message.answer(f"❌ Gagal: {e}")
    finally:
        conn.close()
        await state.clear()

# Command /cari manual
@dp.message(Command("cari"))
async def cari_blacklist_command(message: Message, state: FSMContext):
    args = message.text.replace("/cari", "").strip()
    phone = "".join([c for c in args if c.isdigit()])
    
    if not phone:
        await message.answer(
            "🔍 <b>CARI BLACKLIST</b>\n\n"
            "Format: <code>/cari 083812345678</code>\n\n"
            "Contoh: <code>/cari 083812345678</code>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🚫 LIHAT BLACKLIST", callback_data="no_blacklist")]
            ])
        )
        return
    
    conn = db()
    rows_like = conn.execute("SELECT * FROM blacklist_numbers WHERE phone LIKE ? ORDER BY id DESC LIMIT 10", (f"%{phone}%",)).fetchall()
    conn.close()
    
    if rows_like:
        text = f"🔍 <b>HASIL CARI: {phone}</b>\n\n"
        for r in rows_like:
            text += f"🚫 <code>{r['phone']}</code> - BLACKLIST\n   Alasan: {r['reason'] or '-'}\n   Tgl: {r['created_at'][:10]}\n\n"
        text += f"⚠️ Nomor {phone} ada di blacklist!"
    else:
        text = f"✅ <b>HASIL CARI: {phone}</b>\n\nNomor <code>{phone}</code> TIDAK ada di blacklist.\n✅ AMAN!"
    
    await message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔍 CARI LAGI", callback_data="blacklist_search")],
            [InlineKeyboardButton(text="🚫 LIHAT BLACKLIST", callback_data="no_blacklist")],
            [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
        ])
    )



@dp.callback_query(F.data == "admin_blacklist_add")
async def admin_blacklist_add(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Hanya admin!", show_alert=True)
        return
    await state.set_state(BlacklistState.waiting_add)
    await callback.message.edit_text(
        "➕ <b>TAMBAH NO BLACKLIST (ADMIN)</b>\n\n"
        "Kirim format:\n"
        "<code>nomor|alasan</code>\n\n"
        "Contoh:\n"
        "<code>083812345678|Penipuan</code>\n"
        "<code>083812345678</code>\n\n"
        "Bisa kirim banyak sekaligus, pisah baris baru:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ KEMBALI KE PANEL ADMIN", callback_data="admin_panel")]
        ])
    )
    await callback.answer()

@dp.callback_query(F.data == "admin_blacklist_del")
async def admin_blacklist_del_menu(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Hanya admin!", show_alert=True)
        return
    conn = db()
    rows = conn.execute("SELECT * FROM blacklist_numbers ORDER BY id DESC LIMIT 30").fetchall()
    conn.close()
    
    if not rows:
        await callback.message.edit_text(
            "🗑️ <b>HAPUS NO BLACKLIST</b>\n\nBelum ada data blacklist.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ KEMBALI KE PANEL ADMIN", callback_data="admin_panel")]
            ])
        )
        await callback.answer()
        return
    
    text = "🗑️ <b>HAPUS NO BLACKLIST</b>\n\nPilih nomor yang mau dihapus:\n\n"
    kb = []
    for r in rows:
        kb.append([InlineKeyboardButton(text=f"🗑️ {r['phone']} - {r['reason'][:15]}", callback_data=f"bl_del_{r['id']}")])
    
    kb.append([InlineKeyboardButton(text="🗑️ HAPUS SEMUA BLACKLIST", callback_data="bl_del_all")])
    kb.append([InlineKeyboardButton(text="⬅️ KEMBALI KE PANEL ADMIN", callback_data="admin_panel")])
    
    await callback.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=kb))
    await callback.answer()

@dp.callback_query(F.data.startswith("bl_del_"))
async def admin_blacklist_del_process(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Hanya admin!", show_alert=True)
        return
    
    data = callback.data.replace("bl_del_", "")
    conn = db()
    if data == "all":
        conn.execute("DELETE FROM blacklist_numbers")
        conn.commit()
        conn.close()
        await callback.message.edit_text(
            "✅ <b>SEMUA BLACKLIST DIHAPUS</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔐 PANEL ADMIN", callback_data="admin_panel")]
            ])
        )
        await callback.answer("✅ Semua dihapus")
        return
    
    try:
        bl_id = int(data)
        row = conn.execute("SELECT phone FROM blacklist_numbers WHERE id=?", (bl_id,)).fetchone()
        if row:
            conn.execute("DELETE FROM blacklist_numbers WHERE id=?", (bl_id,))
            conn.commit()
            await callback.message.edit_text(
                f"✅ Nomor <code>{row['phone']}</code> berhasil dihapus dari blacklist!",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="🗑️ HAPUS LAIN", callback_data="admin_blacklist_del")],
                    [InlineKeyboardButton(text="🔐 PANEL ADMIN", callback_data="admin_panel")]
                ])
            )
            await callback.answer(f"✅ {row['phone']} dihapus")
        else:
            await callback.answer("❌ Tidak ditemukan", show_alert=True)
    except Exception as e:
        await callback.answer(f"❌ Error: {e}", show_alert=True)
    finally:
        conn.close()

# Hapus handler blacklist_add lama yang bisa diakses user


# ==================== ADMIN PANEL ====================

@dp.message(Command("admin"))
async def admin_command(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("❌ Kamu bukan Admin.")
        return
    await state.clear()
    await message.answer("🔐 <b>PANEL ADMIN</b>\n\nPilih menu:", reply_markup=admin_menu())


@dp.callback_query(F.data == "admin_cancel")
async def admin_cancel(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await callback.message.edit_text(
        "🔐 <b>PANEL ADMIN</b>\n\nPilih menu:",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.callback_query(F.data == "admin_panel")
async def admin_panel(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Kamu bukan Admin.", show_alert=True)
        return
    await callback.message.edit_text("🔐 <b>PANEL ADMIN</b>\n\nPilih menu:", reply_markup=admin_menu())
    await callback.answer()


@dp.callback_query(F.data == "admin_active")
async def admin_active(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return

    conn = db()
    rows = conn.execute("""
        SELECT u.telegram_id,u.name,u.username,s.package_name,s.expiry_date,s.status
        FROM users u JOIN subscriptions s ON s.telegram_id=u.telegram_id
        WHERE s.status='unlimited' OR (s.status='active' AND s.expiry_date>?)
        ORDER BY u.telegram_id
    """, (datetime.now().isoformat(),)).fetchall()
    conn.close()

    if not rows:
        text = "👥 <b>USER AKTIF</b>\n\nTidak ada user aktif."
    else:
        text = "👥 <b>USER AKTIF</b>\n\n" + "\n\n".join(
            f"👤 {r['name']}\n🆔 <code>{r['telegram_id']}</code>\n"
            f"📦 {r['package_name']}\n"
            f"📅 {'SELAMANYA' if r['status']=='unlimited' else r['expiry_date'][:10]}"
            for r in rows
        )
        text += f"\n\n📊 Total user aktif: <b>{len(rows)}</b>"

    await callback.message.edit_text(text[:3900], reply_markup=admin_menu())
    await callback.answer()


@dp.callback_query(F.data == "admin_pending")
async def admin_pending(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return

    conn = db()
    rows = conn.execute(
        "SELECT * FROM transactions WHERE status='pending' ORDER BY id DESC"
    ).fetchall()
    conn.close()

    if not rows:
        await callback.message.edit_text(
            "💰 <b>TRANSAKSI PENDING</b>\n\nTidak ada transaksi pending.",
            reply_markup=admin_menu()
        )
        await callback.answer()
        return

    for row in rows:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ SETUJU", callback_data=f"approve_{row['id']}"),
            InlineKeyboardButton(text="❌ TOLAK", callback_data=f"reject_{row['id']}")
        ]])
        caption = (
            "💰 <b>TRANSAKSI PENDING</b>\n\n"
            f"🧾 #{row['id']}\n🆔 <code>{row['telegram_id']}</code>\n"
            f"💰 {rupiah(row['amount'])}\n📦 {row['package_name'] or 'TOP UP SALDO'}"
        )
        if row["proof_file_id"]:
            await callback.bot.send_photo(callback.from_user.id, row["proof_file_id"],
                                          caption=caption, reply_markup=keyboard)
        else:
            await callback.bot.send_message(callback.from_user.id, caption, reply_markup=keyboard)

    await callback.message.edit_text(
        f"💰 <b>TRANSAKSI PENDING</b>\n\nDitemukan <b>{len(rows)}</b> transaksi.",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.callback_query(F.data == "admin_add_balance")
async def admin_add_balance_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_user_amount)
    await state.update_data(action="add")
    await callback.message.edit_text(
        "➕ <b>TAMBAH SALDO</b>\n\nKirim:\n<code>ID_USER NOMINAL</code>\n\nContoh: <code>123456789 50000</code>",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.callback_query(F.data == "admin_sub_balance")
async def admin_sub_balance_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_user_amount)
    await state.update_data(action="sub")
    await callback.message.edit_text(
        "➖ <b>KURANGI SALDO</b>\n\nKirim:\n<code>ID_USER NOMINAL</code>\n\nContoh: <code>123456789 50000</code>",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.message(AdminState.waiting_user_amount, F.text)
async def admin_balance_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return

    data = await state.get_data()
    parts = message.text.strip().split()
    if len(parts) != 2 or not all(p.isdigit() for p in parts):
        await message.answer("❌ Format salah. Contoh: <code>123456789 50000</code>")
        return

    user_id, amount = int(parts[0]), int(parts[1])
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE telegram_id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        await message.answer("❌ User tidak ditemukan.")
        return

    if data.get("action") == "add":
        conn.execute("UPDATE users SET balance=balance+? WHERE telegram_id=?", (amount, user_id))
        action_text = f"➕ Ditambah {rupiah(amount)}"
    else:
        conn.execute("UPDATE users SET balance=MAX(balance-?,0) WHERE telegram_id=?", (amount, user_id))
        action_text = f"➖ Dikurangi {rupiah(amount)}"

    new_balance = conn.execute(
        "SELECT balance FROM users WHERE telegram_id=?", (user_id,)
    ).fetchone()["balance"]
    conn.commit()
    conn.close()
    await state.clear()

    await message.answer(
        "✅ <b>BERHASIL</b>\n\n"
        f"👤 User : <code>{user_id}</code>\n{action_text}\n"
        f"💰 Saldo sekarang : <b>{rupiah(new_balance)}</b>",
        reply_markup=admin_menu()
    )
    try:
        await message.bot.send_message(
            user_id,
            "💰 <b>PERUBAHAN SALDO</b>\n\n"
            f"{action_text}\nSaldo sekarang : <b>{rupiah(new_balance)}</b>"
        )
    except Exception:
        pass


@dp.callback_query(F.data == "admin_delete_user")
async def admin_delete_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_delete_user)
    await callback.message.edit_text(
        "🗑️ <b>HAPUS USER</b>\n\nKirim Telegram ID user.\nContoh: <code>123456789</code>",
        reply_markup=admin_menu()
    )
    await callback.answer()


@dp.message(AdminState.waiting_delete_user, F.text)
async def admin_delete_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return
    if not message.text.strip().isdigit():
        await message.answer("❌ Telegram ID harus angka.")
        return

    user_id = int(message.text.strip())
    if user_id in ADMIN_IDS:
        await message.answer("❌ User Admin tidak boleh dihapus.")
        return

    conn = db()
    user = conn.execute("SELECT telegram_id FROM users WHERE telegram_id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        await message.answer("❌ User tidak ditemukan.")
        return

    for table in ["users", "subscriptions", "transactions", "format_settings", "format_results", "format_history"]:
        conn.execute(f"DELETE FROM {table} WHERE telegram_id=?", (user_id,))
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(
        f"🗑️ User <code>{user_id}</code> berhasil dihapus.",
        reply_markup=admin_menu()
    )


@dp.callback_query(F.data == "admin_broadcast")
async def admin_broadcast_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_broadcast)
    await callback.message.edit_text(
        "📢 <b>BROADCAST</b>\n\nKirim pesan/foto/dokumen yang ingin dikirim ke semua user.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ BATAL", callback_data="admin_cancel")]
        ])
    )
    await callback.answer()


@dp.message(AdminState.waiting_broadcast)
async def admin_broadcast_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return

    conn = db()
    users = conn.execute("SELECT telegram_id FROM users").fetchall()
    conn.close()

    success = failed = 0
    for row in users:
        try:
            if message.text:
                await message.bot.send_message(row["telegram_id"], message.text)
            elif message.photo:
                await message.bot.send_photo(row["telegram_id"], message.photo[-1].file_id, caption=message.caption or "")
            elif message.document:
                await message.bot.send_document(row["telegram_id"], message.document.file_id, caption=message.caption or "")
            else:
                failed += 1
                continue
            success += 1
            await asyncio.sleep(0.05)
        except Exception:
            failed += 1

    await state.clear()
    await message.answer(
        "📢 <b>BROADCAST SELESAI</b>\n\n"
        f"✅ Berhasil : {success}\n❌ Gagal : {failed}",
        reply_markup=admin_menu()
    )



@dp.callback_query(F.data == "admin_list_admin")
async def admin_list_admin(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    conn=db()
    rows=conn.execute("SELECT telegram_id, added_by, created_at FROM admins ORDER BY telegram_id").fetchall()
    conn.close()
    env_admins="\n".join([f"• <code>{aid}</code> (ENV)" for aid in ADMIN_IDS])
    db_text="\n".join([f"• <code>{r['telegram_id']}</code>" for r in rows]) if rows else "-"
    await callback.message.edit_text(f"👥 <b>DAFTAR ADMIN</b>\n\n<b>ENV:</b>\n{env_admins}\n\n<b>DB:</b>\n{db_text}", reply_markup=admin_menu())
    await callback.answer()
@dp.callback_query(F.data == "admin_tambah_admin")
async def admin_tambah_admin_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_add_admin)
    await callback.message.edit_text("➕ <b>TAMBAH ADMIN</b>", reply_markup=admin_menu())
    await callback.answer()
@dp.message(AdminState.waiting_add_admin, F.text)
async def admin_tambah_admin_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return
    added=[]
    for p in message.text.replace(',', ' ').split():
        if p.isdigit():
            aid=int(p)
            conn=db()
            conn.execute("INSERT INTO admins(telegram_id, added_by, created_at) VALUES(?,?,?) ON CONFLICT(telegram_id) DO NOTHING", (aid, message.from_user.id, __import__('datetime').datetime.now().isoformat()))
            conn.commit()
            conn.close()
            added.append(aid)
    await state.clear()
    await message.answer(f"✅ Tambah: {added}", reply_markup=admin_menu())
@dp.callback_query(F.data == "admin_hapus_admin")
async def admin_hapus_admin_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Bukan Admin.", show_alert=True)
        return
    await state.clear()
    await state.set_state(AdminState.waiting_remove_admin)
    await callback.message.edit_text("➖ <b>HAPUS ADMIN</b>", reply_markup=admin_menu())
    await callback.answer()
@dp.message(AdminState.waiting_remove_admin, F.text)
async def admin_hapus_admin_process(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await state.clear()
        return
    removed=[]
    for p in message.text.replace(',', ' ').split():
        if p.isdigit():
            aid=int(p)
            conn=db()
            cur=conn.execute("DELETE FROM admins WHERE telegram_id=?", (aid,))
            conn.commit()
            conn.close()
            if cur.rowcount>0:
                removed.append(aid)
    await state.clear()
    await message.answer(f"✅ Hapus: {removed}", reply_markup=admin_menu())



# ==================== MENU KOTA BARU - TAMBAH KOTA / KOTA SAYA / CARI KOTA LAINNYA ====================


@dp.callback_query(F.data == "kota_add")
async def kota_add_handler(callback: CallbackQuery, state: FSMContext):
    # Cek akses TAMBAH KOTA - wajib top up dulu
    can_add, reason = check_kota_can_add(callback.from_user.id)
    
    if not can_add:
        if reason == "belum_langganan":
            await callback.message.edit_text(
                "🔒 <b>TAMBAH KOTA TERKUNCI</b>

"
                "Untuk menambah kota, wajib TOP UP dulu bos!

"
                "💰 <b>HARGA PAKET TAMBAH KOTA:</b>
"
                "• 1 Minggu - <b>Rp 35.000</b> (bisa tambah 2 kota)
"
                "• 1 Bulan - <b>Rp 120.000</b> (bisa tambah 2 kota)
"
                "• 2 Bulan - <b>Rp 200.000</b> (bisa tambah 2 kota)
"
                "• 6 Bulan - <b>Rp 500.000</b> (bisa tambah 2 kota)
"
                "• Unlimited - <b>Rp 2.000.000</b> (tambah sepuasnya)

"
                "Setelah top up, kamu bisa tambah kota 2x.
"
                "Kalau sudah 3x, wajib top up lagi!

"
                "Silakan pilih paket:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="📅 1 MINGGU - Rp 35K", callback_data="kota_1w")],
                    [InlineKeyboardButton(text="📅 1 BULAN - Rp 120K", callback_data="kota_1m")],
                    [InlineKeyboardButton(text="📅 2 BULAN - Rp 200K", callback_data="kota_2m")],
                    [InlineKeyboardButton(text="📅 6 BULAN - Rp 500K", callback_data="kota_6m")],
                    [InlineKeyboardButton(text="♾️ UNLIMITED - Rp 2JT", callback_data="kota_unlimited")],
                    [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="back_main")]
                ])
            )
            await callback.answer("🔒 Wajib Top Up dulu bos!")
            return
        elif reason == "quota_habis":
            await callback.message.edit_text(
                "⚠️ <b>QUOTA TAMBAH KOTA HABIS</b>

"
                "Kamu sudah tambah kota 2x bos!
"
                "Untuk tambah lagi yang ke-3, wajib TOP UP lagi.

"
                "💰 <b>HARGA PAKET:</b>
"
                "• 1 Minggu - Rp 35.000
"
                "• 1 Bulan - Rp 120.000
"
                "• 2 Bulan - Rp 200.000
"
                "• 6 Bulan - Rp 500.000
"
                "• Unlimited - Rp 2.000.000

"
                "Pilih paket untuk lanjut:",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="📅 1 MINGGU - Rp 35K", callback_data="kota_1w")],
                    [InlineKeyboardButton(text="📅 1 BULAN - Rp 120K", callback_data="kota_1m")],
                    [InlineKeyboardButton(text="📅 2 BULAN - Rp 200K", callback_data="kota_2m")],
                    [InlineKeyboardButton(text="📅 6 BULAN - Rp 500K", callback_data="kota_6m")],
                    [InlineKeyboardButton(text="♾️ UNLIMITED - Rp 2JT", callback_data="kota_unlimited")],
                    [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="back_main")]
                ])
            )
            await callback.answer("⚠️ Quota habis, top up lagi!")
            return
    
    # Jika lolos, lanjut ke PILIH PROVINSI - 1 KOTAK PER BARIS
    used, quota = get_kota_usage(callback.from_user.id)
    sisa = quota - used
    await state.set_state(WilayahState.provinsi)
    
    # Build 1 kolom biar jelas tulisannya
    kb = []
    for prov in PROVINSI_LIST:
        kb.append([InlineKeyboardButton(text=f"📍 {prov}", callback_data=f"prov_{prov}")])
    kb.append([InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="back_main")])
    
    await callback.message.edit_text(
        f"📍 <b>PILIH PROVINSI</b>

"
        f"✅ Akses aktif! Sisa quota: <b>{sisa}x</b>

"
        f"Silakan pilih provinsi (1 kotak per baris biar jelas):",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("prov_"))
async def pilih_provinsi_handler(callback: CallbackQuery, state: FSMContext):
    provinsi = callback.data.replace("prov_", "", 1)
    await state.update_data(provinsi=provinsi)
    await state.set_state(WilayahState.kabupaten)
    
    kabupaten_list = get_kabupaten_list(provinsi)
    
    kb = []
    for kab in kabupaten_list:
        kb.append([InlineKeyboardButton(text=f"🏙️ {kab}", callback_data=f"kab_{kab}")])
    kb.append([InlineKeyboardButton(text="⬅️ KEMBALI KE PROVINSI", callback_data="kota_add")])
    
    await callback.message.edit_text(
        f"🏙️ <b>PILIH KABUPATEN/KOTA</b>

"
        f"📍 Provinsi: <b>{provinsi}</b>

"
        f"Pilih kabupaten/kota (1 baris biar jelas):",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer()



@dp.callback_query(F.data.startswith("kab_"))
async def pilih_kabupaten_handler(callback: CallbackQuery, state: FSMContext):
    kabupaten = callback.data.replace("kab_", "", 1)
    data = await state.get_data()
    provinsi = data.get("provinsi", "-")
    await state.update_data(kabupaten=kabupaten, kecamatan_selected=[])
    await state.set_state(WilayahState.kecamatan_multi)
    
    kecamatan_list = get_kecamatan_list(kabupaten)
    
    # 1 KOTAK PER BARIS UNTUK KECAMATAN JUGA - BIAR JELAS
    kb = []
    for kec in kecamatan_list:
        kb.append([InlineKeyboardButton(text=f"🔶 {kec}", callback_data=f"kec_toggle_{kec}")])
    
    kb.append([InlineKeyboardButton(text="💾 SIMPAN 0 KECAMATAN ✅", callback_data="kec_done")])
    kb.append([InlineKeyboardButton(text="🗑️ HAPUS PILIHAN (0)", callback_data="kec_clear")])
    kb.append([InlineKeyboardButton(text="⬅️ Kembali", callback_data=f"prov_{provinsi}")])
    
    await callback.message.edit_text(
        f"📍 <b>PILIH KECAMATAN</b>

"
        f"📍 Provinsi: <b>{provinsi}</b>
"
        f"🏙️ Kab/Kota: <b>{kabupaten}</b>

"
        f"Dipilih: 0 kecamatan",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("kec_toggle_"))
async def toggle_kecamatan_handler(callback: CallbackQuery, state: FSMContext):
    kecamatan = callback.data.replace("kec_toggle_", "", 1)
    data = await state.get_data()
    selected = data.get("kecamatan_selected", [])
    provinsi = data.get("provinsi", "-")
    kabupaten = data.get("kabupaten", "-")
    
    if kecamatan in selected:
        selected.remove(kecamatan)
    else:
        selected.append(kecamatan)
    
    await state.update_data(kecamatan_selected=selected)
    
    kecamatan_list = get_kecamatan_list(kabupaten)
    
    kb = []
    for kec in kecamatan_list:
        if kec in selected:
            icon = "✅"
        else:
            icon = "🔶"
        kb.append([InlineKeyboardButton(text=f"{icon} {kec}", callback_data=f"kec_toggle_{kec}")])
    
    count = len(selected)
    kb.append([InlineKeyboardButton(text=f"💾 SIMPAN {count} KECAMATAN ✅", callback_data="kec_done")])
    kb.append([InlineKeyboardButton(text=f"🗑️ HAPUS PILIHAN ({count})", callback_data="kec_clear")])
    kb.append([InlineKeyboardButton(text="⬅️ Kembali", callback_data=f"prov_{provinsi}")])
    
    await callback.message.edit_text(
        f"Dipilih: {count} kecamatan",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer(f"{'✅' if kecamatan in selected else '🔶'} {kecamatan}")

@dp.callback_query(F.data == "kec_clear")
async def clear_kecamatan_handler(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    provinsi = data.get("provinsi", "-")
    kabupaten = data.get("kabupaten", "-")
    await state.update_data(kecamatan_selected=[])
    
    kecamatan_list = get_kecamatan_list(kabupaten)
    
    kb = []
    for kec in kecamatan_list:
        kb.append([InlineKeyboardButton(text=f"🔶 {kec}", callback_data=f"kec_toggle_{kec}")])
    
    kb.append([InlineKeyboardButton(text="💾 SIMPAN 0 KECAMATAN ✅", callback_data="kec_done")])
    kb.append([InlineKeyboardButton(text="🗑️ HAPUS PILIHAN (0)", callback_data="kec_clear")])
    kb.append([InlineKeyboardButton(text="⬅️ Kembali", callback_data=f"prov_{provinsi}")])
    
    await callback.message.edit_text(
        f"Dipilih: 0 kecamatan",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer("🗑️ Pilihan dihapus")



        return
    
    # Simpan ke database
    conn = db()
    for kec in selected:
        conn.execute(
            "INSERT INTO user_kota(telegram_id,kota_name,kab,kec,kel,catatan,created_at) VALUES(?,?,?,?,?,?,?)",
            (callback.from_user.id, f"{kabupaten} - {kec}", provinsi, kabupaten, kec, "", __import__('datetime').datetime.now().isoformat())
        )
    conn.commit()
    conn.close()
    
    await state.clear()
    
    kec_list_str = "
".join([f"• {k}" for k in selected])
    
    await callback.message.edit_text(
        f"✅ <b>KOTA BERHASIL DISIMPAN</b>

"
        f"📍 Provinsi: <b>{provinsi}</b>
"
        f"🏙️ Kab/Kota: <b>{kabupaten}</b>
"
        f"📋 Kecamatan ({len(selected)}):
{kec_list_str}

"
        f"Tersimpan di KOTA SAYA!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🌆 LIHAT KOTA SAYA", callback_data="kota_list")],
            [InlineKeyboardButton(text="➕ TAMBAH LAGI", callback_data="kota_add")],
            [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer(f"✅ {len(selected)} kecamatan disimpan!")



        return
    
    # Simpan ke database
    conn = db()
    for kec in selected:
        conn.execute(
            "INSERT INTO user_kota(telegram_id,kota_name,kab,kec,kel,catatan,created_at) VALUES(?,?,?,?,?,?,?)",
            (callback.from_user.id, f"{kabupaten} - {kec}", provinsi, kabupaten, kec, "", __import__('datetime').datetime.now().isoformat())
        )
    conn.commit()
    conn.close()
    
    await state.clear()
    
    kec_list_str = "
".join([f"• {k}" for k in selected])
    
    await callback.message.edit_text(
        f"✅ <b>KOTA BERHASIL DITAMBAHKAN</b>

"
        f"📍 Provinsi: <b>{provinsi}</b>
"
        f"🏙️ Kab/Kota: <b>{kabupaten}</b>
"
        f"📋 Kecamatan ({len(selected)}):
{kec_list_str}

"
        f"Tersimpan di KOTA SAYA!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🌆 LIHAT KOTA SAYA", callback_data="kota_list")],
            [InlineKeyboardButton(text="➕ TAMBAH LAGI", callback_data="kota_add")],
            [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
        ])
    )
    await callback.answer(f"✅ {len(selected)} kecamatan ditambahkan!")


@dp.callback_query(F.data.in_(set(KOTA_PACKAGES.keys())))
async def kota_package_handler(callback: CallbackQuery, state: FSMContext):
    code, name, price, days = KOTA_PACKAGES[callback.data]
    await state.update_data(amount=price, package_code=code, package_name=name, kota_days=days)
    await state.set_state(PaymentState.waiting_proof)
    await callback.message.edit_text(
        f"💳 <b>PEMBAYARAN {name}</b>

"
        f"📦 Paket: <b>{name}</b>
"
        f"💰 Harga: <b>{rupiah(price)}</b>
"
        f"⏰ Durasi: {days} hari (Unlimited jika 2JT)
"
        f"🎯 Quota: Bisa tambah kota <b>2x</b> (Unlimited jika paket Unlimited)

"
        f"Silakan transfer ke:

"
        f"🏦 <b>SEABANK</b>
901040978290
A/N HAMBALI

"
        f"💰 <b>DANA</b>
083824101264
A/N HAMBALI

"
        f"Setelah transfer, tekan SUDAH BAYAR lalu upload bukti.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ SUDAH BAYAR", callback_data="payment_done")],
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="kota_add")]
        ])
    )
    await callback.answer()


@dp.message(KotaState.waiting_kota_add, F.text)
async def kota_add_process(message: Message, state: FSMContext):
    parts = [p.strip() for p in message.text.split("|")]
    if not parts[0]:
        await message.answer("❌ Nama kota tidak boleh kosong.")
        return
    kota_name = parts[0]
    kab = parts[1] if len(parts) > 1 else ""
    kec = parts[2] if len(parts) > 2 else ""
    kel = parts[3] if len(parts) > 3 else ""
    catatan = parts[4] if len(parts) > 4 else ""
    conn = db()
    conn.execute(
        "INSERT INTO user_kota(telegram_id,kota_name,kab,kec,kel,catatan,created_at) VALUES(?,?,?,?,?,?,?)",
        (message.from_user.id, kota_name, kab, kec, kel, catatan, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(
        f"✅ <b>KOTA DITAMBAHKAN</b>\n\n🏙️ Kota: <b>{html.escape(kota_name)}</b>\n📍 KAB: {html.escape(kab) or '-'} | KEC: {html.escape(kec) or '-'} | KEL: {html.escape(kel) or '-'}",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🌆 LIHAT KOTA SAYA", callback_data="kota_list")],
            [InlineKeyboardButton(text="➕ TAMBAH LAGI", callback_data="kota_add")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )



@dp.callback_query(F.data == "kota_list")
async def kota_list_handler(callback: CallbackQuery):
    conn = db()
    rows = conn.execute("SELECT * FROM user_kota WHERE telegram_id=? ORDER BY id DESC", (callback.from_user.id,)).fetchall()
    
    # Get subscriptions for KOTA and CARI
    kota_sub = get_kota_subscription(callback.from_user.id)
    cari_sub = get_cari_subscription(callback.from_user.id)
    
    text = "🌆 <b>KOTA SAYA</b>

"
    
    if not rows:
        text += "Belum ada kota yang dipilih.
Silakan tambah kota dulu di menu TAMBAH KOTA.

"
    else:
        for i, r in enumerate(rows, 1):
            kota_name = r["kota_name"] or "-"
            kab = r["kab"] or "-"
            kec = r["kec"] or "-"
            kel = r["kel"] or "-"
            created = r["created_at"][:10] if r["created_at"] else "-"
            text += f"<b>{i}. {kota_name}</b>
   📍 {kab} - {kec} - {kel}
   📅 Ditambah: {created}

"
    
    text += "━━━━━━━━━━━━━━━━
"
    text += "📊 <b>STATUS LANGGANAN:</b>

"
    
    # KOTA status
    if kota_sub:
        if kota_sub["status"] == "unlimited":
            text += f"🏙️ TAMBAH KOTA: <b>{kota_sub['package_name']}</b>
"
            text += f"   ⏰ Expired: <b>SELAMANYA</b>
"
            text += f"   📊 Status: <b>🟢 ACTIVE</b>

"
        else:
            try:
                from datetime import datetime as dt
                exp = dt.fromisoformat(kota_sub["expiry_date"])
                exp_str = exp.strftime("%d-%m-%Y %H:%M")
                is_active = dt.now() < exp
                status_emoji = "🟢 ACTIVE" if is_active else "🔴 EXPIRED"
                text += f"🏙️ TAMBAH KOTA: <b>{kota_sub['package_name']}</b>
"
                text += f"   ⏰ Expired: <b>{exp_str}</b>
"
                text += f"   📊 Status: <b>{status_emoji}</b>
"
                used, quota = get_kota_usage(callback.from_user.id)
                text += f"   🎯 Terpakai: {used}/{quota}

"
            except:
                text += f"🏙️ TAMBAH KOTA: <b>{kota_sub['package_name']}</b> - ACTIVE

"
    else:
        text += "🏙️ TAMBAH KOTA: <b>❌ Belum langganan</b>

"
    
    # CARI status
    if cari_sub:
        try:
            from datetime import datetime as dt
            exp = dt.fromisoformat(cari_sub["expiry_date"])
            exp_str = exp.strftime("%d-%m-%Y %H:%M")
            is_active = dt.now() < exp
            status_emoji = "🟢 ACTIVE" if is_active else "🔴 EXPIRED"
            text += f"🔍 CARI KOTA LAIN: <b>{cari_sub['package_name']}</b>
"
            text += f"   ⏰ Expired: <b>{exp_str}</b>
"
            text += f"   📊 Status: <b>{status_emoji}</b>

"
        except:
            text += f"🔍 CARI KOTA LAIN: <b>{cari_sub['package_name']}</b> - ACTIVE

"
    else:
        text += "🔍 CARI KOTA LAIN: <b>❌ Belum langganan</b>

"
    
    conn.close()
    
    kb = []
    kb.append([InlineKeyboardButton(text="🗑️ HAPUS KOTA SAYA", callback_data="hapus_kota_saya")])
    kb.append([InlineKeyboardButton(text="🏙️ TAMBAH KOTA", callback_data="kota_add")])
    kb.append([InlineKeyboardButton(text="🔍 CARI KOTA LAIN", callback_data="kota_search_lain")])
    kb.append([InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")])
    
    await callback.message.edit_text(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await callback.answer()

@dp.callback_query(F.data == "hapus_kota_saya")
async def hapus_kota_saya_menu(callback: CallbackQuery):
    conn = db()
    rows = conn.execute("SELECT * FROM user_kota WHERE telegram_id=? ORDER BY id DESC", (callback.from_user.id,)).fetchall()
    conn.close()
    
    if not rows:
        await callback.message.edit_text(
            "🗑️ <b>HAPUS KOTA SAYA</b>

Belum ada kota untuk dihapus.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ KEMBALI KE KOTA SAYA", callback_data="kota_list")]
            ])
        )
        await callback.answer()
        return
    
    text = "🗑️ <b>HAPUS KOTA SAYA</b>

Pilih kota yang mau dihapus:

"
    kb = []
    for i, r in enumerate(rows, 1):
        kota_name = r["kota_name"] or "-"
        kab = r["kab"] or "-"
        text += f"{i}. <b>{kota_name}</b> - {kab}
"
        kb.append([InlineKeyboardButton(text=f"🗑️ HAPUS {kota_name[:20]}", callback_data=f"kota_del_{r['id']}")])
    
    kb.append([InlineKeyboardButton(text="🗑️ HAPUS SEMUA KOTA", callback_data="kota_del_all")])
    kb.append([InlineKeyboardButton(text="⬅️ KEMBALI KE KOTA SAYA", callback_data="kota_list")])
    
    await callback.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=kb))
    await callback.answer()

@dp.callback_query(F.data.startswith("kota_del_"))
async def kota_del_handler(callback: CallbackQuery):
    data = callback.data.replace("kota_del_", "")
    if data == "all":
        conn = db()
        conn.execute("DELETE FROM user_kota WHERE telegram_id=?", (callback.from_user.id,))
        conn.commit()
        conn.close()
        await callback.message.edit_text(
            "✅ <b>SEMUA KOTA DIHAPUS</b>

Semua kota saya berhasil dihapus.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🌆 KOTA SAYA", callback_data="kota_list")],
                [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
            ])
        )
        await callback.answer("✅ Semua kota dihapus")
        return
    
    try:
        kota_id = int(data)
    except:
        await callback.answer("❌ ID tidak valid", show_alert=True)
        return
    
    conn = db()
    row = conn.execute("SELECT kota_name FROM user_kota WHERE id=? AND telegram_id=?", (kota_id, callback.from_user.id)).fetchone()
    if row:
        conn.execute("DELETE FROM user_kota WHERE id=? AND telegram_id=?", (kota_id, callback.from_user.id))
        conn.commit()
        conn.close()
        await callback.message.edit_text(
            f"✅ <b>KOTA DIHAPUS</b>

Kota <b>{row['kota_name']}</b> berhasil dihapus.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🌆 LIHAT KOTA SAYA", callback_data="kota_list")],
                [InlineKeyboardButton(text="🗑️ HAPUS KOTA LAIN", callback_data="hapus_kota_saya")],
                [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
            ])
        )
        await callback.answer(f"✅ {row['kota_name']} dihapus")
    else:
        conn.close()
        await callback.answer("❌ Kota tidak ditemukan", show_alert=True)


        return
    text = f"🌆 <b>KOTA SAYA - {len(rows)} DATA</b>\n\n"
    buttons = []
    for r in rows:
        text += f"• <b>{html.escape(r['kota_name'])}</b> | {html.escape(r['kab']) or '-'} | {html.escape(r['kec']) or '-'} | {html.escape(r['kel']) or '-'}\n"
        buttons.append([InlineKeyboardButton(text=f"🗑️ HAPUS {r['kota_name'][:12]}", callback_data=f"kota_del_{r['id']}")])
    buttons.append([InlineKeyboardButton(text="🏙️ TAMBAH KOTA", callback_data="kota_add")])
    buttons.append([InlineKeyboardButton(text="🔍 CARI KOTA LAINNYA", callback_data="kota_search_lain")])
    buttons.append([InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="auto_format")])
    await callback.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@dp.callback_query(F.data.startswith("kota_del_"))
async def kota_delete(callback: CallbackQuery):
    try:
        kota_id = int(callback.data.split("_")[-1])
    except:
        await callback.answer("❌ ID salah")
        return
    conn = db()
    conn.execute("DELETE FROM user_kota WHERE id=? AND telegram_id=?", (kota_id, callback.from_user.id))
    conn.commit()
    conn.close()
    await callback.answer("🗑️ Dihapus")
    await kota_list(callback)


@dp.callback_query(F.data == "kota_search_lain")
async def kota_search_lain_handler(callback: CallbackQuery, state: FSMContext):
    if not has_cari_access(callback.from_user.id):
        await callback.message.edit_text(
            "🔒 <b>CARI KOTA LAIN TERKUNCI</b>

"
            "Untuk cari kota lain, wajib TOP UP dulu bos!

"
            "💰 <b>HARGA PAKET CARI KOTA LAIN:</b>
"
            "• 1 Minggu - <b>Rp 15.000</b>
"
            "• 1 Bulan - <b>Rp 50.000</b>
"
            "• 2 Bulan - <b>Rp 80.000</b>

"
            "Silakan pilih paket:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📅 1 MINGGU - Rp 15K", callback_data="cari_1w")],
                [InlineKeyboardButton(text="📅 1 BULAN - Rp 50K", callback_data="cari_1m")],
                [InlineKeyboardButton(text="📅 2 BULAN - Rp 80K", callback_data="cari_2m")],
                [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="back_main")]
            ])
        )
        await callback.answer("🔒 Wajib Top Up dulu bos!")
        return
    
    await state.set_state(KotaState.waiting_lain_search)
    await callback.message.edit_text(
        "🔎 <b>CARI DATA LAINNYA</b>

"
        "📍 MASUKAN NAMA KOTA
"
        "💡 Contoh: BANDUNG

"
        "✍️ Ketik kota yang mau dicari
"
        "Bot akan cari di history WA yang dishare pengirim!

"
        "❌ Ketik /batal untuk batal.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ BATAL", callback_data="batal_cari")],
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="back_main")]
        ])
    )
    await callback.answer()



@dp.message(KotaState.waiting_lain_search, F.text)
async def kota_search_lain_process(message: Message, state: FSMContext):
    query = message.text.strip()
    if len(query) < 3:
        await message.answer("❌ Ketik minimal 3 huruf bos. Contoh: SERANG CIPOCOK JAYA")
        return
    
    parts = query.upper().split()
    if len(parts) < 2:
        await message.answer(
            "❌ Harus KOTA dan KECAMATAN bos!

"
            "Contoh: <code>SERANG CIPOCOK JAYA</code>
"
            "Bukan cuma <code>SERANG</code> saja!

"
            "Ketik ulang dengan KOTA dan KECAMATAN:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ BATAL", callback_data="back_main")]
            ])
        )
        return
    
    kota_keyword = parts[0]
    kec_keyword = " ".join(parts[1:])
    
    conn = db()
    all_results = conn.execute("SELECT * FROM format_results ORDER BY id DESC LIMIT 200").fetchall()
    
    matched = []
    for r in all_results:
        txt_upper = ((r["result_text"] or "") + " " + (r["input_text"] or "")).upper()
        if kota_keyword in txt_upper and kec_keyword in txt_upper:
            matched.append(r)
    
    kota_refs = conn.execute(
        "SELECT * FROM user_kota WHERE (kota_name LIKE ? OR kab LIKE ? OR kec LIKE ?) AND (kota_name LIKE ? OR kab LIKE ? OR kec LIKE ?) LIMIT 50",
        (f"%{kota_keyword}%", f"%{kota_keyword}%", f"%{kota_keyword}%", f"%{kec_keyword}%", f"%{kec_keyword}%", f"%{kec_keyword}%")
    ).fetchall()
    
    conn.close()
    
    if not matched and not kota_refs:
        await message.answer(
            f"🔍 <b>HASIL CARI: {query}</b>

"
            f"❌ Tidak ada data yang mengandung
"
            f"KOTA: <b>{kota_keyword}</b> dan KECAMATAN: <b>{kec_keyword}</b>

"
            f"Coba kata kunci lain atau cek ejaan.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔍 CARI LAGI", callback_data="kota_search_lain")],
                [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
            ])
        )
        await state.clear()
        return
    
    text = f"🔍 <b>HASIL CARI KOTA LAIN</b>

"
    text += f"Kata kunci: <b>{query}</b>
"
    text += f"KOTA: <b>{kota_keyword}</b> | KECAMATAN: <b>{kec_keyword}</b>
"
    text += f"Ditemukan: <b>{len(matched)} data</b>

"
    text += "━━━━━━━━━━━━━━━━

"
    
    for i, r in enumerate(matched[:10], 1):
        result_preview = (r["result_text"] or "")[:200].replace("
", " ")
        text += f"<b>{i}. {result_preview[:100]}...</b>
"
        phone = extract_phone_from_any(r["result_text"] or "")
        if phone:
            text += f"   📱 {phone}
"
        text += "
"
    
    if kota_refs:
        text += f"📍 <b>REFERENSI KOTA:</b>
"
        for ref in kota_refs[:5]:
            text += f"• {ref['kota_name']} - {ref['kec']} - {ref['kab']}
"
        text += "
"
    
    text += f"Total {len(matched)} kata kunci mengandung {kota_keyword} dan {kec_keyword}"
    
    await state.clear()
    await message.answer(
        text[:3900],
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔍 CARI LAGI", callback_data="kota_search_lain")],
            [InlineKeyboardButton(text="📋 LIHAT SEMUA", callback_data="format_results")],
            [InlineKeyboardButton(text="⬅️ MENU UTAMA", callback_data="back_main")]
        ])
    )


@dp.callback_query(F.data.in_(set(CARI_KOTA_PACKAGES.keys())))
async def cari_package_handler(callback: CallbackQuery, state: FSMContext):
    code, name, price, days = CARI_KOTA_PACKAGES[callback.data]
    await state.update_data(amount=price, package_code=code, package_name=name, cari_days=days)
    await state.set_state(PaymentState.waiting_proof)
    await callback.message.edit_text(
        f"💳 <b>PEMBAYARAN {name}</b>

"
        f"📦 Paket: <b>{name}</b>
"
        f"💰 Harga: <b>{rupiah(price)}</b>
"
        f"⏰ Durasi: {days} hari

"
        f"Silakan transfer ke:

"
        f"🏦 <b>SEABANK</b>
901040978290
A/N HAMBALI

"
        f"💰 <b>DANA</b>
083824101264
A/N HAMBALI

"
        f"Setelah transfer, tekan SUDAH BAYAR lalu upload bukti.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ SUDAH BAYAR", callback_data="payment_done")],
            [InlineKeyboardButton(text="⬅️ KEMBALI", callback_data="kota_search_lain")]
        ])
    )
    await callback.answer()


@dp.message(KotaState.waiting_lain_search, F.text)
async def kota_search_lain_process(message: Message, state: FSMContext):
    kw = message.text.strip()
    like = f"%{kw}%"
    conn = db()
    kota_rows = conn.execute("SELECT * FROM user_kota WHERE telegram_id=? AND (kota_name LIKE ? OR kab LIKE ? OR kec LIKE ? OR kel LIKE ? OR catatan LIKE ?) LIMIT 20", (message.from_user.id, like, like, like, like, like)).fetchall()
    format_rows = conn.execute("SELECT * FROM format_results WHERE telegram_id=? AND result_text LIKE ? ORDER BY created_at DESC LIMIT 10", (message.from_user.id, like)).fetchall()
    conn.close()
    await state.clear()
    text = f"🔍 <b>HASIL CARI: {html.escape(kw)}</b>\n\n"
    if kota_rows:
        text += f"🏙️ <b>KOTA SAYA ({len(kota_rows)})</b>:\n"
        for r in kota_rows:
            text += f"• {html.escape(r['kota_name'])} | {html.escape(r['kab'])} | {html.escape(r['kec'])} | {html.escape(r['kel'])}\n"
        text += "\n"
    if format_rows:
        text += f"📄 <b>FORMAT ({len(format_rows)})</b>:\n"
        for r in format_rows[:5]:
            snippet = (r['result_text'][:70] + "...") if len(r['result_text'])>70 else r['result_text']
            text += f"• {html.escape(snippet)}\n"
        text += "\n"
    if not kota_rows and not format_rows:
        text += "❌ Tidak ada data yang cocok."
    await message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔍 CARI LAGI", callback_data="kota_search_lain")],
            [InlineKeyboardButton(text="🌆 KOTA SAYA", callback_data="kota_list")],
            [InlineKeyboardButton(text="⬅️ MENU AUTO FORMAT", callback_data="auto_format")]
        ])
    )


@dp.message(F.text)
async def fallback(message: Message):
    # Masalah JMO tetap dijawab walaupun user mengetik tanpa membuka menu
    words = ["jmo", "error", "kode", "025", "026", "027", "028", "029", "030", "031", "032", "033", 
             "login", "verifikasi", "kpj", "jht", "otp", "wajah", "password", "email", "aktivasi", 
             "klaim", "bpu", "kamera", "server", "lemot", "notifikasi", "daftar", "cair"]
    if any(w in message.text.lower() for w in words):
        await message.answer(get_jmo_solution(message.text))
    else:
        await message.answer("Silakan gunakan menu utama dengan /start", reply_markup=main_menu())



# ==================== WEBHOOK WA GROUP -> TELEGRAM 1 DETIK ====================
async def wablas_webhook_handler(request):
    try:
        data = await request.json()
        print(f"[WABLAS WEBHOOK] {data}")
        message = data.get('message') or data.get('content') or data.get('text') or data.get('pushMessage') or ''
        sender = data.get('pushName') or data.get('sender') or data.get('fromName') or 'Unknown'
        group_name = data.get('groupName') or data.get('group') or data.get('chatName') or 'Grup WA'
        phone = data.get('phone') or data.get('fromNumber') or data.get('senderPhone') or ''
        is_group = data.get('isGroup') or data.get('isGroupMessage') or False
        
        if not message:
            # coba ambil dari message object
            if isinstance(data.get('message'), dict):
                message = data['message'].get('conversation') or data['message'].get('extendedTextMessage',{}).get('text') or ''
        
        if not message:
            return web.Response(text="no message")
        
        conn = db()
        users = conn.execute("SELECT telegram_id FROM users").fetchall()
        conn.close()
        
        fwd = (
            f"🛡️GRUP: {group_name}\n"
            f"🧑‍💻PENGIRIM: {sender}\n"
            f"📱NO HP: {phone}\n\n"
            f"{message}\n\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"⚠️ PERHATIAN :\n"
            f"Tetap hati-hati dan waspada dalam bertransaksi untuk lebih aman gunakan jasa rekber.Terimakasih🙏\n"
            f"Sumber: https//t.me/seduluranjht_bot"
        )
        
        bot_inst = request.app['bot']
        for u in users:
            try:
                await bot_inst.send_message(u['telegram_id'], fwd)
            except Exception as e:
                print(f"fail send to {u['telegram_id']}: {e}")
                pass
        
        return web.json_response({"status": "ok", "message": "forwarded"})
    except Exception as e:
        print(f"Webhook error: {e}")
        import traceback; traceback.print_exc()
        return web.Response(text=str(e), status=500)

async def health_handler(request):
    return web.Response(text="Bot Auto Format + Wablas Webhook Active! REDMI BALI")

async def start_webhook_server(bot_instance):
    app = web.Application()
    app['bot'] = bot_instance
    app.router.add_post('/webhook', wablas_webhook_handler)
    app.router.add_post('/webhook/wablas', wablas_webhook_handler)
    app.router.add_get('/', health_handler)
    app.router.add_get('/webhook', health_handler)
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.getenv("PORT", "8080"))
    site = web.TCPSite(runner, '0.0.0.0', port)
    await site.start()
    print(f"🌐 Webhook server running on port {port}")



async def main():
    init_database()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
    logging.info("🤖 JMO LINTAS TERKINI + WEBHOOK WA->TG jalan!")

    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML)
    )
    
    # Start webhook server untuk Wablas
    asyncio.create_task(start_webhook_server(bot))
    
    try:
        await dp.start_polling(bot)
    finally:
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
